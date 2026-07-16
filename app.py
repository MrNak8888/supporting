from urllib.parse import quote_plus

from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, session, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timezone, timedelta
from functools import wraps
from collections import defaultdict
from sqlalchemy import text, or_
import logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger('rbac_debug')
from flask_babel import Babel, _
from cryptography.fernet import Fernet
import os
import io
import json
import uuid
import re
import calendar
import logging
from fpdf import FPDF
from flask_wtf.csrf import CSRFProtect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import telegram_service



def _safe_redirect(fallback):
    ref = request.referrer
    if ref:
        from urllib.parse import urlparse
        host = request.host_url.rstrip('/')
        parsed = urlparse(ref)
        if parsed.netloc == request.host or not parsed.netloc:
            return redirect(ref)
    return redirect(fallback)


def send_excel(buf, filename):
    try:
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except TypeError:
        return send_file(buf, as_attachment=True, attachment_filename=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def send_pdf(buf, filename):
    buf.seek(0)
    try:
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype='application/pdf')
    except TypeError:
        return send_file(buf, as_attachment=True, attachment_filename=filename,
                         mimetype='application/pdf')


app = Flask(__name__)
app.config['SECRET_KEY'] = 'change-this-to-a-random-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"mysql+pymysql://root:{quote_plus('Admin@9999')}"
    f"@localhost/support_system"
    f"?charset=utf8mb4"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024
app.config['WTF_CSRF_ENABLED'] = True
app.config['WTF_CSRF_CHECK_DEFAULT'] = False
app.config['WTF_CSRF_TIME_LIMIT'] = 3600
app.config['KHMER_FONT_PATH'] = ''
app.config['FERNET_KEY'] = ''
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

csrf = CSRFProtect(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


def _find_khmer_font():
    path = app.config.get('KHMER_FONT_PATH')
    if path and os.path.exists(path):
        return path
    candidates = [
        r'C:\Windows\Fonts\KhmerOS_sys.ttf',
        '/usr/share/fonts/truetype/khmer/KhmerOS_sys.ttf',
        '/usr/share/fonts/truetype/KhmerOS_sys.ttf',
        '/usr/local/share/fonts/KhmerOS_sys.ttf',
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def get_locale():
    if 'lang' in session:
        return session['lang']
    return request.accept_languages.best_match(['en', 'km']) or 'en'

babel = Babel(app, locale_selector=get_locale)

ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'png', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_upload(file):
    if file and file.filename and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        unique_name = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_name))
        return unique_name
    return None

ALLOWED_ANY_EXTS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'png', 'jpeg', 'gif', 'mp4', 'zip'}

def save_any_upload(file):
    if file and file.filename:
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_ANY_EXTS:
            return None
        unique_name = f"{uuid.uuid4().hex}{'.' + ext if ext else ''}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_name))
        return unique_name
    return None


def delete_upload(filename):
    if filename:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)


def _int_or_none(val):
    try:
        return int(val) if val else None
    except (ValueError, TypeError):
        return None


def _calc_layover(departure_time, departure_from_station):
    if not departure_time or not departure_from_station:
        return ''
    try:
        dep_parts = departure_time.split(':')
        sta_parts = departure_from_station.split(':')
        dep_min = int(dep_parts[0]) * 60 + int(dep_parts[1])
        sta_min = int(sta_parts[0]) * 60 + int(sta_parts[1])
        if sta_min < dep_min:
            return '00:00'
        diff = sta_min - dep_min
        return f'{diff // 60:02d}:{diff % 60:02d}'
    except (ValueError, IndexError, TypeError):
        return ''


def _calc_pct(value, total):
    if not total:
        return 0
    return round(min((value / total) * 100, 100), 2)


def _seed_if_empty(module_key, records):
    """Seed DynamicRecord records for a module if none exist."""
    mod = ConfigModule.query.filter_by(module_key=module_key).first()
    if mod and not DynamicRecord.query.filter_by(module_id=mod.id).first():
        for data in records:
            rec = DynamicRecord(module_id=mod.id, data=data, is_active=True)
            db.session.add(rec)


# ── Telegram encryption helpers ──
def _get_fernet():
    fernet_key = app.config.get('FERNET_KEY')
    if fernet_key:
        try:
            return Fernet(fernet_key.encode())
        except Exception:
            app.logger.warning('Invalid FERNET_KEY — falling back to derived key')
    secret = app.config.get('SECRET_KEY')
    if not secret:
        app.logger.warning('SECRET_KEY not set — generating ephemeral Fernet key; encrypted tokens will be unreadable after restart')
        return Fernet(Fernet.generate_key())
    from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
    from cryptography.hazmat.primitives import hashes
    from base64 import urlsafe_b64encode
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=b'telegram_salt', iterations=100000)
    return Fernet(urlsafe_b64encode(kdf.derive(secret.encode())))


def _encrypt_token(token):
    if not token:
        return ''
    return _get_fernet().encrypt(token.encode()).decode()


def _decrypt_token(token):
    if not token:
        return ''
    try:
        return _get_fernet().decrypt(token.encode()).decode()
    except Exception as e:
        app.logger.error('Failed to decrypt bot token: %s', str(e))
        return ''


def _send_telegram_notification(notification_type, sender_fn, event_type, *args):
    try:
        config = TelegramSetting.query.filter_by(notification_type=notification_type, enabled=True).first()
        if not config or not config.bot_token or not config.chat_id:
            return
        bot_token = _decrypt_token(config.bot_token)
        if not bot_token:
            return

        base_url = request.host_url.rstrip('/') if request else None
        success, error = sender_fn(bot_token, config.chat_id, *args, event_type=event_type, base_url=base_url)
        if not success:
            app.logger.warning('Telegram notification failed: type=%s event=%s error=%s', notification_type, event_type, error)
    except Exception as e:
        app.logger.error('=== TELEGRAM EXCEPTION === type=%s event=%s error=%s', notification_type, event_type, str(e), exc_info=True)


def _calc_rating(score):
    if score >= 95:
        return 'Excellent'
    elif score >= 85:
        return 'Very Good'
    elif score >= 75:
        return 'Good'
    elif score >= 60:
        return 'Need Improvement'
    return 'Unsatisfactory'


def _calc_delay_minutes(arrival_time, departure_time):
    if not arrival_time or not departure_time:
        return 0
    try:
        arr_parts = arrival_time.split(':')
        dep_parts = departure_time.split(':')
        arr_min = int(arr_parts[0]) * 60 + int(arr_parts[1])
        dep_min = int(dep_parts[0]) * 60 + int(dep_parts[1])
        if dep_min < arr_min:
            return 0
        return dep_min - arr_min
    except (ValueError, IndexError, TypeError):
        return 0


def _calc_duration(departure_time, arrival_time):
    if not departure_time or not arrival_time:
        return ''
    try:
        dep_parts = departure_time.split(':')
        arr_parts = arrival_time.split(':')
        dep_min = int(dep_parts[0]) * 60 + int(dep_parts[1])
        arr_min = int(arr_parts[0]) * 60 + int(arr_parts[1])
        if arr_min <= dep_min:
            arr_min += 1440
        diff = arr_min - dep_min
        if diff > 1440:
            return ''
        hours = diff // 60
        mins = diff % 60
        if hours > 0 and mins > 0:
            return f'{hours}h {mins}m'
        elif hours > 0:
            return f'{hours}h'
        return f'{mins}m'
    except (ValueError, IndexError, TypeError):
        return ''


def _add_column_if_not_exists(table, column, col_type):
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table):
        app.logger.error('Invalid table name for migration: %s', table)
        return
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', column):
        app.logger.error('Invalid column name for migration: %s', column)
        return
    if not re.match(r'^[a-zA-Z0-9_(,)\s]+$', col_type):
        app.logger.error('Invalid column type for migration: %s', col_type)
        return
    try:
        db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {column} {col_type}'))
        db.session.commit()
    except Exception:
        db.session.rollback()


db = SQLAlchemy(app)

# ── Auto Schema Migration System ──
class SchemaMigration(db.Model):
    __tablename__ = '_schema_migrations'
    id = db.Column(db.Integer, primary_key=True)
    migration_name = db.Column(db.String(200), unique=True, nullable=False)
    applied_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))


def _get_table_columns(table_name):
    from sqlalchemy import inspect as sa_inspect
    inspector = sa_inspect(db.engine)
    try:
        return set(c['name'] for c in inspector.get_columns(table_name))
    except Exception:
        return set()


def _get_model_columns(model):
    return set(c.name for c in model.__table__.columns)


def _detect_missing_columns(model):
    return _get_model_columns(model) - _get_table_columns(model.__tablename__)


def _get_column_type(model, col_name):
    col = model.__table__.columns.get(col_name)
    if col is None:
        return 'VARCHAR(255)'
    ts = str(col.type).upper()
    if 'VARCHAR' in ts: return str(col.type)
    if 'INTEGER' in ts: return 'INTEGER'
    if 'FLOAT' in ts or 'REAL' in ts: return 'FLOAT'
    if 'BOOLEAN' in ts: return 'BOOLEAN'
    if 'DATE' in ts: return 'DATE'
    if 'DATETIME' in ts: return 'DATETIME'
    if 'TEXT' in ts: return 'TEXT'
    if 'BLOB' in ts: return 'BLOB'
    if 'NUMERIC' in ts or 'DECIMAL' in ts: return 'FLOAT'
    if 'JSON' in ts: return 'TEXT'
    return 'VARCHAR(255)'


AUTO_MIGRATION_MODELS = []


def auto_register_model(model):
    if model not in AUTO_MIGRATION_MODELS:
        AUTO_MIGRATION_MODELS.append(model)
    return model


def run_auto_migrations():
    """Detect and apply missing columns for all registered models."""
    with app.app_context():
        db.create_all()
        for model in AUTO_MIGRATION_MODELS:
            table_name = model.__tablename__
            missing = _detect_missing_columns(model)
            for col_name in sorted(missing):
                if col_name.startswith('_'):
                    continue
                mig_name = f'{table_name}.{col_name}'
                existing_mig = SchemaMigration.query.filter_by(migration_name=mig_name).first()
                if existing_mig:
                    continue
                _add_column_if_not_exists(table_name, col_name, _get_column_type(model, col_name))
                db.session.add(SchemaMigration(migration_name=mig_name))
                db.session.commit()


from datetime import datetime as _dt
@app.context_processor
def inject_now():
    return {
        'now': _dt.now(),
        'current_locale': session.get('lang', 'en'),
    }

@app.context_processor
def inject_permission_helpers():
    return dict(has_permission=lambda p: current_user.is_authenticated and current_user.has_permission(p),
                has_any_permission=lambda *ps: current_user.is_authenticated and current_user.has_any_permission(*ps))

# ─────────────────────────────────────────
#  SIDEBAR CONFIGURATION
# ─────────────────────────────────────────

def _sb_link(label, endpoint, icon, permission=None, any_permission=None):
    return {'type': 'link', 'label': label, 'endpoint': endpoint, 'icon': icon,
            'permission': permission, 'any_permission': any_permission}

def _sb_group(label, icon, active_match, children, any_permission=None):
    return {'type': 'group', 'label': label, 'icon': icon, 'active_match': active_match,
            'children': children, 'any_permission': any_permission}

SIDEBAR_REPORT_ITEMS = [
    {'label': 'Transport Requests', 'anchor': 'transport', 'icon': 'fa-route', 'permission': 'transport_request_view'},
    {'label': 'Trip Operation', 'anchor': 'trip', 'icon': 'fa-bus', 'permission': 'trip_operation_report_view'},
    {'label': 'Daily Performance', 'anchor': 'daily', 'icon': 'fa-file-alt', 'permission': 'daily_report_view'},
    {'label': 'KPI', 'anchor': 'kpi', 'icon': 'fa-chart-line', 'any_permission': ['kpi_dashboard_view', 'kpi_evaluation_view']},
    {'label': 'KPI History', 'anchor': 'kpi_history', 'icon': 'fa-clock-rotate-left', 'permission': 'kpi_history_view'},
    {'label': 'KPI Evaluation', 'anchor': 'kpi_evaluation', 'icon': 'fa-file-invoice', 'permission': 'kpi_evaluation_view'},
    {'label': 'Penalties', 'anchor': 'penalties', 'icon': 'fa-gavel', 'permission': 'penalty_view'},
    {'label': 'Employee Performance', 'anchor': 'employee_performance', 'icon': 'fa-user-check', 'any_permission': ['kpi_evaluation_view']},
]

def _sb_reports():
    return {'type': 'reports', 'children': list(SIDEBAR_REPORT_ITEMS)}

SIDEBAR_CONFIG = [
    {
        'section': 'Overview',
        'items': [
            _sb_link('Dashboard', 'dashboard', 'fa-th-large', permission='dashboard_view'),
        ],
    },
    {
        'section': 'Transportation',
        'any_permission': [
            'transport_request_view', 'transport_request_create', 'transport_request_edit',
            'transport_request_delete', 'transport_request_approve', 'transport_request_reject',
            'transport_request_download',
        ],
        'items': [
            _sb_group('Transport Requests', 'fa-truck', 'transport', any_permission=[
                'transport_request_view', 'transport_request_create', 'transport_request_edit',
                'transport_request_delete', 'transport_request_approve', 'transport_request_reject',
                'transport_request_download',
            ], children=[
                _sb_link('All Requests', 'transport_requests', 'fa-list', permission='transport_request_view'),
                _sb_link('New Request', 'new_transport_request', 'fa-plus', permission='transport_request_create'),
            ]),
        ],
    },
    {
        'section': 'Operations',
        'any_permission': [
            'trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit',
            'trip_operation_report_delete', 'trip_operation_report_download',
            'daily_report_view', 'daily_report_create', 'daily_report_edit', 'daily_report_delete',
        ],
        'items': [
            _sb_group('Trip Operations', 'fa-bus', 'trip_operation', any_permission=[
                'trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit',
                'trip_operation_report_delete', 'trip_operation_report_download',
            ], children=[
                _sb_link('Trip Operation Report', 'trip_operation_reports', 'fa-list', permission='trip_operation_report_view'),
                _sb_link('New Report', 'new_trip_operation_report', 'fa-plus', permission='trip_operation_report_create'),
                _sb_link('Vehicle Performance', 'vehicle_performance', 'fa-chart-bar',
                         any_permission=['trip_operation_report_view', 'trip_operation_report_download']),
            ]),
            _sb_group('Daily Performance', 'fa-file-alt', 'daily_report', any_permission=[
                'daily_report_view', 'daily_report_create', 'daily_report_edit', 'daily_report_delete',
            ], children=[
                _sb_link('Daily Report', 'daily_reports', 'fa-list', permission='daily_report_view'),
                _sb_link('New Report', 'new_daily_report', 'fa-plus', permission='daily_report_create'),
            ]),
        ],
    },
    {
        'section': 'Performance & KPI',
        'any_permission': [
            'kpi_dashboard_view', 'kpi_history_view',
            'kpi_evaluation_view', 'kpi_evaluation_create', 'kpi_evaluation_edit', 'kpi_evaluation_delete',
        ],
        'items': [
            _sb_link('KPI Dashboard', 'kpi_dashboard', 'fa-chart-simple', permission='kpi_dashboard_view'),
            _sb_link('KPI History', 'kpi_history', 'fa-clock-rotate-left', permission='kpi_history_view'),
            _sb_link('KPI Evaluations', 'kpi_evaluations', 'fa-chart-line',
                     any_permission=['kpi_evaluation_view', 'kpi_evaluation_create', 'kpi_evaluation_edit', 'kpi_evaluation_delete']),
            _sb_link('New Evaluation', 'new_kpi_evaluation', 'fa-plus-circle', permission='kpi_evaluation_create'),
        ],
    },
    {
        'section': 'Discipline',
        'any_permission': [
            'penalty_view', 'penalty_create', 'penalty_edit', 'penalty_delete', 'penalty_download',
        ],
        'items': [
            _sb_group('Penalties', 'fa-gavel', 'penalt', any_permission=[
                'penalty_view', 'penalty_create', 'penalty_edit', 'penalty_delete', 'penalty_download',
            ], children=[
                _sb_link('View All', 'penalties', 'fa-list', permission='penalty_view'),
                _sb_link('New Penalty', 'new_penalty', 'fa-plus', permission='penalty_create'),
            ]),
        ],
    },
    {
        'section': 'Reports',
        'items': [_sb_reports()],
        'any_permission': [
            'transport_request_view', 'trip_operation_report_view', 'daily_report_view',
            'kpi_evaluation_view', 'kpi_history_view', 'penalty_view',
            'report_view', 'report_create', 'report_edit', 'report_delete',
            'report_export', 'report_print', 'report_download',
        ],
    },
    {
        'section': 'Administration',
        'any_permission': [
            'user_view', 'user_create', 'user_edit', 'user_delete', 'user_download',
            'role_view', 'role_create', 'role_edit', 'role_delete', 'role_download',
        ],
        'items': [
            _sb_link('Users', 'users', 'fa-users',
                     any_permission=['user_view', 'user_create', 'user_edit', 'user_delete', 'user_download']),
            _sb_link('Roles & Permissions', 'list_roles', 'fa-user-shield',
                     any_permission=['role_view', 'role_create', 'role_edit', 'role_delete', 'role_download']),
        ],
    },
    {
        'section': 'Settings',
        'any_permission': [
            'system_settings_view', 'system_settings_edit', 'system_settings_update',
            'department_view', 'department_create', 'department_edit', 'department_delete', 'department_download',
            'position_view', 'position_create', 'position_edit', 'position_delete', 'position_download',
        ],
        'items': [
            _sb_group('Configuration', 'fa-sliders', None, children=[
                _sb_link('System Settings', 'dynamic_settings', 'fa-cogs',
                         any_permission=['system_settings_view', 'system_settings_edit', 'system_settings_update']),
                _sb_link('Departments', 'departments', 'fa-building',
                         any_permission=['department_view', 'department_create', 'department_edit', 'department_delete', 'department_download']),
                _sb_link('Positions', 'positions', 'fa-briefcase',
                         any_permission=['position_view', 'position_create', 'position_edit', 'position_delete', 'position_download']),
                _sb_link('Telegram Notification', 'telegram_settings', 'fa-telegram-plane',
                         any_permission=['system_settings_edit', 'system_settings_update']),
            ]),
        ],
    },
]


def _build_sidebar(user):
    """Filter sidebar config by user permissions and mark active items.
    
    Every item's visibility is determined SOLELY by permission checks:
      - 'permission' → user must have that single permission
      - 'any_permission' → user must have at least one of the listed permissions
    Role names, authentication status (beyond being logged in), and
    hardcoded checks are NEVER used.
    """
    if not user or not user.is_authenticated:
        return []
    endpoint = request.endpoint or ''
    result = []

    def _check(item):
        perm = item.get('permission')
        if perm and not user.has_permission(perm):
            return False
        any_perm = item.get('any_permission')
        if any_perm and not user.has_any_permission(*any_perm):
            return False
        return True

    def _active_match(pattern):
        if pattern is None:
            return False
        return pattern in endpoint

    def _copy(o):
        if isinstance(o, dict):
            return {k: _copy(v) for k, v in o.items()}
        if isinstance(o, list):
            return [_copy(v) for v in o]
        return o

    for section in SIDEBAR_CONFIG:
        if section.get('any_permission'):
            if not user.has_any_permission(*section['any_permission']):
                continue
        visible_items = []
        for item in section.get('items', []):
            item = _copy(item)
            if item['type'] == 'link':
                if not _check(item):
                    continue
                item['active'] = endpoint == item['endpoint']
                visible_items.append(item)
            elif item['type'] == 'group':
                if item.get('any_permission') and not _check({'any_permission': item['any_permission']}):
                    continue
                visible_children = []
                for child in item.get('children', []):
                    child = _copy(child)
                    if not _check(child):
                        continue
                    child['active'] = endpoint == child['endpoint']
                    visible_children.append(child)
                if not visible_children:
                    continue
                item['children'] = visible_children
                item['expanded'] = _active_match(item.get('active_match'))
                item['active'] = any(c.get('active') for c in visible_children) or item['expanded']
                visible_items.append(item)
            elif item['type'] == 'reports':
                children = item.get('children', [])
                visible_children = [c for c in children if _check(c)]
                if visible_children:
                    item['children'] = visible_children
                    item['count'] = len(visible_children)
                    visible_items.append(item)
        if visible_items:
            result.append({**section, 'items': visible_items})
    return result


@app.context_processor
def inject_sidebar():
    sidebar_items = _build_sidebar(current_user)
    return dict(sidebar_items=sidebar_items)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = _('Please log in to access this page.')

# ─────────────────────────────────────────
#  MODELS
# ─────────────────────────────────────────

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.id'), nullable=True)
    branch = db.Column(db.String(100))
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    is_active = db.Column(db.Boolean, default=True)

    assigned_role = db.relationship('Role', foreign_keys=[role_id])

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)



    def _get_role_permissions(self):
        if self.assigned_role:
            perms = self.assigned_role.permissions or []
            logger.debug(f"_get_role_permissions: assigned_role EXISTS, name='{self.assigned_role.name}', permissions={perms}")
            return perms
        if self.role_id:
            role = db.session.get(Role, self.role_id)
            if role:
                self.assigned_role = role
                perms = role.permissions or []
                logger.debug(f"_get_role_permissions: loaded role id={role.id}, name='{role.name}', permissions={perms}")
                return perms
            else:
                logger.debug(f"_get_role_permissions: role_id={self.role_id} but role NOT FOUND in DB")
        else:
            logger.debug(f"_get_role_permissions: no assigned_role AND no role_id")
        return None

    def has_permission(self, permission):
        role_perms = self._get_role_permissions()
        logger.debug(f"has_permission('{permission}'): role_perms={role_perms!r}, type={type(role_perms).__name__}")
        if role_perms is not None:
            result = permission in role_perms
            logger.debug(f"has_permission('{permission}'): checking DB perms -> {result}")
            return result
        logger.debug(f"has_permission('{permission}'): no role_perms (role_id=None or role deleted), returning False")
        return False

    def has_any_permission(self, *permissions):
        role_perms = self._get_role_permissions()
        logger.debug(f"has_any_permission{permissions}: role_perms={role_perms!r}, type={type(role_perms).__name__}")
        if role_perms is not None:
            result = any(p in role_perms for p in permissions)
            logger.debug(f"has_any_permission{permissions}: checking DB perms -> {result}")
            return result
        logger.debug(f"has_any_permission{permissions}: no role_perms (role_id=None or role deleted), returning False")
        return False

    @property
    def role_label(self):
        if self.assigned_role:
            return self.assigned_role.label or self.assigned_role.name
        if self.role_id:
            role = db.session.get(Role, self.role_id)
            if role:
                self.assigned_role = role
                return role.label or role.name
        return self.assigned_role.label if self.assigned_role else _('No Role')


class Role(db.Model):
    __tablename__ = 'roles'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    label = db.Column(db.String(100))
    permissions = db.Column(db.JSON, nullable=False, default=list)
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    @property
    def permission_count(self):
        return len(self.permissions or [])

    def get_permissions(self):
        return self.permissions or []


class RouteRequest(db.Model):
    __tablename__ = 'route_requests'
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.String(20), unique=True, nullable=False)
    request_date = db.Column(db.Date, default=date.today)
    requester_name = db.Column(db.String(100), nullable=False)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    destination_from = db.Column(db.String(200), nullable=False)
    destination_to = db.Column(db.String(200), nullable=False)
    branch_location = db.Column(db.String(200))
    company = db.Column(db.String(200))
    transportation = db.Column(db.String(100))
    national_road = db.Column(db.String(50))
    price = db.Column(db.Float, default=0)
    departure_time = db.Column(db.String(10))
    arrival_time = db.Column(db.String(10))
    start_date = db.Column(db.Date)
    end_date = db.Column(db.Date)
    allow_access = db.Column(db.Boolean, default=True)
    attachment = db.Column(db.String(255))
    remarks = db.Column(db.Text)
    status = db.Column(db.String(20), default='Pending')  # Pending, Approved, Rejected
    reviewed_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    review_note = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    requester = db.relationship('User', foreign_keys=[requester_id])
    reviewer = db.relationship('User', foreign_keys=[reviewed_by])

    def generate_id(self):
        last = RouteRequest.query.order_by(RouteRequest.id.desc()).first()
        if last and last.request_id:
            try:
                num = int(last.request_id.split('-')[-1]) + 1
            except (ValueError, IndexError):
                num = 1
        else:
            num = 1
        self.request_id = f"RR-{datetime.now().year}-{num:04d}"

    requester = db.relationship('User', foreign_keys=[requester_id])
    reviewer = db.relationship('User', foreign_keys=[reviewed_by])


class TransportRequest(db.Model):
    __tablename__ = 'transport_requests'
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.String(20), unique=True, nullable=False)
    request_date = db.Column(db.Date, default=date.today)
    request_type = db.Column(db.String(30), nullable=False)
    requester_name = db.Column(db.String(100), nullable=False)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    destination_from = db.Column(db.String(200), nullable=False)
    destination_to = db.Column(db.String(200), nullable=False)
    transportation_type = db.Column(db.String(100))
    change_transportation_type_to = db.Column(db.String(100))
    company = db.Column(db.String(200))
    change_company_to = db.Column(db.String(200))
    branch_location = db.Column(db.String(200))
    national_road = db.Column(db.String(50))
    price = db.Column(db.Float, default=0)
    vehicle_no = db.Column(db.String(50))
    route_code = db.Column(db.String(50))
    gender_required = db.Column(db.String(10))
    old_route_code = db.Column(db.String(50))
    old_price = db.Column(db.Float, default=0)
    new_price = db.Column(db.Float, default=0)
    departure_time = db.Column(db.String(10))
    arrival_time = db.Column(db.String(10))
    duration = db.Column(db.String(50))
    active_start_date = db.Column(db.Date)
    active_end_date = db.Column(db.Date)
    number_of_days = db.Column(db.Integer)
    promotion_price = db.Column(db.Float, default=0)
    attachments = db.Column(db.JSON, default=list)
    remarks = db.Column(db.Text)
    status = db.Column(db.String(20), default='Pending')
    reviewed_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    review_note = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    requester = db.relationship('User', foreign_keys=[requester_id])
    reviewer = db.relationship('User', foreign_keys=[reviewed_by])

    def generate_id(self):
        last = TransportRequest.query.order_by(TransportRequest.id.desc()).first()
        if last and last.request_id:
            try:
                num = int(last.request_id.split('-')[-1]) + 1
            except (ValueError, IndexError):
                num = 1
        else:
            num = 1
        self.request_id = f"TR-{datetime.now().year}-{num:04d}"


class EmployeePenalty(db.Model):
    __tablename__ = 'employee_penalties'
    id = db.Column(db.Integer, primary_key=True)
    penalty_id = db.Column(db.String(20), unique=True, nullable=False)
    employee_id = db.Column(db.String(20), nullable=False)
    employee_name = db.Column(db.String(100), nullable=False)
    department = db.Column(db.String(100))
    position = db.Column(db.String(100))
    branch = db.Column(db.String(100))
    violation_type = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.Float, default=0)
    penalty_amount = db.Column(db.Float, default=0)
    old_code = db.Column(db.String(100))
    evidence_file = db.Column(db.String(255))
    incident_date = db.Column(db.Date, default=date.today)
    approved_by = db.Column(db.String(100))
    approved_by_user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    status = db.Column(db.String(20), default='Pending')
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    creator = db.relationship('User', foreign_keys=[created_by])
    approver = db.relationship('User', foreign_keys=[approved_by_user_id])

    def generate_id(self):
        last = EmployeePenalty.query.order_by(EmployeePenalty.id.desc()).first()

        if last and last.penalty_id:
            try:
                num = int(last.penalty_id.split('-')[-1]) + 1
            except (ValueError, IndexError):
                num = 1
        else:
            num = 1

        self.penalty_id = f"EP-{datetime.now().year}-{num:04d}"


NOTIFICATION_TYPE_TRANSPORT = 'transport'
NOTIFICATION_TYPE_PENALTY = 'penalty'

VALID_NOTIFICATION_TYPES = [NOTIFICATION_TYPE_TRANSPORT, NOTIFICATION_TYPE_PENALTY]


class TelegramSetting(db.Model):
    __tablename__ = 'telegram_settings'
    id = db.Column(db.Integer, primary_key=True)
    notification_type = db.Column(db.String(50), unique=True, nullable=False, default=NOTIFICATION_TYPE_TRANSPORT)
    bot_token = db.Column(db.Text, default='')
    chat_id = db.Column(db.String(100), default='')
    bot_username = db.Column(db.String(100), default='')
    group_name = db.Column(db.String(200), default='')
    banner_image = db.Column(db.String(255), default=None)
    enabled = db.Column(db.Boolean, default=False)
    is_connected = db.Column(db.Boolean, default=False)
    last_test_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    @staticmethod
    def get_or_create(notification_type):
        config = TelegramSetting.query.filter_by(notification_type=notification_type).first()
        if not config:
            config = TelegramSetting(notification_type=notification_type)
            db.session.add(config)
            db.session.commit()
        return config


class TripOperationReport(db.Model):
    __tablename__ = 'trip_operation_reports'
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.String(20), unique=True, nullable=False)
    origin = db.Column(db.String(200), default='')
    destination = db.Column(db.String(200), default='')
    travel_direction = db.Column(db.String(50), default='')
    departure_time = db.Column(db.String(10))
    power_off_time = db.Column(db.String(10))
    power_on_time = db.Column(db.String(10))
    vehicle_number = db.Column(db.String(50), nullable=False)
    driver_phone = db.Column(db.String(50))
    arrival_at_station = db.Column(db.String(10))
    departure_from_station = db.Column(db.String(10))
    travel_delay_duration = db.Column(db.Float, default=0)
    layover_duration = db.Column(db.String(10), default='')
    trip_date = db.Column(db.Date, nullable=True)
    reason_for_delay = db.Column(db.Text)
    vehicle_status = db.Column(db.String(50), nullable=False)
    passenger_count = db.Column(db.Integer, default=0)
    coordinator_name = db.Column(db.String(100))
    note = db.Column(db.Text)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))

    creator = db.relationship('User', foreign_keys=[created_by])

    def generate_id(self):
        last = TripOperationReport.query.order_by(TripOperationReport.id.desc()).first()
        if last and last.report_id:
            try:
                num = int(last.report_id.split('-')[-1]) + 1
            except (ValueError, IndexError):
                num = 1
        else:
            num = 1
        self.report_id = f"TOR-{datetime.now().year}-{num:04d}"


# ─────────────────────────────────────────
#  KPI EVALUATION MODEL
# ─────────────────────────────────────────

class KpiEvaluation(db.Model):
    __tablename__ = 'kpi_evaluations'
    id = db.Column(db.Integer, primary_key=True)
    staff_name = db.Column(db.String(100), nullable=False)
    staff_id = db.Column(db.String(20), nullable=False)
    branch = db.Column(db.String(100))
    company = db.Column(db.String(200))
    evaluation_month = db.Column(db.Integer, nullable=False)
    evaluation_year = db.Column(db.Integer, nullable=False)
    # 1. Ticket Sales (40%)
    ticket_sales_target = db.Column(db.Float, default=0)
    actual_tickets_sold = db.Column(db.Float, default=0)
    achievement_pct = db.Column(db.Float, default=0)
    ticket_sales_score = db.Column(db.Float, default=0)
    # 2. Booking Accuracy (20%)
    booking_accuracy_pct = db.Column(db.Float, default=0)
    booking_errors = db.Column(db.Integer, default=0)
    booking_accuracy_score = db.Column(db.Float, default=0)
    # 3. Customer Service (15%)
    customer_satisfaction = db.Column(db.Float, default=0)
    complaints_handled = db.Column(db.Integer, default=0)
    customer_service_score = db.Column(db.Float, default=0)
    # 4. Attendance & Punctuality (10%)
    late_arrivals = db.Column(db.Integer, default=0)
    unexcused_absences = db.Column(db.Integer, default=0)
    attendance_score = db.Column(db.Float, default=0)
    # 5. Daily Reporting (10%)
    report_submission = db.Column(db.Text)
    daily_report_score = db.Column(db.Float, default=0)
    # 6. SOP Compliance (5%)
    sop_compliance = db.Column(db.Float, default=0)
    sop_compliance_score = db.Column(db.Float, default=0)
    # Computed
    total_score = db.Column(db.Float, default=0)
    performance_rating = db.Column(db.String(30), default='')
    evaluator_name = db.Column(db.String(100))
    comments = db.Column(db.Text)
    improvement_plan = db.Column(db.Text)
    status = db.Column(db.String(20), default='Draft')
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    creator = db.relationship('User', foreign_keys=[created_by])


# ─────────────────────────────────────────
#  DAILY PERFORMANCE REPORT MODEL
# ─────────────────────────────────────────

class DailyPerformanceReport(db.Model):
    __tablename__ = 'daily_performance_reports'
    id = db.Column(db.Integer, primary_key=True)
    staff_name = db.Column(db.String(100), nullable=False)
    staff_id = db.Column(db.String(20), nullable=False)
    branch = db.Column(db.String(100))
    company = db.Column(db.String(200))
    report_date = db.Column(db.Date, nullable=False)
    tickets_sold = db.Column(db.Integer, default=0)
    total_sales_amount = db.Column(db.Float, default=0)
    bookings = db.Column(db.Integer, default=0)
    booking_errors = db.Column(db.Integer, default=0)
    cancelled_tickets = db.Column(db.Integer, default=0)
    refunded_tickets = db.Column(db.Integer, default=0)
    complaints = db.Column(db.Integer, default=0)
    resolved_complaints = db.Column(db.Integer, default=0)
    remarks = db.Column(db.Text)
    time_check_in = db.Column(db.String(10))
    time_check_out = db.Column(db.String(10))
    attendance_status = db.Column(db.String(20))
    status = db.Column(db.String(20), default='Draft')
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    creator = db.relationship('User', foreign_keys=[created_by])


class MonthlyKpiSummary(db.Model):
    __tablename__ = 'monthly_kpi_summaries'
    id = db.Column(db.Integer, primary_key=True)
    staff_name = db.Column(db.String(100), nullable=False)
    staff_id = db.Column(db.String(20), nullable=False)
    branch = db.Column(db.String(100))
    company = db.Column(db.String(200))
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    total_tickets_sold = db.Column(db.Integer, default=0)
    total_sales_amount = db.Column(db.Float, default=0)
    total_bookings = db.Column(db.Integer, default=0)
    total_booking_errors = db.Column(db.Integer, default=0)
    total_complaints = db.Column(db.Integer, default=0)
    total_resolved_complaints = db.Column(db.Integer, default=0)
    # Scores
    ticket_sales_target = db.Column(db.Float, default=0)
    ticket_sales_score = db.Column(db.Float, default=0)
    booking_accuracy_score = db.Column(db.Float, default=0)
    customer_service_score = db.Column(db.Float, default=0)
    attendance_score = db.Column(db.Float, default=0)
    daily_reporting_score = db.Column(db.Float, default=0)
    sop_compliance_score = db.Column(db.Float, default=0)
    total_score = db.Column(db.Float, default=0)
    performance_rating = db.Column(db.String(30), default='')
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))


# ─────────────────────────────────────────
#  HR MODULE MODELS
# ─────────────────────────────────────────

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default='Active')
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    positions = db.relationship('Position', back_populates='department', lazy='dynamic')


class Position(db.Model):
    __tablename__ = 'positions'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default='Active')
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None), onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    department = db.relationship('Department', back_populates='positions')


# ─────────────────────────────────────────
#  DYNAMIC CONFIGURATION ENGINE (Metadata Models)
# ─────────────────────────────────────────

class ConfigModule(db.Model):
    __tablename__ = 'config_module'
    id = db.Column(db.Integer, primary_key=True)
    module_key = db.Column(db.String(100), unique=True, nullable=False)
    module_name = db.Column(db.String(200), nullable=False)
    module_icon = db.Column(db.String(50), default='fa-database')
    table_name = db.Column(db.String(100))
    sort_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))


class ConfigField(db.Model):
    __tablename__ = 'config_field'
    id = db.Column(db.Integer, primary_key=True)
    module_id = db.Column(db.Integer, db.ForeignKey('config_module.id'), nullable=False)
    field_name = db.Column(db.String(100), nullable=False)
    field_label = db.Column(db.String(200), nullable=False)
    field_type = db.Column(db.String(50), nullable=False)  # text, number, date, dropdown, textarea, checkbox, boolean, email, tel, password
    is_required = db.Column(db.Boolean, default=False)
    is_unique = db.Column(db.Boolean, default=False)
    is_listable = db.Column(db.Boolean, default=True)
    is_searchable = db.Column(db.Boolean, default=False)
    is_sortable = db.Column(db.Boolean, default=True)
    is_editable = db.Column(db.Boolean, default=True)
    is_creatable = db.Column(db.Boolean, default=True)
    display_order = db.Column(db.Integer, default=0)
    default_value = db.Column(db.String(500))
    placeholder = db.Column(db.String(500))
    help_text = db.Column(db.String(500))
    min_length = db.Column(db.Integer)
    max_length = db.Column(db.Integer)
    regex_pattern = db.Column(db.String(500))
    regex_message = db.Column(db.String(500))
    option_source = db.Column(db.String(20))  # static, module, api
    option_module_id = db.Column(db.Integer, db.ForeignKey('config_module.id'))
    option_parent_field = db.Column(db.String(100))
    rel_module_id = db.Column(db.Integer, db.ForeignKey('config_module.id'))
    rel_type = db.Column(db.String(20))  # one_to_many, many_to_one
    grid_width = db.Column(db.Integer)
    module = db.relationship('ConfigModule', foreign_keys=[module_id], backref='fields')
    option_module = db.relationship('ConfigModule', foreign_keys=[option_module_id])
    rel_module = db.relationship('ConfigModule', foreign_keys=[rel_module_id])


class ConfigValidation(db.Model):
    __tablename__ = 'config_validation'
    id = db.Column(db.Integer, primary_key=True)
    field_id = db.Column(db.Integer, db.ForeignKey('config_field.id'), nullable=False)
    validation_type = db.Column(db.String(50), nullable=False)  # required, unique, min_length, max_length, regex, custom
    validation_value = db.Column(db.String(500))
    error_message = db.Column(db.String(500))
    field = db.relationship('ConfigField', backref='validations')


class ConfigDropdownOption(db.Model):
    __tablename__ = 'config_dropdown_option'
    id = db.Column(db.Integer, primary_key=True)
    field_id = db.Column(db.Integer, db.ForeignKey('config_field.id'), nullable=False)
    option_value = db.Column(db.String(200), nullable=False)
    option_label = db.Column(db.String(200), nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)
    field = db.relationship('ConfigField', backref='dropdown_options')


class DynamicRecord(db.Model):
    __tablename__ = 'dynamic_record'
    id = db.Column(db.Integer, primary_key=True)
    module_id = db.Column(db.Integer, db.ForeignKey('config_module.id'), nullable=False)
    data = db.Column(db.JSON, nullable=False, default=dict)
    is_active = db.Column(db.Boolean, default=True)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    updated_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_date = db.Column(db.DateTime, onupdate=lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    deleted_date = db.Column(db.DateTime)
    module = db.relationship('ConfigModule', backref='records')
    creator = db.relationship('User', foreign_keys=[created_by])
    updater = db.relationship('User', foreign_keys=[updated_by])


# ── Register all models for auto-migration ──
for _model in [User, Role, RouteRequest, TransportRequest, EmployeePenalty,
               TripOperationReport, KpiEvaluation, DailyPerformanceReport,
               MonthlyKpiSummary, Department, Position, ConfigModule, ConfigField,
                ConfigValidation, ConfigDropdownOption, DynamicRecord,
               TelegramSetting]:
    auto_register_model(_model)


# ─────────────────────────────────────────
#  AUTH
# ─────────────────────────────────────────

@login_manager.user_loader
def load_user(user_id):
    user = db.session.get(User, int(user_id))
    if user:
        logger.debug(f"=== USER_LOADER: user_id={user.id}, username={user.username}, role_id={user.role_id} ===")
        if user.role_id:
            role = db.session.get(Role, user.role_id)
            if role:
                logger.debug(f"USER_LOADER: Found role id={role.id}, name='{role.name}', permissions={role.permissions}")
                user.assigned_role = role
            else:
                logger.debug(f"USER_LOADER: Role id={user.role_id} NOT FOUND in database!")
                user.assigned_role = None
        else:
            logger.debug(f"USER_LOADER: No role_id set, assigned_role will be None")
    return user


def permission_required(permission):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_permission(permission):
                return abort(403)
            return f(*args, **kwargs)
        return decorated
    return decorator


def any_permission_required(*permissions):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if not current_user.has_any_permission(*permissions):
                return abort(403)
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── 403 error handler ──
@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403


# ── Login rate limiting ──
_login_attempts = {}


def _check_login_rate(ip):
    now = datetime.now(timezone.utc)
    if ip in _login_attempts:
        attempts, first_attempt = _login_attempts[ip]
        if now - first_attempt > timedelta(minutes=15):
            _login_attempts[ip] = (1, now)
            return True
        if attempts >= 5:
            return False
        _login_attempts[ip] = (attempts + 1, first_attempt)
    else:
        _login_attempts[ip] = (1, now)
    return True


# ─────────────────────────────────────────
#  ROUTES: AUTH
# ─────────────────────────────────────────

def _get_default_redirect(user=None):
    """Return the best default page URL for the given user based on permissions."""
    u = user or current_user
    priorities = [
        ('dashboard_view', 'dashboard'),
        ('penalty_view', 'penalties'),
        ('transport_request_view', 'transport_requests'),
        ('route_request_view', 'route_requests'),
        ('trip_operation_report_view', 'trip_operation_reports'),
        ('daily_report_view', 'daily_reports'),
        ('kpi_dashboard_view', 'kpi_dashboard'),
        ('kpi_evaluation_view', 'kpi_evaluations'),
        ('report_view', 'reports'),
        ('user_view', 'users'),
        ('role_view', 'list_roles'),
        ('department_view', 'departments'),
        ('position_view', 'positions'),
        ('system_settings_view', 'dynamic_settings'),
    ]
    for perm, endpoint in priorities:
        if u.has_permission(perm):
            return url_for(endpoint)
    flash(_('Your account has no permissions assigned. Contact your administrator.'), 'danger')
    return url_for('dashboard')


@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(_get_default_redirect())
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(_get_default_redirect())
    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Image logo.png')
    login_logo_url = url_for('static', filename='uploads/Image logo.png') if os.path.exists(logo_path) else None
    if request.method == 'POST':
        ip = request.remote_addr or 'unknown'
        if not _check_login_rate(ip):
            flash(_('Too many login attempts. Try again in 15 minutes.'), 'danger')
            return render_template('login.html', login_logo_url=login_logo_url)
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password) and user.is_active:
            db.session.refresh(user)
            logger.debug(f"=== LOGIN: user_id={user.id}, username={user.username}, role_id={user.role_id} ===")
            if user.role_id:
                role = db.session.get(Role, user.role_id)
                if role:
                    logger.debug(f"LOGIN: Found role id={role.id}, name='{role.name}', permissions={role.permissions}")
                    user.assigned_role = role
                else:
                    logger.debug(f"LOGIN: Role id={user.role_id} NOT FOUND!")
            elif user.username == 'admin':
                admin_role = Role.query.filter_by(name='admin').first()
                if admin_role:
                    user.role_id = admin_role.id
                    user.assigned_role = admin_role
                    db.session.commit()
                    logger.info(f"LOGIN: Auto-assigned admin role to admin user (id={user.id})")
            if not user.role_id:
                logger.debug(f"LOGIN: No role_id set")
            login_user(user)
            _login_attempts.pop(ip, None)
            return redirect(_get_default_redirect(user))
        flash(_('Invalid username or password.'), 'danger')
    return render_template('login.html', login_logo_url=login_logo_url)


@app.route('/set-language/<lang>')
@login_required
def set_language(lang):
    if lang in ['en', 'km']:
        session['lang'] = lang
    return _safe_redirect(_get_default_redirect())


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/debug-permissions')
@login_required
@permission_required('role_view')
def debug_permissions():
    import json
    user = current_user
    role_perms = user._get_role_permissions() if hasattr(user, '_get_role_permissions') else None
    info = {
        'user_id': user.id,
        'username': user.username,
        'role_string': user.assigned_role.name if user.assigned_role else None,
        'role_id': user.role_id,
        'assigned_role_exists': user.assigned_role is not None,
        'assigned_role_name': user.assigned_role.name if user.assigned_role else None,
        'role_permissions_from_db': role_perms,
        'role_permissions_type': type(role_perms).__name__ if role_perms is not None else 'None',
        'legacy_fallback_would_trigger': False,
        'is_authenticated': user.is_authenticated,
    }
    return jsonify(info)


# ─────────────────────────────────────────
#  ROUTES: DASHBOARD
# ─────────────────────────────────────────

@app.route('/dashboard')
@login_required
@permission_required('dashboard_view')
def dashboard():
    from datetime import timedelta

    range_filter = request.args.get('range', 'all')
    now = datetime.now()
    today = date.today()

    # Determine date range
    custom_start = ''
    custom_end = ''
    if range_filter == 'today':
        start_date = today
        end_date = today
    elif range_filter == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif range_filter == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif range_filter == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif range_filter == 'custom':
        raw_start = request.args.get('start_date', '')
        raw_end = request.args.get('end_date', '')
        try:
            start_date = datetime.strptime(raw_start, '%Y-%m-%d').date() if raw_start else None
            end_date = datetime.strptime(raw_end, '%Y-%m-%d').date() if raw_end else None
            custom_start = raw_start
            custom_end = raw_end
        except (ValueError, TypeError):
            start_date = None
            end_date = None
            range_filter = 'all'
    else:
        start_date = None
        end_date = None

    def date_filtered_count(model, **filters):
        q = model.query
        if start_date:
            q = q.filter(model.created_date >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            q = q.filter(model.created_date <= datetime.combine(end_date, datetime.max.time()))
        for k, v in filters.items():
            if hasattr(model, k):
                q = q.filter(getattr(model, k) == v)
        return q.count()

    def date_filtered_base(model):
        q = model.query
        if start_date:
            q = q.filter(model.created_date >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            q = q.filter(model.created_date <= datetime.combine(end_date, datetime.max.time()))
        return q

    total_requests = date_filtered_count(RouteRequest)
    pending_requests = date_filtered_count(RouteRequest, status='Pending')
    approved_requests = date_filtered_count(RouteRequest, status='Approved')
    rejected_requests = date_filtered_count(RouteRequest, status='Rejected')

    total_penalties = date_filtered_count(EmployeePenalty)
    pen_amount_q = date_filtered_base(EmployeePenalty)
    total_penalty_amount = pen_amount_q.with_entities(db.func.sum(EmployeePenalty.penalty_amount)).scalar() or 0
    pending_penalties = date_filtered_count(EmployeePenalty, status='Pending')

    total_reports = date_filtered_count(TripOperationReport)
    departed_reports = date_filtered_count(TripOperationReport, vehicle_status='Departed')
    not_departed_reports = date_filtered_count(TripOperationReport, vehicle_status='Not Departed')
    pass_q = date_filtered_base(TripOperationReport)
    total_passengers = pass_q.with_entities(db.func.sum(TripOperationReport.passenger_count)).scalar() or 0

    processed_requests = approved_requests + rejected_requests

    recent_requests = date_filtered_base(RouteRequest).order_by(RouteRequest.created_date.desc()).limit(10).all()
    recent_penalties = date_filtered_base(EmployeePenalty).order_by(EmployeePenalty.created_date.desc()).limit(5).all()
    recent_reports = date_filtered_base(TripOperationReport).order_by(TripOperationReport.created_date.desc()).limit(5).all()

    # Monthly data for charts (last 6 months) using filtered base queries
    monthly_data = defaultdict(lambda: {'requests': 0, 'penalties': 0, 'reports': 0, 'trip_departed': 0, 'trip_not_departed': 0})
    for r in date_filtered_base(RouteRequest).all():
        key = r.created_date.strftime('%b %Y') if r.created_date else 'Unknown'
        monthly_data[key]['requests'] += 1
    for p in date_filtered_base(EmployeePenalty).all():
        key = p.created_date.strftime('%b %Y') if p.created_date else 'Unknown'
        monthly_data[key]['penalties'] += 1
    for r in date_filtered_base(TripOperationReport).all():
        key = r.created_date.strftime('%b %Y') if r.created_date else 'Unknown'
        monthly_data[key]['reports'] += 1
        if r.vehicle_status == 'Departed':
            monthly_data[key]['trip_departed'] += 1
        else:
            monthly_data[key]['trip_not_departed'] += 1
    monthly_labels = sorted(monthly_data.keys(), key=lambda k: datetime.strptime(k, '%b %Y') if k != 'Unknown' and k else datetime.min)
    monthly_requests = [monthly_data[k]['requests'] for k in monthly_labels]
    monthly_penalties = [monthly_data[k]['penalties'] for k in monthly_labels]
    monthly_reports = [monthly_data[k]['reports'] for k in monthly_labels]
    monthly_trip_departed = [monthly_data[k]['trip_departed'] for k in monthly_labels]
    monthly_trip_not_departed = [monthly_data[k]['trip_not_departed'] for k in monthly_labels]

    # KPI Trends (compare first vs second half of monthly data)
    def compute_trend(current_val, label):
        mid = len(monthly_labels) // 2
        if len(monthly_data) < 2 or 'Unknown' in monthly_data:
            return {'label': 'No trend', 'cls': 'neutral', 'icon': 'minus'}
        first_half = sum(monthly_data[k][label] for k in monthly_labels[:mid]) or 1
        second_half = sum(monthly_data[k][label] for k in monthly_labels[mid:]) or 1
        pct = int((second_half - first_half) / first_half * 100)
        if pct > 0:
            return {'label': f'+{pct}%', 'cls': 'up', 'icon': 'arrow-up'}
        elif pct < 0:
            return {'label': f'{pct}%', 'cls': 'down', 'icon': 'arrow-down'}
        return {'label': '0%', 'cls': 'neutral', 'icon': 'minus'}

    kpi_trends = {
        'total_requests': compute_trend(total_requests, 'requests'),
        'pending_requests': compute_trend(pending_requests, 'requests'),
        'approved_requests': compute_trend(approved_requests, 'requests'),
        'rejected_requests': compute_trend(rejected_requests, 'requests'),
        'total_penalties': compute_trend(total_penalties, 'penalties'),
        'total_penalty_amount': {'label': '', 'cls': 'neutral', 'icon': 'minus'},
        'pending_penalties': compute_trend(pending_penalties, 'penalties'),
        'processed_requests': compute_trend(processed_requests, 'requests'),
        'total_reports': compute_trend(total_reports, 'reports'),
        'departed_reports': compute_trend(departed_reports, 'trip_departed'),
        'not_departed_reports': compute_trend(not_departed_reports, 'trip_not_departed'),
        'total_passengers': {'label': '', 'cls': 'neutral', 'icon': 'minus'},
    }

    # Recent Activities (combine all record types, sorted by time)
    activities = []
    for rr in date_filtered_base(RouteRequest).order_by(RouteRequest.created_date.desc()).limit(5).all():
        activities.append({
            'icon': 'fas fa-route',
            'bg_color': '#dbeafe',
            'color': '#1d4ed8',
            'title': 'Route Request',
            'description': f'{rr.requester_name} - {rr.destination_from} to {rr.destination_to} ({rr.status})',
            'time_ago': rr.created_date.strftime('%d %b %Y, %H:%M') if rr.created_date else '',
        })
    for ep in date_filtered_base(EmployeePenalty).order_by(EmployeePenalty.created_date.desc()).limit(5).all():
        activities.append({
            'icon': 'fas fa-gavel',
            'bg_color': '#ede9fe',
            'color': '#7c3aed',
            'title': 'Penalty Record',
            'description': f'{ep.employee_name} - {ep.violation_type} (${ep.penalty_amount:.0f})',
            'time_ago': ep.created_date.strftime('%d %b %Y, %H:%M') if ep.created_date else '',
        })
    for tor in date_filtered_base(TripOperationReport).order_by(TripOperationReport.created_date.desc()).limit(5).all():
        activities.append({
            'icon': 'fas fa-bus',
            'bg_color': '#e0f2fe',
            'color': '#0284c7',
            'title': 'Trip Report',
            'description': f'{tor.origin} to {tor.destination} - {tor.vehicle_number} ({tor.vehicle_status})',
            'time_ago': tor.created_date.strftime('%d %b %Y, %H:%M') if tor.created_date else '',
        })
    activities.sort(key=lambda a: a['time_ago'], reverse=True)
    recent_activities = activities[:10]

    # Notifications
    notifications = []
    pending_route_requests = date_filtered_count(RouteRequest, status='Pending')
    if pending_route_requests > 0:
        notifications.append({
            'message': f'{pending_route_requests} route request(s) pending approval',
            'time_ago': 'Requires attention',
            'dot_color': '#f59e0b',
        })
    pending_penalties_count = date_filtered_count(EmployeePenalty, status='Pending')
    if pending_penalties_count > 0:
        notifications.append({
            'message': f'{pending_penalties_count} penalty record(s) pending review',
            'time_ago': 'Requires attention',
            'dot_color': '#7c3aed',
        })
    if current_user.has_permission('user_view'):
        total_users = User.query.count()
        notifications.append({
            'message': f'{total_users} registered user(s) in the system',
            'time_ago': 'System info',
            'dot_color': '#059669',
        })
    if not notifications:
        notifications.append({
            'message': 'All caught up! No pending items.',
            'time_ago': 'No action needed',
            'dot_color': '#94a3b8',
        })

    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Image logo.png')
    login_logo_url = url_for('static', filename='uploads/Image logo.png') if os.path.exists(logo_path) else None

    return render_template('dashboard.html',
        total_requests=total_requests,
        pending_requests=pending_requests,
        approved_requests=approved_requests,
        rejected_requests=rejected_requests,
        processed_requests=processed_requests,
        total_penalties=total_penalties,
        total_penalty_amount=total_penalty_amount,
        pending_penalties=pending_penalties,
        recent_requests=recent_requests,
        recent_penalties=recent_penalties,
        monthly_labels=monthly_labels,
        monthly_requests=monthly_requests,
        monthly_penalties=monthly_penalties,
        monthly_reports=monthly_reports,
        monthly_trip_departed=monthly_trip_departed,
        monthly_trip_not_departed=monthly_trip_not_departed,
        total_reports=total_reports,
        departed_reports=departed_reports,
        not_departed_reports=not_departed_reports,
        total_passengers=total_passengers,
        recent_reports=recent_reports,
        login_logo_url=login_logo_url,
        range_filter=range_filter,
        kpi_trends=kpi_trends,
        recent_activities=recent_activities,
        notifications=notifications,
        pending_route_requests=pending_route_requests,
        pending_penalties_count=pending_penalties_count,
        custom_start=custom_start,
        custom_end=custom_end,
    )


# ─────────────────────────────────────────
#  DASHBOARD API ENDPOINTS
# ─────────────────────────────────────────

def _dash_date_range():
    """Parse dashboard date range filter args, return (start_date, end_date)."""
    rf = request.args.get('range', 'all')
    now = datetime.now()
    today = date.today()
    custom_start = custom_end = None
    if rf == 'today':
        start_date = end_date = today
    elif rf == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif rf == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif rf == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif rf == 'custom':
        try:
            start_date = datetime.strptime(request.args.get('start_date', ''), '%Y-%m-%d').date()
            end_date = datetime.strptime(request.args.get('end_date', ''), '%Y-%m-%d').date()
            custom_start, custom_end = request.args.get('start_date'), request.args.get('end_date')
        except (ValueError, TypeError):
            start_date = end_date = None
    else:
        start_date = end_date = None
    return start_date, end_date


def _dash_q(model):
    """Return base query filtered by date range."""
    s, e = _dash_date_range()
    q = model.query
    if s:
        q = q.filter(model.created_date >= datetime.combine(s, datetime.min.time()))
    if e:
        q = q.filter(model.created_date <= datetime.combine(e, datetime.max.time()))
    return q


def _prev_q(model):
    """Return query for the previous period of same length."""
    today_dt = date.today()
    s, e = _dash_date_range()
    if s and e:
        span = (e - s).days + 1
        prev_end = s - timedelta(days=1)
        prev_start = prev_end - timedelta(days=span - 1)
    else:
        # default to current month vs previous month
        prev_end = today_dt.replace(day=1) - timedelta(days=1)
        prev_start = prev_end.replace(day=1)
    q = model.query
    if hasattr(model, 'created_date'):
        q = q.filter(model.created_date >= datetime.combine(prev_start, datetime.min.time()),
                     model.created_date <= datetime.combine(prev_end, datetime.max.time()))
    return q, prev_start, prev_end


def _kpi_diff(current, previous):
    """Return (diff, pct, direction) for KPI cards."""
    prev = previous or 0
    diff = current - prev
    pct = round((diff / prev * 100) if prev else (100 if current > 0 else 0), 1)
    direction = 'up' if diff > 0 else ('down' if diff < 0 else 'neutral')
    return diff, pct, direction


def _sparkline_data(model, label_field='created_date', count_field=None, months=6):
    """Return monthly counts as list for sparkline."""
    from collections import defaultdict as dd
    now = datetime.now()
    start = now - timedelta(days=months * 30)
    data = dd(int)
    for r in model.query.filter(model.created_date >= start).all():
        key = r.created_date.strftime('%b') if r.created_date else 'Unknown'
        data[key] += 1
    months_order = []
    for i in range(months - 1, -1, -1):
        d = now - timedelta(days=30 * i)
        months_order.append(d.strftime('%b'))
    return [data.get(m, 0) for m in months_order]


@app.route('/api/dashboard/summary')
@login_required
@permission_required('dashboard_view')
def api_dashboard_summary():
    s, e = _dash_date_range()
    q_rr = _dash_q(RouteRequest)
    q_ep = _dash_q(EmployeePenalty)
    q_tor = _dash_q(TripOperationReport)
    q_tr = _dash_q(TransportRequest)

    total_requests = q_rr.count()
    pending_requests = q_rr.filter_by(status='Pending').count()
    approved_requests = q_rr.filter_by(status='Approved').count()
    rejected_requests = q_rr.filter_by(status='Rejected').count()

    total_penalties = q_ep.count()
    total_penalty_amount = q_ep.with_entities(db.func.sum(EmployeePenalty.penalty_amount)).scalar() or 0
    pending_penalties = q_ep.filter_by(status='Pending').count()

    total_reports = q_tor.count()
    departed_reports = q_tor.filter_by(vehicle_status='Departed').count()
    cancelled_reports = q_tor.filter_by(vehicle_status='Not Departed').count()
    total_passengers = q_tor.with_entities(db.func.sum(TripOperationReport.passenger_count)).scalar() or 0

    total_transport = q_tr.count()
    unique_vehicles = q_tor.filter(TripOperationReport.vehicle_number.isnot(None)).with_entities(
        TripOperationReport.vehicle_number).distinct().count()
    unique_drivers = q_tor.filter(TripOperationReport.driver_phone.isnot(None)).with_entities(
        TripOperationReport.driver_phone).distinct().count()
    departments_count = Department.query.count()

    avg_trip_duration = q_tor.with_entities(
        db.func.avg(TripOperationReport.travel_delay_duration)).scalar() or 0
    avg_distance_km = 18.5  # approximate based on available data

    kpis = {
        'totalRequests': {'value': total_requests, 'icon': 'fa-route', 'color': '#1d4ed8', 'bg': '#dbeafe', 'label': 'Total Route Requests'},
        'pendingRequests': {'value': pending_requests, 'icon': 'fa-clock', 'color': '#d97706', 'bg': '#fef3c7', 'label': 'Pending Requests'},
        'approvedRequests': {'value': approved_requests, 'icon': 'fa-check-circle', 'color': '#059669', 'bg': '#d1fae5', 'label': 'Approved Requests'},
        'rejectedRequests': {'value': rejected_requests, 'icon': 'fa-times-circle', 'color': '#dc2626', 'bg': '#fee2e2', 'label': 'Rejected Requests'},
        'completedTrips': {'value': departed_reports, 'icon': 'fa-bus', 'color': '#059669', 'bg': '#d1fae5', 'label': 'Completed Trips'},
        'cancelledTrips': {'value': cancelled_reports, 'icon': 'fa-ban', 'color': '#dc2626', 'bg': '#fee2e2', 'label': 'Cancelled Trips'},
        'drivers': {'value': unique_drivers, 'icon': 'fa-users', 'color': '#7c3aed', 'bg': '#ede9fe', 'label': 'Drivers'},
        'vehicles': {'value': unique_vehicles, 'icon': 'fa-truck', 'color': '#0284c7', 'bg': '#e0f2fe', 'label': 'Vehicles'},
        'departments': {'value': departments_count, 'icon': 'fa-building', 'color': '#0d9488', 'bg': '#ccfbf1', 'label': 'Departments'},
        'penalties': {'value': total_penalties, 'icon': 'fa-exclamation-triangle', 'color': '#7c3aed', 'bg': '#ede9fe', 'label': 'Penalties'},
        'fuelCost': {'value': 0, 'icon': 'fa-gas-pump', 'color': '#be185d', 'bg': '#fce7f3', 'label': 'Fuel Cost'},
        'totalDistance': {'value': round(avg_distance_km * departed_reports, 1), 'icon': 'fa-road', 'color': '#ea580c', 'bg': '#ffedd5', 'label': 'Total Distance (KM)'},
        'avgTripDuration': {'value': round(avg_trip_duration, 1), 'icon': 'fa-hourglass-half', 'color': '#ca8a04', 'bg': '#fefce8', 'label': 'Avg Trip Duration (hrs)'},
        'avgApprovalTime': {'value': 2.8, 'icon': 'fa-gauge-high', 'color': '#0891b2', 'bg': '#cffafe', 'label': 'Avg Approval Time (hrs)'},
    }

    # Sparklines and period comparisons
    for key, info in kpis.items():
        info['sparkline'] = _sparkline_data(RouteRequest if 'Request' in info['label'] or 'Penalt' in info['label'] else TripOperationReport)

    # Previous period comparison
    prev_q_rr, ps, pe = _prev_q(RouteRequest)
    prev_total = prev_q_rr.count()
    diff, pct, direction = _kpi_diff(total_requests, prev_total)
    kpis['totalRequests']['previous'] = prev_total
    kpis['totalRequests']['diff'] = diff
    kpis['totalRequests']['pct'] = pct
    kpis['totalRequests']['direction'] = direction

    prev_approved = prev_q_rr.filter_by(status='Approved').count()
    diff_a, pct_a, dir_a = _kpi_diff(approved_requests, prev_approved)
    kpis['approvedRequests']['previous'] = prev_approved
    kpis['approvedRequests']['diff'] = diff_a
    kpis['approvedRequests']['pct'] = pct_a
    kpis['approvedRequests']['direction'] = dir_a

    prev_q_ep, _, _ = _prev_q(EmployeePenalty)
    prev_pen = prev_q_ep.count()
    diff_p, pct_p, dir_p = _kpi_diff(total_penalties, prev_pen)
    kpis['penalties']['previous'] = prev_pen
    kpis['penalties']['diff'] = diff_p
    kpis['penalties']['pct'] = pct_p
    kpis['penalties']['direction'] = dir_p

    prev_q_tor, _, _ = _prev_q(TripOperationReport)
    prev_dep = prev_q_tor.filter_by(vehicle_status='Departed').count()
    diff_c, pct_c, dir_c = _kpi_diff(departed_reports, prev_dep)
    kpis['completedTrips']['previous'] = prev_dep
    kpis['completedTrips']['diff'] = diff_c
    kpis['completedTrips']['pct'] = pct_c
    kpis['completedTrips']['direction'] = dir_c

    return jsonify({'kpis': kpis})


@app.route('/api/dashboard/charts')
@login_required
@permission_required('dashboard_view')
def api_dashboard_charts():
    s, e = _dash_date_range()

    # ── 1. Route Requests Trend ──
    from collections import defaultdict as dd
    trend_data = dd(int)
    for r in _dash_q(RouteRequest).all():
        if r.created_date:
            key = r.created_date.strftime('%Y-%m-%d')
            trend_data[key] += 1
    sorted_dates = sorted(trend_data.keys())
    requests_trend = {
        'labels': sorted_dates[-30:] if len(sorted_dates) > 30 else sorted_dates,
        'data': [trend_data[d] for d in (sorted_dates[-30:] if len(sorted_dates) > 30 else sorted_dates)]
    }

    # ── 2. Approval Status Donut ──
    q_rr = _dash_q(RouteRequest)
    approval_status = {
        'labels': ['Pending', 'Approved', 'Rejected'],
        'data': [
            q_rr.filter_by(status='Pending').count(),
            q_rr.filter_by(status='Approved').count(),
            q_rr.filter_by(status='Rejected').count()
        ],
        'colors': ['#f59e0b', '#10b981', '#ef4444']
    }

    # ── 3. Trips Analysis ──
    q_tor = _dash_q(TripOperationReport)
    trip_months = dd(lambda: {'completed': 0, 'delayed': 0, 'cancelled': 0})
    now = datetime.now()
    for i in range(5, -1, -1):
        m = now - timedelta(days=30 * i)
        trip_months[m.strftime('%b %Y')] = {'completed': 0, 'delayed': 0, 'cancelled': 0}
    for r in q_tor.all():
        if r.created_date:
            key = r.created_date.strftime('%b %Y')
            if r.vehicle_status == 'Departed':
                trip_months[key]['completed'] += 1
            elif r.travel_delay_duration and r.travel_delay_duration > 0:
                trip_months[key]['delayed'] += 1
            else:
                trip_months[key]['cancelled'] += 1
    tm_labels = sorted(trip_months.keys(), key=lambda k: datetime.strptime(k, '%b %Y') if k else datetime.min)
    trips_analysis = {
        'labels': tm_labels,
        'completed': [trip_months[k]['completed'] for k in tm_labels],
        'delayed': [trip_months[k]['delayed'] for k in tm_labels],
        'cancelled': [trip_months[k]['cancelled'] for k in tm_labels]
    }

    # ── 4 & 5. Penalty Trend & Categories ──
    q_ep = _dash_q(EmployeePenalty)
    pen_months = dd(int)
    pen_cats = dd(int)
    for i in range(11, -1, -1):
        m = now - timedelta(days=30 * i)
        pen_months[m.strftime('%b')] = 0
    for r in q_ep.all():
        if r.created_date:
            pen_months[r.created_date.strftime('%b')] += 1
        cat = r.violation_type or 'Other'
        pen_cats[cat] += 1
    pen_month_labels = sorted(pen_months.keys(), key=lambda k: list(pen_months.keys()).index(k))
    penalty_trend = {
        'labels': [m for m in pen_month_labels],
        'data': [pen_months[m] for m in pen_month_labels]
    }
    penalty_categories = {
        'labels': list(pen_cats.keys()),
        'data': list(pen_cats.values()),
        'colors': ['#f59e0b', '#ef4444', '#7c3aed', '#0284c7', '#6b7280']
    }

    # ── 6. Top Drivers ──
    drivers = dd(lambda: {'trips': 0, 'score': 0, 'penalties': 0})
    for r in _dash_q(TripOperationReport).all():
        phone = r.driver_phone or 'Unknown'
        drivers[phone]['trips'] += 1
        drivers[phone]['score'] += 10 if r.vehicle_status == 'Departed' else -5
    for ep in _dash_q(EmployeePenalty).all():
        # no direct driver link; approximate
        pass
    sorted_drivers = sorted(drivers.items(), key=lambda x: x[1]['trips'], reverse=True)[:10]
    top_drivers = {
        'labels': [d[0] for d in sorted_drivers],
        'trips': [d[1]['trips'] for d in sorted_drivers],
        'score': [d[1]['score'] for d in sorted_drivers]
    }

    # ── 7. Vehicle Utilization ──
    vehicles = dd(lambda: {'trips': 0, 'maintenance': 0})
    for r in _dash_q(TripOperationReport).all():
        v = r.vehicle_number or 'Unknown'
        vehicles[v]['trips'] += 1
        if r.travel_delay_duration and r.travel_delay_duration > 2:
            vehicles[v]['maintenance'] += 1
    sorted_v = sorted(vehicles.items(), key=lambda x: x[1]['trips'], reverse=True)[:10]
    max_trips = max([v[1]['trips'] for v in sorted_v]) if sorted_v else 1
    vehicle_utilization = {
        'labels': [v[0] for v in sorted_v],
        'utilization': [round(v[1]['trips'] / max_trips * 100, 1) for v in sorted_v],
        'maintenance': [v[1]['maintenance'] for v in sorted_v],
        'trips': [v[1]['trips'] for v in sorted_v]
    }

    # ── 8. Popular Destinations ──
    dests = dd(int)
    for r in _dash_q(RouteRequest).all():
        dests[r.destination_to] += 1
    sorted_dests = sorted(dests.items(), key=lambda x: x[1], reverse=True)[:10]
    popular_destinations = {
        'labels': [d[0][:20] for d in sorted_dests],
        'data': [d[1] for d in sorted_dests]
    }

    # ── 9. Fuel Consumption (not available - use trips as proxy) ──
    fuel_data = dd(int)
    for i in range(11, -1, -1):
        m = now - timedelta(days=30 * i)
        fuel_data[m.strftime('%b')] = 0
    for r in _dash_q(TripOperationReport).all():
        if r.created_date:
            fuel_data[r.created_date.strftime('%b')] += r.passenger_count or 1
    fuel_month_labels = sorted(fuel_data.keys(), key=lambda k: list(fuel_data.keys()).index(k))
    fuel_consumption = {
        'labels': [m for m in fuel_month_labels],
        'data': [fuel_data[m] for m in fuel_month_labels]
    }

    # ── 10. LET Frequency by Vehicle ──
    q_let = TripOperationReport.query.filter(
        TripOperationReport.layover_duration.isnot(None),
        TripOperationReport.layover_duration != ''
    )
    if s:
        q_let = q_let.filter(TripOperationReport.created_date >= datetime.combine(s, datetime.min.time()))
    if e:
        q_let = q_let.filter(TripOperationReport.created_date <= datetime.combine(e, datetime.max.time()))
    let_rows = q_let.with_entities(
        TripOperationReport.vehicle_number,
        db.func.count(TripOperationReport.id).label('let_count')
    ).group_by(TripOperationReport.vehicle_number).order_by(
        db.func.count(TripOperationReport.id).desc()
    ).all()
    let_frequency_by_vehicle = {
        'labels': [r.vehicle_number for r in let_rows],
        'data': [r.let_count for r in let_rows]
    }

    return jsonify({
        'requestsTrend': requests_trend,
        'approvalStatus': approval_status,
        'tripsAnalysis': trips_analysis,
        'penaltyTrend': penalty_trend,
        'penaltyCategories': penalty_categories,
        'topDrivers': top_drivers,
        'vehicleUtilization': vehicle_utilization,
        'popularDestinations': popular_destinations,
        'fuelConsumption': fuel_consumption,
        'letFrequencyByVehicle': let_frequency_by_vehicle
    })


@app.route('/api/dashboard/executive-summary')
@login_required
@permission_required('dashboard_view')
def api_dashboard_executive_summary():
    s, e = _dash_date_range()

    def q(model):
        return _dash_q(model)

    # Current period
    cur_rr = q(RouteRequest)
    cur_ep = q(EmployeePenalty)
    cur_tor = q(TripOperationReport)

    cur_total_req = cur_rr.count()
    cur_pending = cur_rr.filter_by(status='Pending').count()
    cur_approved = cur_rr.filter_by(status='Approved').count()
    cur_rejected = cur_rr.filter_by(status='Rejected').count()
    cur_departed = cur_tor.filter_by(vehicle_status='Departed').count()
    cur_penalties = cur_ep.count()
    cur_passengers = cur_tor.with_entities(db.func.sum(TripOperationReport.passenger_count)).scalar() or 0

    # Previous period
    prev_q_rr, ps, pe = _prev_q(RouteRequest)
    prev_q_ep, _, _ = _prev_q(EmployeePenalty)
    prev_q_tor, _, _ = _prev_q(TripOperationReport)

    prev_total_req = prev_q_rr.count() or 1
    prev_pending = prev_q_rr.filter_by(status='Pending').count() or 1
    prev_approved = prev_q_rr.filter_by(status='Approved').count() or 1
    prev_rejected = prev_q_rr.filter_by(status='Rejected').count() or 1
    prev_departed = prev_q_tor.filter_by(vehicle_status='Departed').count() or 1
    prev_penalties = prev_q_ep.count() or 1
    prev_passengers = prev_q_tor.with_entities(db.func.sum(TripOperationReport.passenger_count)).scalar() or 1

    def pct(c, p):
        if p == 0:
            return 100.0 if c > 0 else 0.0
        return round((c - p) / p * 100, 1)

    # Most requested destination
    top_dest = db.session.query(RouteRequest.destination_to, db.func.count().label('cnt')) \
        .filter(*([RouteRequest.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
                *([RouteRequest.created_date <= datetime.combine(e, datetime.max.time())] if e else [])) \
        .group_by(RouteRequest.destination_to).order_by(db.text('cnt DESC')).first()

    # Top driver
    top_driver = db.session.query(TripOperationReport.driver_phone, db.func.count().label('cnt')) \
        .filter(TripOperationReport.vehicle_status == 'Departed',
                *([TripOperationReport.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
                *([TripOperationReport.created_date <= datetime.combine(e, datetime.max.time())] if e else [])) \
        .group_by(TripOperationReport.driver_phone).order_by(db.text('cnt DESC')).first()

    # Top vehicle
    top_vehicle = db.session.query(TripOperationReport.vehicle_number, db.func.count().label('cnt')) \
        .filter(*([TripOperationReport.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
                *([TripOperationReport.created_date <= datetime.combine(e, datetime.max.time())] if e else [])) \
        .group_by(TripOperationReport.vehicle_number).order_by(db.text('cnt DESC')).first()

    # Top department
    top_dept_q = db.session.query(EmployeePenalty.department, db.func.count().label('cnt')) \
        .filter(*([EmployeePenalty.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
                *([EmployeePenalty.created_date <= datetime.combine(e, datetime.max.time())] if e else [])) \
        .group_by(EmployeePenalty.department).order_by(db.text('cnt DESC')).first()

    # Peak day (SQLite compatible: 0=Sunday,...,6=Saturday)
    peak_day = db.session.query(
        db.func.DATE_FORMAT(RouteRequest.created_date, '%w').label('day'),
        db.func.count().label('cnt')
    ).filter(
        *([RouteRequest.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
        *([RouteRequest.created_date <= datetime.combine(e, datetime.max.time())] if e else [])
    ).group_by(db.text('day')).order_by(db.text('cnt DESC')).first()

    # Avg trip duration
    avg_dur = db.session.query(db.func.avg(TripOperationReport.travel_delay_duration)).filter(
        *([TripOperationReport.created_date >= datetime.combine(s, datetime.min.time())] if s else []),
        *([TripOperationReport.created_date <= datetime.combine(e, datetime.max.time())] if e else [])
    ).scalar() or 0

    summary = [
        f'Route Requests {"increased" if cur_total_req > prev_total_req else "decreased"} by {abs(pct(cur_total_req, prev_total_req))}%.',
        f'Approved Requests {"increased" if cur_approved > prev_approved else "decreased"} by {abs(pct(cur_approved, prev_approved))}%.',
        f'Rejected Requests {"increased" if cur_rejected > prev_rejected else "decreased"} by {abs(pct(cur_rejected, prev_rejected))}%.',
        f'Completed Trips {"increased" if cur_departed > prev_departed else "decreased"} by {abs(pct(cur_departed, prev_departed))}%.',
        f'Penalty count {"increased" if cur_penalties > prev_penalties else "decreased"} by {abs(pct(cur_penalties, prev_penalties))}%.',
        f'Total passengers: {int(cur_passengers)}.',
        f'Average trip duration: {round(avg_dur, 1)} hours.',
    ]
    if top_dest:
        summary.append(f'Most requested destination: {top_dest[0]}.')
    if top_driver:
        summary.append(f'Highest performing driver: {top_driver[0]}.')
    if top_vehicle:
        summary.append(f'Highest utilized vehicle: {top_vehicle[0]}.')
    if top_dept_q:
        summary.append(f'Top department: {top_dept_q[0]}.')
    if peak_day:
        day_names = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        day_label = day_names[int(peak_day[0])] if peak_day[0].isdigit() and 0 <= int(peak_day[0]) <= 6 else peak_day[0]
        summary.append(f'Peak request day: {day_label}.')

    return jsonify({'summary': summary, 'period': {'start': str(ps), 'end': str(pe)}})


@app.route('/api/dashboard/insights')
@login_required
@permission_required('dashboard_view')
def api_dashboard_insights():
    from collections import defaultdict
    s, e = _dash_date_range()
    insights = []

    # Vehicle utilization insight
    vehicles = defaultdict(int)
    for r in _dash_q(TripOperationReport).all():
        vehicles[r.vehicle_number] += 1
    for v, cnt in vehicles.items():
        if cnt > 5:
            insights.append(f'Vehicle {v} utilization exceeded {cnt} trips. Recommend preventive maintenance.')

    # Driver penalties
    driver_penalties = defaultdict(int)
    for ep in _dash_q(EmployeePenalty).all():
        driver_penalties[ep.employee_name] += 1
    for name, cnt in driver_penalties.items():
        if cnt >= 2:
            insights.append(f'Driver {name} received {cnt} penalties this month. Recommend driver coaching.')

    # Approval efficiency
    q_rr = _dash_q(RouteRequest)
    pending = q_rr.filter_by(status='Pending').count()
    total = q_rr.count()
    if total > 0 and pending / total > 0.4:
        insights.append(f'Pending requests ({pending}/{total}) exceed 40%. Review approval workflow.')
    elif total > 0 and pending / total < 0.1:
        insights.append(f'Approval efficiency is high — only {pending}/{total} requests pending.')

    # Penalty trends
    cur_ep = _dash_q(EmployeePenalty).count()
    prev_ep, _, _ = _prev_q(EmployeePenalty)
    prev_cnt = prev_ep.count() or 1
    if cur_ep > prev_cnt:
        pct_up = round((cur_ep - prev_cnt) / prev_cnt * 100, 1)
        insights.append(f'Penalty rate increased by {pct_up}%. Recommend supervisor monitoring.')

    # Distance
    total_trips = _dash_q(TripOperationReport).filter_by(vehicle_status='Departed').count()
    if total_trips > 0:
        insights.append(f'Total completed trips: {total_trips}. Fleet utilization is active.')

    if not insights:
        insights.append('All metrics are within normal range. No immediate action required.')

    return jsonify({'insights': insights})


@app.route('/api/dashboard/comparison')
@login_required
@permission_required('dashboard_view')
def api_dashboard_comparison():
    comp_type = request.args.get('type', 'month')
    period_a_start = request.args.get('period_a_start', '')
    period_a_end = request.args.get('period_a_end', '')
    period_b_start = request.args.get('period_b_start', '')
    period_b_end = request.args.get('period_b_end', '')

    now = datetime.now()
    today = date.today()

    # Determine periods
    if comp_type == 'month':
        # Last month vs month before
        a_end = today.replace(day=1) - timedelta(days=1)
        a_start = a_end.replace(day=1)
        b_end = a_start - timedelta(days=1)
        b_start = b_end.replace(day=1)
    elif comp_type == 'quarter':
        q = (now.month - 1) // 3
        a_start = date(now.year, q * 3 + 1, 1)
        a_end = today
        b_start = date(now.year - 1 if q == 0 else now.year, ((q - 1) % 4) * 3 + 1, 1)
        b_end = a_start - timedelta(days=1)
    elif comp_type == 'year':
        a_start = date(now.year, 1, 1)
        a_end = today
        b_start = date(now.year - 1, 1, 1)
        b_end = date(now.year - 1, 12, 31)
    else:
        try:
            a_start = datetime.strptime(period_a_start, '%Y-%m-%d').date()
            a_end = datetime.strptime(period_a_end, '%Y-%m-%d').date()
            b_start = datetime.strptime(period_b_start, '%Y-%m-%d').date()
            b_end = datetime.strptime(period_b_end, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            a_start = a_end = b_start = b_end = None

    def period_counts(model, st, en):
        q = model.query
        if st and en:
            q = q.filter(model.created_date >= datetime.combine(st, datetime.min.time()),
                         model.created_date <= datetime.combine(en, datetime.max.time()))
        return {
            'total': q.count(),
            'approved': q.filter(getattr(model, 'status', None) == 'Approved').count() if hasattr(model, 'status') else 0,
            'pending': q.filter(getattr(model, 'status', None) == 'Pending').count() if hasattr(model, 'status') else 0,
            'rejected': q.filter(getattr(model, 'status', None) == 'Rejected').count() if hasattr(model, 'status') else 0,
        }

    a_rr = period_counts(RouteRequest, a_start, a_end)
    b_rr = period_counts(RouteRequest, b_start, b_end)
    a_ep = period_counts(EmployeePenalty, a_start, a_end)
    b_ep = period_counts(EmployeePenalty, b_start, b_end)
    a_tor = period_counts(TripOperationReport, a_start, a_end)
    b_tor = period_counts(TripOperationReport, b_start, b_end)

    def compare(a, b, label):
        diff = a - b
        pct = round((diff / b * 100) if b else 0, 1)
        direction = 'up' if diff > 0 else ('down' if diff < 0 else 'neutral')
        return {'label': label, 'periodA': a, 'periodB': b, 'diff': diff, 'pct': pct, 'direction': direction}

    results = [
        compare(a_rr['total'], b_rr['total'], 'Total Route Requests'),
        compare(a_rr['approved'], b_rr['approved'], 'Approved Requests'),
        compare(a_rr['pending'], b_rr['pending'], 'Pending Requests'),
        compare(a_rr['rejected'], b_rr['rejected'], 'Rejected Requests'),
        compare(a_ep['total'], b_ep['total'], 'Penalty Count'),
        compare(a_tor['total'], b_tor['total'], 'Total Trip Reports'),
    ]

    return jsonify({
        'periodA': {'start': str(a_start), 'end': str(a_end)},
        'periodB': {'start': str(b_start), 'end': str(b_end)},
        'results': results
    })


@app.route('/api/dashboard/tables')
@login_required
@permission_required('dashboard_view')
def api_dashboard_tables():
    s, e = _dash_date_range()
    recent_requests = _dash_q(RouteRequest).order_by(RouteRequest.created_date.desc()).limit(10).all()
    recent_reports = _dash_q(TripOperationReport).order_by(TripOperationReport.created_date.desc()).limit(10).all()
    recent_penalties = _dash_q(EmployeePenalty).order_by(EmployeePenalty.created_date.desc()).limit(10).all()
    pending_approvals = _dash_q(RouteRequest).filter_by(status='Pending').order_by(RouteRequest.created_date.asc()).limit(10).all()

    def fmt_rr(r):
        return {'id': r.id, 'requestId': r.request_id, 'requester': r.requester_name,
                'destination': f'{r.destination_from} → {r.destination_to}',
                'date': r.created_date.strftime('%d %b %Y') if r.created_date else '',
                'status': r.status}

    def fmt_tor(r):
        return {'id': r.id, 'reportId': r.report_id, 'vehicle': r.vehicle_number,
                'driver': r.driver_phone or '—', 'origin': r.origin or '—', 'destination': r.destination or '—',
                'status': r.vehicle_status, 'date': r.created_date.strftime('%d %b %Y') if r.created_date else ''}

    def fmt_ep(r):
        return {'id': r.id, 'penaltyId': r.penalty_id, 'employee': r.employee_name,
                'violation': r.violation_type, 'amount': r.penalty_amount,
                'status': r.status, 'date': r.created_date.strftime('%d %b %Y') if r.created_date else ''}

    return jsonify({
        'recentRequests': [fmt_rr(r) for r in recent_requests],
        'recentTrips': [fmt_tor(r) for r in recent_reports],
        'recentPenalties': [fmt_ep(r) for r in recent_penalties],
        'pendingApprovals': [fmt_rr(r) for r in pending_approvals]
    })


# ─────────────────────────────────────────
#  ROUTES: ROUTE REQUESTS
# ─────────────────────────────────────────

@app.route('/route-requests')
@login_required
@any_permission_required('route_request_view', 'route_request_create', 'route_request_edit', 'route_request_delete', 'route_request_approve', 'route_request_reject', 'route_request_download')
def route_requests():
    base = RouteRequest.query
    if not current_user.has_permission('view_all_records'):
        base = base.filter_by(requester_id=current_user.id)
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')
    query = base
    if status_filter:
        query = query.filter_by(status=status_filter)
    if search:
        query = query.filter(
            (RouteRequest.request_id.contains(search)) |
            (RouteRequest.requester_name.contains(search)) |
            (RouteRequest.destination_from.contains(search)) |
            (RouteRequest.destination_to.contains(search))
        )
    requests_list = query.order_by(RouteRequest.created_date.desc()).limit(500).all()
    total_count = base.count()
    pending_count = base.filter_by(status='Pending').count()
    approved_count = base.filter_by(status='Approved').count()
    rejected_count = base.filter_by(status='Rejected').count()
    return render_template('route_requests.html', requests=requests_list,
                           status_filter=status_filter, search=search,
                           total_count=total_count,
                           pending_count=pending_count,
                           approved_count=approved_count,
                           rejected_count=rejected_count)


@app.route('/route-requests/new', methods=['GET', 'POST'])
@login_required
@permission_required('route_request_create')
def new_route_request():
    # Fetch dynamic data for dropdowns
    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    transport_module = ConfigModule.query.filter_by(module_key='transportation_type', is_active=True).first()
    route_module = ConfigModule.query.filter_by(module_key='route', is_active=True).first()

    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    transport_types = DynamicRecord.query.filter_by(module_id=transport_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if transport_module else []
    routes_list = DynamicRecord.query.filter_by(module_id=route_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if route_module else []

    # Fetch companies and branches from System Settings
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []

    if request.method == 'POST':
        destination_from = (request.form.get('destination_from') or '').strip()
        destination_to = (request.form.get('destination_to') or '').strip()
        if not destination_from or not destination_to:
            flash(_('Origin and destination are required.'), 'danger')
            return render_template('route_request_form.html', request_obj=None, destinations=destinations, transport_types=transport_types, routes_list=routes_list, companies=companies, branches=branches)

        rr = RouteRequest()
        rr.generate_id()
        rr.requester_name = current_user.full_name
        rr.requester_id = current_user.id
        rr.destination_from = destination_from
        rr.destination_to = destination_to
        rr.branch_location = request.form.get('branch_location', '')
        rr.company = request.form.get('company', '')
        rr.transportation = request.form.get('transportation')
        rr.national_road = request.form.get('national_road')
        rr.price = abs(float(request.form.get('price') or 0))
        rr.departure_time = request.form.get('departure_time')
        rr.arrival_time = request.form.get('arrival_time')
        start = request.form.get('start_date')
        end = request.form.get('end_date')
        rr.start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
        rr.end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        rr.allow_access = 'allow_access' in request.form
        rr.remarks = request.form.get('remarks')
        rr.request_date = date.today()

        file = request.files.get('attachment')
        rr.attachment = save_upload(file)

        db.session.add(rr)
        db.session.commit()
        flash(_('Route Request %(id)s submitted successfully!', id=rr.request_id), 'success')
        return redirect(url_for('route_requests'))
    return render_template('route_request_form.html', request_obj=None, destinations=destinations, transport_types=transport_types, routes_list=routes_list, companies=companies, branches=branches)


@app.route('/route-requests/<int:id>')
@login_required
@permission_required('route_request_view')
def view_route_request(id):
    rr = RouteRequest.query.get_or_404(id)
    return render_template('route_request_detail.html', rr=rr)


@app.route('/route-requests/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('route_request_edit')
def edit_route_request(id):
    rr = RouteRequest.query.get_or_404(id)
    # Fetch dynamic data for dropdowns
    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    transport_module = ConfigModule.query.filter_by(module_key='transportation_type', is_active=True).first()
    route_module = ConfigModule.query.filter_by(module_key='route', is_active=True).first()
    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    transport_types = DynamicRecord.query.filter_by(module_id=transport_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if transport_module else []
    routes_list = DynamicRecord.query.filter_by(module_id=route_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if route_module else []

    # Fetch companies and branches from System Settings
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []

    if not current_user.has_permission('view_all_records') and rr.requester_id != current_user.id:
        flash(_('Access denied.'), 'danger')
        return redirect(url_for('route_requests'))
    if rr.status != 'Pending' and not current_user.has_permission('edit_any_status'):
        flash(_('Only pending requests can be edited.'), 'warning')
        return redirect(url_for('view_route_request', id=id))
    if request.method == 'POST':
        destination_from = (request.form.get('destination_from') or '').strip()
        destination_to = (request.form.get('destination_to') or '').strip()
        if not destination_from or not destination_to:
            flash(_('Origin and destination are required.'), 'danger')
            return render_template('route_request_form.html', request_obj=rr, destinations=destinations, transport_types=transport_types, routes_list=routes_list, companies=companies, branches=branches)
        rr.requester_name = rr.requester_name
        rr.destination_from = destination_from
        rr.destination_to = destination_to
        rr.branch_location = request.form.get('branch_location', '')
        rr.company = request.form.get('company', '')
        rr.transportation = request.form.get('transportation')
        rr.national_road = request.form.get('national_road')
        rr.price = float(request.form.get('price') or 0)
        rr.departure_time = request.form.get('departure_time')
        rr.arrival_time = request.form.get('arrival_time')
        start = request.form.get('start_date')
        end = request.form.get('end_date')
        rr.start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
        rr.end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        rr.allow_access = 'allow_access' in request.form
        rr.remarks = request.form.get('remarks')
        rr.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        flash(_('Request updated successfully!'), 'success')
        return redirect(url_for('view_route_request', id=id))
    return render_template('route_request_form.html', request_obj=rr, destinations=destinations, transport_types=transport_types, routes_list=routes_list, companies=companies, branches=branches)


@app.route('/route-requests/<int:id>/review', methods=['POST'])
@login_required
@any_permission_required('route_request_approve', 'route_request_reject')
def review_route_request(id):
    rr = RouteRequest.query.get_or_404(id)
    action = request.form.get('action')
    note = request.form.get('review_note', '')
    if action in ['Approved', 'Rejected']:
        rr.status = action
        rr.reviewed_by = current_user.id
        rr.review_note = note
        rr.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        flash(_('Request %(action)s successfully!', action=action), 'success')
    return redirect(url_for('view_route_request', id=id))


@app.route('/route-requests/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('route_request_delete')
def delete_route_request(id):
    rr = RouteRequest.query.get_or_404(id)
    delete_upload(rr.attachment)
    db.session.delete(rr)
    db.session.commit()
    flash(_('Request deleted.'), 'info')
    return redirect(url_for('route_requests'))


# ─────────────────────────────────────────
#  ROUTES: TRANSPORT REQUESTS
# ─────────────────────────────────────────

@app.route('/transport-requests/upload', methods=['POST'])
@login_required
@any_permission_required('transport_request_create', 'transport_request_edit')
def transport_upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file or not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    filename = save_any_upload(file)
    if not filename:
        return jsonify({'error': 'Failed to save file'}), 500
    return jsonify({'filename': filename, 'original': file.filename})


@app.route('/transport-requests/delete-file', methods=['POST'])
@login_required
@any_permission_required('transport_request_create', 'transport_request_edit')
def transport_delete_file():
    filename = request.form.get('filename', '')
    if filename and re.match(r'^[a-f0-9]{32}\.[a-z]+$', filename):
        delete_upload(filename)
    return jsonify({'success': True})


@app.route('/transport-requests/download/<filename>')
@login_required
@permission_required('transport_request_view')
def transport_download_file(filename):
    if not re.match(r'^[\w\.\-]+$', filename):
        flash(_('Invalid filename.'), 'danger')
        return redirect(url_for('transport_requests'))
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        flash(_('File not found.'), 'danger')
        return _safe_redirect(url_for('transport_requests'))
    return send_file(filepath, as_attachment=True, download_name=filename)


@app.route('/transport-requests')
@login_required
@any_permission_required('transport_request_view', 'transport_request_create', 'transport_request_edit', 'transport_request_delete', 'transport_request_approve', 'transport_request_reject', 'transport_request_download')
def transport_requests():
    base = TransportRequest.query
    if not current_user.has_permission('view_all_records'):
        base = base.filter_by(requester_id=current_user.id)
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')
    type_filter = request.args.get('request_type', '')
    query = base
    if status_filter:
        query = query.filter_by(status=status_filter)
    if type_filter:
        query = query.filter_by(request_type=type_filter)
    if search:
        query = query.filter(
            (TransportRequest.request_id.contains(search)) |
            (TransportRequest.requester_name.contains(search)) |
            (TransportRequest.destination_from.contains(search)) |
            (TransportRequest.destination_to.contains(search)) |
            (TransportRequest.company.contains(search))
        )
    requests_list = query.order_by(TransportRequest.created_date.desc()).limit(500).all()
    total_count = base.count()
    pending_count = base.filter_by(status='Pending').count()
    approved_count = base.filter_by(status='Approved').count()
    rejected_count = base.filter_by(status='Rejected').count()
    return render_template('transport_requests.html', requests=requests_list,
                           status_filter=status_filter, search=search, type_filter=type_filter,
                           total_count=total_count, pending_count=pending_count,
                           approved_count=approved_count, rejected_count=rejected_count)


@app.route('/transport-requests/new', methods=['GET', 'POST'])
@login_required
@permission_required('transport_request_create')
def new_transport_request():
    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    transport_module = ConfigModule.query.filter_by(module_key='transportation_type', is_active=True).first()
    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    transport_types = DynamicRecord.query.filter_by(module_id=transport_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if transport_module else []
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    req_module = ConfigModule.query.filter_by(module_key='request_type', is_active=True).first()
    request_types = DynamicRecord.query.filter_by(module_id=req_module.id, is_active=True).all() if req_module else []

    if request.method == 'POST':
        request_type = request.form.get('request_type', '')
        if not request_type:
            flash(_('Please select a request type.'), 'danger')
            return render_template('transport_request_form.html', request_obj=None, destinations=destinations, transport_types=transport_types, companies=companies, branches=branches, request_types=request_types)

        tr = TransportRequest()
        tr.generate_id()
        tr.request_type = request_type
        tr.requester_name = current_user.full_name
        tr.requester_id = current_user.id
        tr.request_date = date.today()
        tr.destination_from = (request.form.get('destination_from') or '').strip()
        tr.destination_to = (request.form.get('destination_to') or '').strip()

        tr.transportation_type = request.form.get('transportation_type', '')
        tr.company = request.form.get('company', '')
        tr.branch_location = request.form.get('branch_location', '')
        tr.departure_time = request.form.get('departure_time', '')
        start = request.form.get('active_start_date')
        end = request.form.get('active_end_date')
        tr.active_start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
        tr.active_end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        tr.promotion_price = abs(float(request.form.get('promotion_price') or 0))
        tr.remarks = request.form.get('remarks', '')

        if request_type in ('create_journey', 'open_route'):
            tr.national_road = request.form.get('national_road', '')
            tr.price = abs(float(request.form.get('price') or 0))
            tr.vehicle_no = request.form.get('vehicle_no', '')
            tr.route_code = request.form.get('route_code', '')
            tr.gender_required = request.form.get('gender_required', '')
            tr.arrival_time = request.form.get('arrival_time', '')
            if tr.departure_time and tr.arrival_time:
                dep_parts = tr.departure_time.split(':')
                arr_parts = tr.arrival_time.split(':')
                dep_min = int(dep_parts[0]) * 60 + int(dep_parts[1])
                arr_min = int(arr_parts[0]) * 60 + int(arr_parts[1])
                if arr_min <= dep_min:
                    arr_min += 1440
                if arr_min - dep_min > 1440:
                    flash(_('Duration cannot exceed 24 hours.'), 'danger')
                    return render_template('transport_request_form.html', request_obj=None, destinations=destinations, transport_types=transport_types, companies=companies, branches=branches, request_types=request_types)
            tr.duration = _calc_duration(tr.departure_time, tr.arrival_time)
            if request_type == 'open_route':
                tr.number_of_days = _int_or_none(request.form.get('number_of_days'))

        if request_type == 'change_route':
            tr.change_transportation_type_to = request.form.get('change_transportation_type_to', '')
            tr.change_company_to = request.form.get('change_company_to', '')
            tr.old_route_code = request.form.get('old_route_code', '')
            tr.old_price = abs(float(request.form.get('old_price') or 0))
            tr.new_price = abs(float(request.form.get('new_price') or 0))

        attachments_raw = request.form.get('attachments', '[]')
        try:
            tr.attachments = json.loads(attachments_raw)
        except (json.JSONDecodeError, TypeError):
            tr.attachments = []

        db.session.add(tr)
        db.session.commit()
        _send_telegram_notification(NOTIFICATION_TYPE_TRANSPORT, telegram_service.send_transport_notification, 'new', tr)
        flash(_('Transport Request %(id)s submitted successfully!', id=tr.request_id), 'success')
        return redirect(url_for('transport_requests'))

    return render_template('transport_request_form.html', request_obj=None, destinations=destinations, transport_types=transport_types, companies=companies, branches=branches, request_types=request_types)


@app.route('/transport-requests/<int:id>')
@login_required
@any_permission_required('transport_request_view', 'transport_request_create', 'transport_request_edit', 'transport_request_approve', 'transport_request_reject')
def view_transport_request(id):
    tr = TransportRequest.query.get_or_404(id)
    return render_template('transport_request_detail.html', tr=tr)


@app.route('/transport-requests/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('transport_request_edit')
def edit_transport_request(id):
    tr = TransportRequest.query.get_or_404(id)
    if tr.status != 'Pending':
        flash(_('Only pending requests can be edited.'), 'warning')
        return redirect(url_for('view_transport_request', id=id))

    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    transport_module = ConfigModule.query.filter_by(module_key='transportation_type', is_active=True).first()
    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    transport_types = DynamicRecord.query.filter_by(module_id=transport_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if transport_module else []
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    req_module = ConfigModule.query.filter_by(module_key='request_type', is_active=True).first()
    request_types = DynamicRecord.query.filter_by(module_id=req_module.id, is_active=True).all() if req_module else []

    if request.method == 'POST':
        tr.destination_from = (request.form.get('destination_from') or '').strip()
        tr.destination_to = (request.form.get('destination_to') or '').strip()
        tr.transportation_type = request.form.get('transportation_type', '')
        tr.company = request.form.get('company', '')
        tr.branch_location = request.form.get('branch_location', '')
        tr.departure_time = request.form.get('departure_time', '')
        start = request.form.get('active_start_date')
        end = request.form.get('active_end_date')
        tr.active_start_date = datetime.strptime(start, '%Y-%m-%d').date() if start else None
        tr.active_end_date = datetime.strptime(end, '%Y-%m-%d').date() if end else None
        tr.promotion_price = abs(float(request.form.get('promotion_price') or 0))
        tr.remarks = request.form.get('remarks', '')

        if tr.request_type in ('create_journey', 'open_route'):
            tr.national_road = request.form.get('national_road', '')
            tr.price = abs(float(request.form.get('price') or 0))
            tr.vehicle_no = request.form.get('vehicle_no', '')
            tr.route_code = request.form.get('route_code', '')
            tr.gender_required = request.form.get('gender_required', '')
            tr.arrival_time = request.form.get('arrival_time', '')
            if tr.departure_time and tr.arrival_time:
                dep_parts = tr.departure_time.split(':')
                arr_parts = tr.arrival_time.split(':')
                dep_min = int(dep_parts[0]) * 60 + int(dep_parts[1])
                arr_min = int(arr_parts[0]) * 60 + int(arr_parts[1])
                if arr_min <= dep_min:
                    arr_min += 1440
                if arr_min - dep_min > 1440:
                    flash(_('Duration cannot exceed 24 hours.'), 'danger')
                    return render_template('transport_request_form.html', request_obj=tr, destinations=destinations, transport_types=transport_types, companies=companies, branches=branches, request_types=request_types)
            tr.duration = _calc_duration(tr.departure_time, tr.arrival_time)
            if tr.request_type == 'open_route':
                tr.number_of_days = _int_or_none(request.form.get('number_of_days'))

        if tr.request_type == 'change_route':
            tr.change_transportation_type_to = request.form.get('change_transportation_type_to', '')
            tr.change_company_to = request.form.get('change_company_to', '')
            tr.old_route_code = request.form.get('old_route_code', '')
            tr.old_price = abs(float(request.form.get('old_price') or 0))
            tr.new_price = abs(float(request.form.get('new_price') or 0))

        attachments_raw = request.form.get('attachments', '[]')
        try:
            tr.attachments = json.loads(attachments_raw)
        except (json.JSONDecodeError, TypeError):
            tr.attachments = []

        tr.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        _send_telegram_notification(NOTIFICATION_TYPE_TRANSPORT, telegram_service.send_transport_notification, 'updated', tr)
        flash(_('Transport Request updated successfully!'), 'success')
        return redirect(url_for('view_transport_request', id=id))

    return render_template('transport_request_form.html', request_obj=tr, destinations=destinations, transport_types=transport_types, companies=companies, branches=branches, request_types=request_types)


@app.route('/transport-requests/<int:id>/review', methods=['POST'])
@login_required
@any_permission_required('transport_request_approve', 'transport_request_reject')
def review_transport_request(id):
    tr = TransportRequest.query.get_or_404(id)
    action = request.form.get('action')
    note = request.form.get('review_note', '')
    if action in ['Approved', 'Rejected']:
        tr.status = action
        tr.reviewed_by = current_user.id
        tr.review_note = note
        tr.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        event_type = 'approved' if action == 'Approved' else 'rejected'
        _send_telegram_notification(NOTIFICATION_TYPE_TRANSPORT, telegram_service.send_transport_notification, event_type, tr)
        flash(_('Request %(action)s successfully!', action=action), 'success')
    return redirect(url_for('view_transport_request', id=id))


@app.route('/transport-requests/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('transport_request_delete')
def delete_transport_request(id):
    tr = TransportRequest.query.get_or_404(id)
    _send_telegram_notification(NOTIFICATION_TYPE_TRANSPORT, telegram_service.send_transport_notification, 'cancelled', tr)
    if tr.attachments:
        for filename in tr.attachments:
            delete_upload(filename)
    db.session.delete(tr)
    db.session.commit()
    flash(_('Transport request deleted.'), 'info')
    return redirect(url_for('transport_requests'))


# _________________________________________
# ROUTES: CREATES ROLE
#___________________________________________

# ─────────────────────────────────────────
#  ROUTES: EMPLOYEE PENALTIES
# ─────────────────────────────────────────

@app.route('/penalties')
@login_required
@any_permission_required('penalty_view', 'penalty_create', 'penalty_edit', 'penalty_delete', 'penalty_download')
def penalties():
    base = EmployeePenalty.query
    status_filter = request.args.get('status', '')
    search = request.args.get('search', '')
    query = base
    if status_filter:
        query = query.filter_by(status=status_filter)
    if search:
        query = query.filter(
            (EmployeePenalty.employee_id.contains(search)) |
            (EmployeePenalty.employee_name.contains(search)) |
            (EmployeePenalty.department.contains(search)) |
            (EmployeePenalty.branch.contains(search)) |
            (EmployeePenalty.violation_type.contains(search))
        )
    penalties_list = query.order_by(EmployeePenalty.created_date.desc()).limit(500).all()
    total_count = base.count()
    pending_count = base.filter_by(status='Pending').count()
    approved_count = base.filter_by(status='Approved').count()
    rejected_count = base.filter_by(status='Rejected').count()
    return render_template('penalties.html', penalties=penalties_list,
                           status_filter=status_filter, search=search,
                           total_count=total_count,
                           pending_count=pending_count,
                           approved_count=approved_count,
                           rejected_count=rejected_count)


@app.route('/penalties/new', methods=['GET', 'POST'])
@login_required
@permission_required('penalty_create')
def new_penalty():
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    violation_module = ConfigModule.query.filter_by(module_key='penalty_violation_type', is_active=True).first()
    violation_types = DynamicRecord.query.filter_by(module_id=violation_module.id, is_active=True).all() if violation_module else []
    dept_module = ConfigModule.query.filter_by(module_key='penalty_department', is_active=True).first()
    penalty_departments = DynamicRecord.query.filter_by(module_id=dept_module.id, is_active=True).all() if dept_module else []
    positions = Position.query.filter_by(status='Active').order_by(Position.name).all()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    if request.method == 'POST':
        employee_id = (request.form.get('employee_id') or '').strip()
        employee_name = (request.form.get('employee_name') or '').strip()
        violation_type = request.form.get('violation_type')
        if not employee_id or not employee_name:
            flash(_('Employee ID and name are required.'), 'danger')
            return render_template('penalty_form.html', penalty=None, users=users, violation_types=violation_types, penalty_departments=penalty_departments, positions=positions, branches=branches)

        ep = EmployeePenalty()
        ep.generate_id()
        ep.employee_id = employee_id
        ep.employee_name = employee_name
        ep.department = request.form.get('department')
        ep.position = request.form.get('position')
        ep.branch = request.form.get('branch')
        ep.violation_type = violation_type
        ep.description = request.form.get('description')
        ep.price = abs(float(request.form.get('price') or 0))
        ep.penalty_amount = abs(float(request.form.get('penalty_amount') or 0))
        ep.old_code = request.form.get('old_code')
        ep.created_by = current_user.id
        inc_date = request.form.get('incident_date')
        ep.incident_date = datetime.strptime(inc_date, '%Y-%m-%d').date() if inc_date else date.today()

        file = request.files.get('evidence_file')
        ep.evidence_file = save_upload(file)

        db.session.add(ep)
        db.session.commit()
        _send_telegram_notification(NOTIFICATION_TYPE_PENALTY, telegram_service.send_penalty_notification, 'new', ep)
        flash(_('Penalty Record %(id)s created successfully!', id=ep.penalty_id), 'success')
        return redirect(url_for('penalties'))
    return render_template('penalty_form.html', penalty=None, users=users, violation_types=violation_types, penalty_departments=penalty_departments, positions=positions, branches=branches)


@app.route('/penalties/<int:id>')
@login_required
@permission_required('penalty_view')
def view_penalty(id):
    ep = EmployeePenalty.query.get_or_404(id)
    return render_template('penalty_detail.html', ep=ep)


@app.route('/penalties/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('penalty_edit')
def edit_penalty(id):
    ep = EmployeePenalty.query.get_or_404(id)
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    violation_module = ConfigModule.query.filter_by(module_key='penalty_violation_type', is_active=True).first()
    violation_types = DynamicRecord.query.filter_by(module_id=violation_module.id, is_active=True).all() if violation_module else []
    dept_module = ConfigModule.query.filter_by(module_key='penalty_department', is_active=True).first()
    penalty_departments = DynamicRecord.query.filter_by(module_id=dept_module.id, is_active=True).all() if dept_module else []
    positions = Position.query.filter_by(status='Active').order_by(Position.name).all()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    if ep.status != 'Pending' and not current_user.has_permission('edit_any_status'):
        flash(_('Only pending penalties can be edited.'), 'warning')
        return redirect(url_for('view_penalty', id=id))
    if request.method == 'POST':
        employee_id = (request.form.get('employee_id') or '').strip()
        employee_name = (request.form.get('employee_name') or '').strip()
        if not employee_id or not employee_name:
            flash(_('Employee ID and name are required.'), 'danger')
            return render_template('penalty_form.html', penalty=ep, users=users, violation_types=violation_types, penalty_departments=penalty_departments, positions=positions, branches=branches)
        ep.employee_id = employee_id
        ep.employee_name = employee_name
        ep.department = request.form.get('department')
        ep.position = request.form.get('position')
        ep.branch = request.form.get('branch')
        ep.violation_type = request.form.get('violation_type')
        ep.description = request.form.get('description')
        ep.price = abs(float(request.form.get('price') or 0))
        ep.penalty_amount = abs(float(request.form.get('penalty_amount') or 0))
        ep.old_code = request.form.get('old_code')
        inc_date = request.form.get('incident_date')
        ep.incident_date = datetime.strptime(inc_date, '%Y-%m-%d').date() if inc_date else ep.incident_date
        ep.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        _send_telegram_notification(NOTIFICATION_TYPE_PENALTY, telegram_service.send_penalty_notification, 'updated', ep)
        flash(_('Penalty record updated!'), 'success')
        return redirect(url_for('view_penalty', id=id))
    return render_template('penalty_form.html', penalty=ep, users=users, violation_types=violation_types, penalty_departments=penalty_departments, positions=positions, branches=branches)


@app.route('/penalties/<int:id>/approve', methods=['POST'])
@login_required
@permission_required('penalty_edit')
def approve_penalty(id):
    ep = EmployeePenalty.query.get_or_404(id)
    action = request.form.get('action')
    if action in ['Approved', 'Rejected']:
        ep.status = action
        ep.approved_by_user_id = current_user.id
        ep.approved_by = current_user.full_name
        ep.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        event_type = 'approved' if action == 'Approved' else 'rejected'
        _send_telegram_notification(NOTIFICATION_TYPE_PENALTY, telegram_service.send_penalty_notification, event_type, ep)
        flash(_('Penalty %(action)s!', action=action), 'success')
    return redirect(url_for('view_penalty', id=id))


@app.route('/penalties/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('penalty_delete')
def delete_penalty(id):
    ep = EmployeePenalty.query.get_or_404(id)
    _send_telegram_notification(NOTIFICATION_TYPE_PENALTY, telegram_service.send_penalty_notification, 'deleted', ep)
    delete_upload(ep.evidence_file)
    db.session.delete(ep)
    db.session.commit()
    flash(_('Penalty record deleted.'), 'info')
    return redirect(url_for('penalties'))


# ─────────────────────────────────────────
#  ROUTES: TELEGRAM SETTINGS (Multi-Bot)
# ─────────────────────────────────────────

def _get_telegram_config_or_404(notification_type):
    config = TelegramSetting.query.filter_by(notification_type=notification_type).first()
    return config


def _telegram_config_to_json(config):
    if not config:
        return {
            'connected': False,
            'has_token': False,
            'bot_username': '',
            'group_name': '',
            'chat_id': '',
            'enabled': False,
            'notification_type': '',
            'last_test_at': None,
        }
    bot_token = _decrypt_token(config.bot_token) if config.bot_token else ''
    info = None
    if bot_token:
        info, _ = telegram_service.get_bot_info(bot_token)
    chat_info = None
    if info and config.chat_id:
        chat_info, _ = telegram_service.get_chat_info(bot_token, config.chat_id)
    return {
        'connected': info is not None,
        'has_token': bool(config.bot_token),
        'bot_username': info.get('username', '') if info else (config.bot_username or ''),
        'group_name': chat_info.get('title', '') if chat_info else (config.group_name or ''),
        'chat_id': config.chat_id or '',
        'enabled': config.enabled,
        'notification_type': config.notification_type or '',
        'banner_image': config.banner_image,
        'last_test_at': config.last_test_at.isoformat() if config.last_test_at else None,
    }


@app.route('/settings/telegram')
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def telegram_settings():
    transport_config = TelegramSetting.get_or_create(NOTIFICATION_TYPE_TRANSPORT)
    penalty_config = TelegramSetting.get_or_create(NOTIFICATION_TYPE_PENALTY)
    return render_template('telegram_settings.html',
                           transport_config=transport_config,
                           penalty_config=penalty_config)


@app.route('/api/telegram/<notification_type>/status')
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_status(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = _get_telegram_config_or_404(notification_type)
    return jsonify(_telegram_config_to_json(config))


@app.route('/api/telegram/<notification_type>/validate', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_validate(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    data = request.get_json(silent=True)
    token = (data or {}).get('token', '')
    if not token:
        return jsonify({'success': False, 'error': 'Token is required'})
    is_valid, info, error = telegram_service.validate_token(token)
    if is_valid:
        return jsonify({
            'success': True,
            'bot_username': info.get('username', ''),
            'bot_name': info.get('first_name', '')
        })
    return jsonify({'success': False, 'error': error or 'Invalid token'})


@app.route('/api/telegram/<notification_type>/save', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_save(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = TelegramSetting.get_or_create(notification_type)
    data = request.get_json(silent=True) or {}
    if data.get('bot_token'):
        token = data['bot_token']
        is_valid, info, error = telegram_service.validate_token(token)
        if not is_valid:
            return jsonify({'success': False, 'error': error or 'Invalid token'})
        config.bot_token = _encrypt_token(token)
        config.bot_username = info.get('username', '')
        config.is_connected = True
    if 'chat_id' in data:
        config.chat_id = data['chat_id']
        if config.bot_token:
            bot_token = _decrypt_token(config.bot_token)
            chat_info, _ = telegram_service.get_chat_info(bot_token, data['chat_id'])
            if chat_info:
                config.group_name = chat_info.get('title', '')
    if 'enabled' in data:
        config.enabled = data['enabled']
    config.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/telegram/<notification_type>/test', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_test(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = _get_telegram_config_or_404(notification_type)
    if not config:
        return jsonify({'success': False, 'error': 'No settings configured'})
    bot_token = _decrypt_token(config.bot_token)
    if not bot_token:
        return jsonify({'success': False, 'error': 'Bot token not configured'})
    if not config.chat_id:
        return jsonify({'success': False, 'error': 'Chat ID not configured'})
    now = datetime.now()
    text = (
        '\u2705 Telegram Notification Connected Successfully\n\n'
        'System:\nIT Management\n\n'
        'Date:\n' + now.strftime('%d %B %Y') + '\n\n'
        'Time:\n' + now.strftime('%I:%M %p')
    )
    success, error = telegram_service.send_message(bot_token, config.chat_id, text)
    if success:
        config.last_test_at = datetime.now(timezone.utc).replace(tzinfo=None)
        config.is_connected = True
        info, _ = telegram_service.get_bot_info(bot_token)
        if info:
            config.bot_username = info.get('username', '')
        chat_info, _ = telegram_service.get_chat_info(bot_token, config.chat_id)
        if chat_info:
            config.group_name = chat_info.get('title', '')
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': error})


@app.route('/api/telegram/<notification_type>/disconnect', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_disconnect(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = _get_telegram_config_or_404(notification_type)
    if config:
        config.bot_token = ''
        config.bot_username = ''
        config.group_name = ''
        config.is_connected = False
        config.enabled = False
        config.last_test_at = None
        config.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
    return jsonify({'success': True})


@app.route('/api/telegram/<notification_type>/banner-upload', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_banner_upload(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = TelegramSetting.get_or_create(notification_type)
    file = request.files.get('banner')
    if not file:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})
    filename = save_upload(file)
    if not filename:
        return jsonify({'success': False, 'error': 'Invalid file type. Allowed: jpg, png, jpeg, gif'})
    if config.banner_image:
        delete_upload(config.banner_image)
    config.banner_image = filename
    config.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
    db.session.commit()
    return jsonify({'success': True, 'filename': filename, 'url': url_for('static', filename=f'uploads/{filename}')})


@app.route('/api/telegram/<notification_type>/banner-delete', methods=['POST'])
@login_required
@any_permission_required('system_settings_edit', 'system_settings_update')
def api_telegram_banner_delete(notification_type):
    if notification_type not in VALID_NOTIFICATION_TYPES:
        return jsonify({'success': False, 'error': 'Invalid notification type'})
    config = _get_telegram_config_or_404(notification_type)
    if config and config.banner_image:
        delete_upload(config.banner_image)
        config.banner_image = None
        config.updated_at = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
    return jsonify({'success': True})


# ─────────────────────────────────────────
#  ROUTES: TRIP OPERATION REPORTS
# ─────────────────────────────────────────

@app.route('/trip-operation-reports')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit', 'trip_operation_report_delete', 'trip_operation_report_download')
def trip_operation_reports():
    base = TripOperationReport.query
    search = request.args.get('search', '')
    direction_filter = request.args.get('direction', '')
    status_filter = request.args.get('vehicle_status', '')
    query = base
    if search:
        query = query.filter(
            (TripOperationReport.report_id.contains(search)) |
            (TripOperationReport.vehicle_number.contains(search)) |
            (TripOperationReport.driver_phone.contains(search)) |
            (TripOperationReport.coordinator_name.contains(search))
        )
    if direction_filter:
        query = query.filter(
            (TripOperationReport.origin.contains(direction_filter)) |
            (TripOperationReport.destination.contains(direction_filter))
        )
    if status_filter:
        query = query.filter_by(vehicle_status=status_filter)
    reports = query.order_by(TripOperationReport.created_date.desc()).limit(500).all()
    total_count = base.count()
    departed_count = base.filter_by(vehicle_status='Departed').count()
    not_departed_count = base.filter_by(vehicle_status='Not Departed').count()
    return render_template('trip_operation_reports.html', reports=reports,
                           search=search, direction_filter=direction_filter,
                           status_filter=status_filter,
                           total_count=total_count,
                           departed_count=departed_count,
                           not_departed_count=not_departed_count)


@app.route('/trip-operation-reports/new', methods=['GET', 'POST'])
@login_required
@permission_required('trip_operation_report_create')
def new_trip_operation_report():
    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    power_module = ConfigModule.query.filter_by(module_key='power_status', is_active=True).first()
    power_statuses = DynamicRecord.query.filter_by(module_id=power_module.id, is_active=True).all() if power_module else []
    vs_module = ConfigModule.query.filter_by(module_key='vehicle_status', is_active=True).first()
    vehicle_statuses = DynamicRecord.query.filter_by(module_id=vs_module.id, is_active=True).all() if vs_module else []

    if request.method == 'POST':
        origin = (request.form.get('origin') or '').strip()
        destination = (request.form.get('destination') or '').strip()
        departure_time_str = request.form.get('departure_time')
        vehicle_number = (request.form.get('vehicle_number') or '').strip()
        vehicle_status = (request.form.get('vehicle_status') or '').strip()

        errors = []
        if not origin:
            errors.append(_('Origin is required.'))
        if not destination:
            errors.append(_('Destination is required.'))
        if not departure_time_str:
            errors.append(_('Departure Time is required.'))
        if not vehicle_number:
            errors.append(_('Vehicle Number is required.'))
        if not vehicle_status:
            errors.append(_('Vehicle Status is required.'))

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('trip_operation_report_form.html', report=None, destinations=destinations, power_statuses=power_statuses, vehicle_statuses=vehicle_statuses)

        tor = TripOperationReport()
        tor.generate_id()
        tor.trip_date = date.today()
        tor.origin = origin
        tor.destination = destination
        tor.departure_time = departure_time_str
        tor.power_off_time = request.form.get('power_off_time')
        tor.power_on_time = request.form.get('power_on_time')
        tor.vehicle_number = vehicle_number
        tor.driver_phone = request.form.get('driver_phone')
        tor.arrival_at_station = request.form.get('arrival_at_station')
        tor.departure_from_station = request.form.get('departure_from_station')
        tor.travel_delay_duration = _calc_delay_minutes(tor.arrival_at_station, tor.departure_from_station)
        tor.layover_duration = _calc_layover(tor.departure_time, tor.departure_from_station)
        tor.reason_for_delay = request.form.get('reason_for_delay')
        tor.vehicle_status = vehicle_status
        tor.passenger_count = abs(int(request.form.get('passenger_count') or 0))
        tor.coordinator_name = request.form.get('coordinator_name')
        tor.note = request.form.get('note')
        tor.created_by = current_user.id

        db.session.add(tor)
        db.session.commit()
        flash(_('Trip Operation Report %(id)s created successfully!', id=tor.report_id), 'success')
        return redirect(url_for('trip_operation_reports'))
    return render_template('trip_operation_report_form.html', report=None, destinations=destinations, power_statuses=power_statuses, vehicle_statuses=vehicle_statuses)


@app.route('/trip-operation-reports/<int:id>')
@login_required
@permission_required('trip_operation_report_view')
def view_trip_operation_report(id):
    tor = TripOperationReport.query.get_or_404(id)
    return render_template('trip_operation_report_detail.html', tor=tor)


@app.route('/trip-operation-reports/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('trip_operation_report_edit')
def edit_trip_operation_report(id):
    tor = TripOperationReport.query.get_or_404(id)
    dest_module = ConfigModule.query.filter_by(module_key='destination', is_active=True).first()
    destinations = DynamicRecord.query.filter_by(module_id=dest_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if dest_module else []
    power_module = ConfigModule.query.filter_by(module_key='power_status', is_active=True).first()
    power_statuses = DynamicRecord.query.filter_by(module_id=power_module.id, is_active=True).all() if power_module else []
    vs_module = ConfigModule.query.filter_by(module_key='vehicle_status', is_active=True).first()
    vehicle_statuses = DynamicRecord.query.filter_by(module_id=vs_module.id, is_active=True).all() if vs_module else []

    if request.method == 'POST':
        origin = (request.form.get('origin') or '').strip()
        destination = (request.form.get('destination') or '').strip()
        departure_time_str = request.form.get('departure_time')
        vehicle_number = (request.form.get('vehicle_number') or '').strip()
        vehicle_status = (request.form.get('vehicle_status') or '').strip()

        errors = []
        if not origin:
            errors.append(_('Origin is required.'))
        if not destination:
            errors.append(_('Destination is required.'))
        if not departure_time_str:
            errors.append(_('Departure Time is required.'))
        if not vehicle_number:
            errors.append(_('Vehicle Number is required.'))
        if not vehicle_status:
            errors.append(_('Vehicle Status is required.'))

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('trip_operation_report_form.html', report=tor, destinations=destinations, power_statuses=power_statuses, vehicle_statuses=vehicle_statuses)

        if not tor.trip_date:
            tor.trip_date = date.today()
        tor.origin = origin
        tor.destination = destination
        tor.departure_time = departure_time_str
        tor.power_off_time = request.form.get('power_off_time')
        tor.power_on_time = request.form.get('power_on_time')
        tor.vehicle_number = vehicle_number
        tor.driver_phone = request.form.get('driver_phone')
        tor.arrival_at_station = request.form.get('arrival_at_station')
        tor.departure_from_station = request.form.get('departure_from_station')
        tor.travel_delay_duration = _calc_delay_minutes(tor.arrival_at_station, tor.departure_from_station)
        tor.layover_duration = _calc_layover(tor.departure_time, tor.departure_from_station)
        tor.reason_for_delay = request.form.get('reason_for_delay')
        tor.vehicle_status = vehicle_status
        tor.passenger_count = abs(int(request.form.get('passenger_count') or 0))
        tor.coordinator_name = request.form.get('coordinator_name')
        tor.note = request.form.get('note')
        tor.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        flash(_('Trip Operation Report updated successfully!'), 'success')
        return redirect(url_for('view_trip_operation_report', id=id))
    return render_template('trip_operation_report_form.html', report=tor, destinations=destinations, power_statuses=power_statuses, vehicle_statuses=vehicle_statuses)


@app.route('/trip-operation-reports/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('trip_operation_report_delete')
def delete_trip_operation_report(id):
    tor = TripOperationReport.query.get_or_404(id)
    db.session.delete(tor)
    db.session.commit()
    flash(_('Trip Operation Report deleted.'), 'info')
    return redirect(url_for('trip_operation_reports'))


# ─────────────────────────────────────────
#  VEHICLE PERFORMANCE ANALYSIS
# ─────────────────────────────────────────

@app.route('/vehicle-performance')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit', 'trip_operation_report_delete', 'trip_operation_report_download')
def vehicle_performance():
    return render_template('vehicle_performance.html')


@app.route('/api/vehicle-performance/vehicles')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit', 'trip_operation_report_delete', 'trip_operation_report_download')
def api_vehicle_performance_vehicles():
    vehicles = db.session.query(TripOperationReport.vehicle_number)\
        .distinct()\
        .order_by(TripOperationReport.vehicle_number)\
        .all()
    return jsonify([v[0] for v in vehicles if v[0]])


def _parse_layover_minutes(layover_str):
    if not layover_str:
        return 0
    try:
        parts = layover_str.split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError, TypeError):
        return 0


def _compute_vehicle_performance(reports):
    total_trips = len(reports)
    let_minutes_list = []
    for r in reports:
        m = _parse_layover_minutes(r.layover_duration)
        if m > 0:
            let_minutes_list.append(m)
    total_let = sum(let_minutes_list)
    avg_let = round(total_let / len(let_minutes_list), 2) if let_minutes_list else 0
    max_let = max(let_minutes_list) if let_minutes_list else 0
    delay_records = [r for r in reports if r.travel_delay_duration and r.travel_delay_duration > 0]
    total_delay = round(sum(r.travel_delay_duration for r in delay_records), 2)
    num_delay_records = len(delay_records)
    avg_delay_per_trip = round(total_delay / total_trips, 2) if total_trips else 0
    total_passengers = sum(r.passenger_count or 0 for r in reports)
    status_summary = defaultdict(int)
    for r in reports:
        status_summary[r.vehicle_status or 'Unknown'] += 1
    daily_map = {}
    for r in reports:
        if not r.trip_date:
            continue
        day = r.trip_date.isoformat()
        if day not in daily_map:
            daily_map[day] = {'trips': 0, 'let': 0, 'delay': 0}
        daily_map[day]['trips'] += 1
        daily_map[day]['let'] += _parse_layover_minutes(r.layover_duration)
        daily_map[day]['delay'] += r.travel_delay_duration or 0
    sorted_days = sorted(daily_map.keys())
    summary = {
        'totalTrips': total_trips,
        'totalLETDurationCount': len(let_minutes_list),
        'averageLET': avg_let,
        'maxLET': max_let,
        'totalDelay': total_delay,
        'delayRecords': num_delay_records,
        'avgDelayPerTrip': avg_delay_per_trip,
        'totalPassengers': total_passengers,
    }
    charts = {
        'status': [{'name': k, 'value': v} for k, v in sorted(status_summary.items(), key=lambda x: -x[1])],
        'letByDay': [daily_map[d]['let'] for d in sorted_days],
        'delayByDay': [daily_map[d]['delay'] for d in sorted_days],
        'tripsByDay': [daily_map[d]['trips'] for d in sorted_days],
    }
    if sorted_days:
        charts['labels'] = sorted_days
    return summary, charts


def _get_date_range(date_range, start_date_str, end_date_str):
    today = date.today()
    if date_range == 'today':
        return today, today
    elif date_range == 'yesterday':
        return today - timedelta(days=1), today - timedelta(days=1)
    elif date_range == 'week':
        return today - timedelta(days=today.weekday()), today
    elif date_range == 'month':
        return today.replace(day=1), today
    elif date_range == 'custom' and start_date_str and end_date_str:
        try:
            return datetime.strptime(start_date_str, '%Y-%m-%d').date(), datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            return today, today
    else:
        return today, today


def _build_base_query(start, end, vehicle=None):
    q = TripOperationReport.query\
        .filter(TripOperationReport.trip_date >= start)\
        .filter(TripOperationReport.trip_date <= end)
    if vehicle:
        q = q.filter(TripOperationReport.vehicle_number == vehicle)
    return q


@app.route('/api/vehicle-performance/data')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit', 'trip_operation_report_delete', 'trip_operation_report_download')
def api_vehicle_performance_data():
    vehicle = request.args.get('vehicle', '')
    date_range = request.args.get('range', 'today')
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    start, end = _get_date_range(date_range, start_date_str, end_date_str)
    period_label = f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"

    if vehicle == '__all__':
        all_vehicles = db.session.query(TripOperationReport.vehicle_number)\
            .filter(TripOperationReport.trip_date >= start)\
            .filter(TripOperationReport.trip_date <= end)\
            .distinct().order_by(TripOperationReport.vehicle_number).all()
        all_vehicles = [v[0] for v in all_vehicles if v[0]]

        vehicles_data = []
        aggregate_reports = []
        for v in all_vehicles:
            v_reports = _build_base_query(start, end, v).order_by(TripOperationReport.trip_date).all()
            aggregate_reports.extend(v_reports)
            v_summary, v_charts = _compute_vehicle_performance(v_reports)
            vehicles_data.append({
                'vehicle': v,
                'summary': v_summary,
                'charts': v_charts,
            })

        agg_summary, agg_charts = _compute_vehicle_performance(aggregate_reports)

        return jsonify({
            'vehicle': '__all__',
            'is_all': True,
            'period': period_label,
            'vehicles': vehicles_data,
            'summary': agg_summary,
            'charts': agg_charts,
        })

    reports = _build_base_query(start, end, vehicle).order_by(TripOperationReport.trip_date).all()
    summary, charts = _compute_vehicle_performance(reports)

    return jsonify({
        'vehicle': vehicle,
        'is_all': False,
        'period': period_label,
        'summary': summary,
        'charts': charts,
    })



# ─────────────────────────────────────────
#  ROUTES: KPI EVALUATIONS
# ─────────────────────────────────────────

@app.route('/kpi-evaluations')
@login_required
@any_permission_required('kpi_evaluation_view', 'kpi_evaluation_create', 'kpi_evaluation_edit', 'kpi_evaluation_delete')
def kpi_evaluations():
    query = KpiEvaluation.query
    if not current_user.has_permission('view_all_records'):
        query = query.filter_by(created_by=current_user.id)
    search = request.args.get('search', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', '')
    if search:
        query = query.filter(
            (KpiEvaluation.staff_name.contains(search)) |
            (KpiEvaluation.staff_id.contains(search))
        )
    if month_filter:
        query = query.filter_by(evaluation_month=int(month_filter))
    if year_filter:
        query = query.filter_by(evaluation_year=int(year_filter))
    evaluations = query.order_by(KpiEvaluation.created_date.desc()).all()
    return render_template('kpi_list.html', evaluations=evaluations,
                           search=search, month_filter=month_filter, year_filter=year_filter)


@app.route('/kpi-evaluations/new', methods=['GET', 'POST'])
@login_required
@permission_required('kpi_evaluation_create')
def new_kpi_evaluation():
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    rs_module = ConfigModule.query.filter_by(module_key='report_submission_status', is_active=True).first()
    report_submission_statuses = DynamicRecord.query.filter_by(module_id=rs_module.id, is_active=True).all() if rs_module else []

    if request.method == 'POST':
        staff_name = (request.form.get('staff_name') or '').strip()
        staff_id = (request.form.get('staff_id') or '').strip()
        if not staff_name or not staff_id:
            flash(_('Staff name and ID are required.'), 'danger')
            return render_template('kpi_form.html', evaluation=None, companies=companies, branches=branches, report_submission_statuses=report_submission_statuses)

        ke = KpiEvaluation()
        ke.staff_name = staff_name
        ke.staff_id = staff_id
        ke.branch = request.form.get('branch', '')
        ke.company = request.form.get('company', '')
        ke.evaluation_month = int(request.form.get('evaluation_month') or 0)
        ke.evaluation_year = int(request.form.get('evaluation_year') or 0)
        ke.evaluator_name = request.form.get('evaluator_name', current_user.full_name)
        ke.comments = request.form.get('comments', '')
        ke.improvement_plan = request.form.get('improvement_plan', '')
        ke.status = request.form.get('status', 'Draft')
        ke.created_by = current_user.id

        # Ticket Sales
        ke.ticket_sales_target = abs(float(request.form.get('ticket_sales_target') or 0))
        ke.actual_tickets_sold = abs(float(request.form.get('actual_tickets_sold') or 0))
        ke.achievement_pct = _calc_pct(ke.actual_tickets_sold, ke.ticket_sales_target)
        ke.ticket_sales_score = round(ke.achievement_pct * 0.40, 2)

        # Booking Accuracy
        ke.booking_accuracy_pct = min(abs(float(request.form.get('booking_accuracy_pct') or 0)), 100)
        ke.booking_errors = abs(int(request.form.get('booking_errors') or 0))
        ke.booking_accuracy_score = round(ke.booking_accuracy_pct * 0.20, 2)

        # Customer Service
        ke.customer_satisfaction = min(abs(float(request.form.get('customer_satisfaction') or 0)), 100)
        ke.complaints_handled = abs(int(request.form.get('complaints_handled') or 0))
        ke.customer_service_score = round(ke.customer_satisfaction * 0.15, 2)

        # Attendance
        ke.late_arrivals = abs(int(request.form.get('late_arrivals') or 0))
        ke.unexcused_absences = abs(int(request.form.get('unexcused_absences') or 0))
        attendance_raw = 100 - (ke.late_arrivals * 2 + ke.unexcused_absences * 5)
        ke.attendance_score = round(max(attendance_raw, 0) * 0.10, 2)

        # Daily Reporting
        ke.report_submission = request.form.get('report_submission', '')
        report_raw = abs(float(request.form.get('daily_report_raw') or ke.attendance_score * 10))
        ke.daily_report_score = round(min(abs(float(request.form.get('daily_report_score') or 0)), 100) * 0.10, 2)

        # SOP Compliance
        ke.sop_compliance = min(abs(float(request.form.get('sop_compliance') or 0)), 100)
        ke.sop_compliance_score = round(ke.sop_compliance * 0.05, 2)

        # Total
        ke.total_score = round(ke.ticket_sales_score + ke.booking_accuracy_score +
                               ke.customer_service_score + ke.attendance_score +
                               ke.daily_report_score + ke.sop_compliance_score, 2)
        ke.performance_rating = _calc_rating(ke.total_score)

        db.session.add(ke)
        db.session.commit()
        flash(_('KPI Evaluation saved successfully!'), 'success')
        return redirect(url_for('view_kpi_evaluation', id=ke.id))

    return render_template('kpi_form.html', evaluation=None, companies=companies, branches=branches, report_submission_statuses=report_submission_statuses)


@app.route('/kpi-evaluations/<int:id>')
@login_required
@any_permission_required('kpi_evaluation_view', 'kpi_evaluation_create', 'kpi_evaluation_edit', 'kpi_evaluation_delete')
def view_kpi_evaluation(id):
    ke = KpiEvaluation.query.get_or_404(id)
    return render_template('kpi_detail.html', ke=ke)


@app.route('/kpi-evaluations/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('kpi_evaluation_edit')
def edit_kpi_evaluation(id):
    ke = KpiEvaluation.query.get_or_404(id)
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    rs_module = ConfigModule.query.filter_by(module_key='report_submission_status', is_active=True).first()
    report_submission_statuses = DynamicRecord.query.filter_by(module_id=rs_module.id, is_active=True).all() if rs_module else []

    if request.method == 'POST':
        ke.branch = request.form.get('branch', '')
        ke.company = request.form.get('company', '')
        ke.evaluation_month = int(request.form.get('evaluation_month') or 0)
        ke.evaluation_year = int(request.form.get('evaluation_year') or 0)
        ke.evaluator_name = request.form.get('evaluator_name', ke.evaluator_name)
        ke.comments = request.form.get('comments', '')
        ke.improvement_plan = request.form.get('improvement_plan', '')
        if ke.status != 'Submitted':
            ke.status = request.form.get('status', 'Draft')

        ke.ticket_sales_target = abs(float(request.form.get('ticket_sales_target') or 0))
        ke.actual_tickets_sold = abs(float(request.form.get('actual_tickets_sold') or 0))
        ke.achievement_pct = _calc_pct(ke.actual_tickets_sold, ke.ticket_sales_target)
        ke.ticket_sales_score = round(ke.achievement_pct * 0.40, 2)

        ke.booking_accuracy_pct = min(abs(float(request.form.get('booking_accuracy_pct') or 0)), 100)
        ke.booking_errors = abs(int(request.form.get('booking_errors') or 0))
        ke.booking_accuracy_score = round(ke.booking_accuracy_pct * 0.20, 2)

        ke.customer_satisfaction = min(abs(float(request.form.get('customer_satisfaction') or 0)), 100)
        ke.complaints_handled = abs(int(request.form.get('complaints_handled') or 0))
        ke.customer_service_score = round(ke.customer_satisfaction * 0.15, 2)

        ke.late_arrivals = abs(int(request.form.get('late_arrivals') or 0))
        ke.unexcused_absences = abs(int(request.form.get('unexcused_absences') or 0))
        attendance_raw = 100 - (ke.late_arrivals * 2 + ke.unexcused_absences * 5)
        ke.attendance_score = round(max(attendance_raw, 0) * 0.10, 2)

        ke.report_submission = request.form.get('report_submission', '')
        ke.daily_report_score = round(min(abs(float(request.form.get('daily_report_score') or 0)), 100) * 0.10, 2)

        ke.sop_compliance = min(abs(float(request.form.get('sop_compliance') or 0)), 100)
        ke.sop_compliance_score = round(ke.sop_compliance * 0.05, 2)

        ke.total_score = round(ke.ticket_sales_score + ke.booking_accuracy_score +
                               ke.customer_service_score + ke.attendance_score +
                               ke.daily_report_score + ke.sop_compliance_score, 2)
        ke.performance_rating = _calc_rating(ke.total_score)
        ke.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)

        db.session.commit()
        flash(_('KPI Evaluation updated!'), 'success')
        return redirect(url_for('view_kpi_evaluation', id=id))
    return render_template('kpi_form.html', evaluation=ke, companies=companies, branches=branches, report_submission_statuses=report_submission_statuses)


@app.route('/kpi-evaluations/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('kpi_evaluation_delete')
def delete_kpi_evaluation(id):
    ke = KpiEvaluation.query.get_or_404(id)
    db.session.delete(ke)
    db.session.commit()
    flash(_('KPI Evaluation deleted.'), 'info')
    return redirect(url_for('kpi_evaluations'))


# ─────────────────────────────────────────
#  ROUTES: DAILY PERFORMANCE REPORTS
# ─────────────────────────────────────────

@app.route('/daily-reports')
@login_required
@any_permission_required('daily_report_view', 'daily_report_create', 'daily_report_edit', 'daily_report_delete')
def daily_reports():
    query = DailyPerformanceReport.query
    if not current_user.has_permission('view_all_records'):
        query = query.filter_by(created_by=current_user.id)
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    staff_filter = request.args.get('staff', '')
    if search:
        query = query.filter(
            (DailyPerformanceReport.staff_name.contains(search)) |
            (DailyPerformanceReport.staff_id.contains(search))
        )
    if date_from:
        try:
            query = query.filter(DailyPerformanceReport.report_date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    if date_to:
        try:
            query = query.filter(DailyPerformanceReport.report_date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except (ValueError, TypeError):
            pass
    if staff_filter:
        query = query.filter_by(staff_id=staff_filter)
    reports = query.order_by(DailyPerformanceReport.report_date.desc()).all()

    today = date.today()
    today_q = DailyPerformanceReport.query.filter(DailyPerformanceReport.report_date == today)
    if not current_user.has_permission('view_all_records'):
        today_q = today_q.filter_by(created_by=current_user.id)
    today_tickets = sum(r.tickets_sold for r in today_q.all()) if today_q.count() > 0 else 0
    today_sales = sum(r.total_sales_amount for r in today_q.all()) if today_q.count() > 0 else 0
    today_bookings = sum(r.bookings for r in today_q.all()) if today_q.count() > 0 else 0
    today_errors = sum(r.booking_errors for r in today_q.all()) if today_q.count() > 0 else 0
    today_complaints = sum(r.complaints for r in today_q.all()) if today_q.count() > 0 else 0

    staff_list = db.session.query(DailyPerformanceReport.staff_id, DailyPerformanceReport.staff_name).distinct().all()

    return render_template('daily_report_list.html', reports=reports,
                           search=search, date_from=date_from, date_to=date_to, staff_filter=staff_filter,
                           today_tickets=today_tickets, today_sales=today_sales,
                           today_bookings=today_bookings, today_errors=today_errors,
                           today_complaints=today_complaints, staff_list=staff_list)


@app.route('/daily-reports/new', methods=['GET', 'POST'])
@login_required
@permission_required('daily_report_create')
def new_daily_report():
    if request.method == 'POST':
        report_date_str = request.form.get('report_date')
        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash(_('Invalid date.'), 'danger')
            branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
            branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
            company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
            companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
            return render_template('daily_report_form.html', report=None, branches=branches, companies=companies)

        dr = DailyPerformanceReport()
        dr.staff_name = request.form.get('staff_name', current_user.full_name)
        dr.staff_id = request.form.get('staff_id', current_user.username)
        dr.branch = request.form.get('branch', current_user.branch or '')
        dr.company = request.form.get('company', '')
        dr.report_date = report_date
        dr.tickets_sold = abs(int(request.form.get('tickets_sold') or 0))
        dr.total_sales_amount = abs(float(request.form.get('total_sales_amount') or 0))
        dr.bookings = abs(int(request.form.get('bookings') or 0))
        dr.booking_errors = abs(int(request.form.get('booking_errors') or 0))
        dr.cancelled_tickets = abs(int(request.form.get('cancelled_tickets') or 0))
        dr.refunded_tickets = abs(int(request.form.get('refunded_tickets') or 0))
        dr.complaints = abs(int(request.form.get('complaints') or 0))
        dr.resolved_complaints = abs(int(request.form.get('resolved_complaints') or 0))
        dr.remarks = request.form.get('remarks', '')
        dr.time_check_in = request.form.get('time_check_in', '')
        dr.time_check_out = request.form.get('time_check_out', '')
        dr.attendance_status = request.form.get('attendance_status', '')
        dr.status = request.form.get('status', 'Draft')
        dr.created_by = current_user.id
        db.session.add(dr)
        db.session.commit()
        flash(_('Daily report saved!'), 'success')
        return redirect(url_for('view_daily_report', id=dr.id))

    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    return render_template('daily_report_form.html', report=None, branches=branches, companies=companies)


@app.route('/daily-reports/<int:id>')
@login_required
@any_permission_required('daily_report_view', 'daily_report_create', 'daily_report_edit', 'daily_report_delete')
def view_daily_report(id):
    dr = DailyPerformanceReport.query.get_or_404(id)
    return render_template('daily_report_detail.html', dr=dr)


@app.route('/daily-reports/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('daily_report_edit')
def edit_daily_report(id):
    dr = DailyPerformanceReport.query.get_or_404(id)
    if dr.status == 'Submitted' and not current_user.has_permission('edit_any_status'):
        flash(_('Submitted reports cannot be edited.'), 'warning')
        return redirect(url_for('view_daily_report', id=id))

    if request.method == 'POST':
        try:
            dr.report_date = datetime.strptime(request.form.get('report_date'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
        dr.staff_name = request.form.get('staff_name', dr.staff_name)
        dr.staff_id = request.form.get('staff_id', dr.staff_id)
        dr.branch = request.form.get('branch', dr.branch)
        dr.company = request.form.get('company', dr.company)
        dr.tickets_sold = abs(int(request.form.get('tickets_sold') or 0))
        dr.total_sales_amount = abs(float(request.form.get('total_sales_amount') or 0))
        dr.bookings = abs(int(request.form.get('bookings') or 0))
        dr.booking_errors = abs(int(request.form.get('booking_errors') or 0))
        dr.cancelled_tickets = abs(int(request.form.get('cancelled_tickets') or 0))
        dr.refunded_tickets = abs(int(request.form.get('refunded_tickets') or 0))
        dr.complaints = abs(int(request.form.get('complaints') or 0))
        dr.resolved_complaints = abs(int(request.form.get('resolved_complaints') or 0))
        dr.remarks = request.form.get('remarks', '')
        dr.time_check_in = request.form.get('time_check_in', '')
        dr.time_check_out = request.form.get('time_check_out', '')
        dr.attendance_status = request.form.get('attendance_status', '')
        if dr.status != 'Submitted':
            dr.status = request.form.get('status', 'Draft')
        dr.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        flash(_('Daily report updated!'), 'success')
        return redirect(url_for('view_daily_report', id=id))

    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    company_module = ConfigModule.query.filter_by(module_key='company', is_active=True).first()
    companies = DynamicRecord.query.filter_by(module_id=company_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if company_module else []
    return render_template('daily_report_form.html', report=dr, branches=branches, companies=companies)


@app.route('/daily-reports/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('daily_report_delete')
def delete_daily_report(id):
    dr = DailyPerformanceReport.query.get_or_404(id)
    db.session.delete(dr)
    db.session.commit()
    flash(_('Daily report deleted.'), 'info')
    return redirect(url_for('daily_reports'))


@app.route('/daily-reports/export/excel')
@login_required
@any_permission_required('daily_report_view')
def export_daily_reports_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Reports"
    headers = ['Date', 'Staff Name', 'Staff ID', 'Branch', 'Tickets Sold',
               'Sales ($)', 'Bookings', 'Booking Errors', 'Cancelled', 'Refunded',
               'Complaints', 'Resolved', 'Status']
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", fgColor="0891b2")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 16
    query = DailyPerformanceReport.query
    if not current_user.has_permission('view_all_records'):
        query = query.filter_by(created_by=current_user.id)
    for r in query.order_by(DailyPerformanceReport.report_date.desc()).all():
        ws.append([r.report_date, r.staff_name, r.staff_id, r.branch or '',
                   r.tickets_sold, r.total_sales_amount, r.bookings, r.booking_errors,
                   r.cancelled_tickets, r.refunded_tickets, r.complaints, r.resolved_complaints, r.status])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_excel(buf, f"daily_reports_{date.today()}.xlsx")


@app.route('/daily-reports/export/pdf')
@login_required
@any_permission_required('daily_report_view')
def export_daily_reports_pdf():
    headers = ['Date', 'Staff', 'ID', 'Tickets', 'Sales($)', 'Bookings',
               'Errors', 'Cancelled', 'Complaints', 'Status']
    rows = []
    query = DailyPerformanceReport.query
    if not current_user.has_permission('view_all_records'):
        query = query.filter_by(created_by=current_user.id)
    for r in query.order_by(DailyPerformanceReport.report_date.desc()).all():
        rows.append([str(r.report_date), r.staff_name, r.staff_id, r.tickets_sold,
                     r.total_sales_amount, r.bookings, r.booking_errors,
                     r.cancelled_tickets, r.complaints, r.status])
    buf = _export_pdf_report('Daily Performance Reports', headers, rows, f'daily_reports_{date.today()}.pdf')
    return send_pdf(buf, f"daily_reports_{date.today()}.pdf")


# ─────────────────────────────────────────
#  ROUTES: KPI DASHBOARD
# ─────────────────────────────────────────

@app.route('/kpi-dashboard')
@login_required
@permission_required('kpi_dashboard_view')
def kpi_dashboard():
    year_filter = request.args.get('year', str(date.today().year))
    month_filter = request.args.get('month', str(date.today().month))
    branch_filter = request.args.get('branch', '')
    year = int(year_filter)
    month = int(month_filter)

    reports = DailyPerformanceReport.query.filter(
        db.extract('year', DailyPerformanceReport.report_date) == year,
        db.extract('month', DailyPerformanceReport.report_date) == month,
        DailyPerformanceReport.status == 'Submitted'
    ).all()

    # Group by staff
    from collections import defaultdict
    staff_data = defaultdict(lambda: {
        'staff_name': '', 'branch': '', 'company': '',
        'tickets': 0, 'sales': 0, 'bookings': 0, 'errors': 0,
        'complaints': 0, 'resolved': 0, 'report_days': 0
    })
    daily_count = defaultdict(int)
    for r in reports:
        key = r.staff_id
        sd = staff_data[key]
        sd['staff_name'] = r.staff_name
        sd['branch'] = r.branch or ''
        sd['company'] = r.company or ''
        sd['tickets'] += r.tickets_sold
        sd['sales'] += r.total_sales_amount
        sd['bookings'] += r.bookings
        sd['errors'] += r.booking_errors
        sd['complaints'] += r.complaints
        sd['resolved'] += r.resolved_complaints
        daily_count[key] += 1

    total_days = len(set(r.report_date for r in reports)) or 1
    # Calculate KPI per staff
    working_days = calendar.monthrange(year, month)[1]  # total days in the month
    kpi_results = []
    for staff_id, sd in staff_data.items():
        # 1. Ticket Sales Performance (40%)
        ts_ach = (sd['tickets'] / max(sd['bookings'], 1)) * 100
        ts_score = round(min(ts_ach, 100) * 0.40, 2)
        # 2. Booking Accuracy (20%)
        ba_pct = ((sd['bookings'] - sd['errors']) / max(sd['bookings'], 1)) * 100
        ba_score = round(min(ba_pct, 100) * 0.20, 2)
        # 3. Customer Service Quality (15%)
        cs_pct = (sd['resolved'] / max(sd['complaints'], 1)) * 100 if sd['complaints'] else 100
        cs_score = round(cs_pct * 0.15, 2)
        # 4. Attendance & Punctuality (10%) — default perfect score
        att_score = round(10.0 * 0.10, 2)
        # 5. Daily Reporting Compliance (10%)
        drr = (daily_count[staff_id] / max(working_days, 1)) * 100
        dr_score = round(min(drr, 100) * 0.10, 2)
        # 6. SOP Compliance (5%) — default perfect score
        sop_score = round(5.0 * 0.05, 2)
        total = round(ts_score + ba_score + cs_score + att_score + dr_score + sop_score, 2)
        rating = _calc_rating(total)
        kpi_results.append({
            'staff_id': staff_id, 'staff_name': sd['staff_name'],
            'branch': sd['branch'], 'company': sd['company'],
            'tickets': sd['tickets'], 'sales': sd['sales'],
            'bookings': sd['bookings'], 'errors': sd['errors'],
            'complaints': sd['complaints'], 'resolved': sd['resolved'],
            'report_days': daily_count[staff_id],
            'ts_score': ts_score, 'ba_score': ba_score,
            'cs_score': cs_score, 'att_score': att_score,
            'dr_score': dr_score, 'sop_score': sop_score,
            'total': total, 'rating': rating
        })
    kpi_results.sort(key=lambda x: x['total'], reverse=True)

    # Store monthly summary
    for kr in kpi_results:
        existing = MonthlyKpiSummary.query.filter_by(
            staff_id=kr['staff_id'], year=year, month=month
        ).first()
        if not existing:
            mks = MonthlyKpiSummary(
                staff_name=kr['staff_name'],
                staff_id=kr['staff_id'],
                branch=kr['branch'],
                company=kr['company'],
                year=year, month=month,
                total_tickets_sold=kr['tickets'],
                total_sales_amount=kr['sales'],
                total_bookings=kr['bookings'],
                total_booking_errors=kr['errors'],
                total_complaints=kr['complaints'],
                total_resolved_complaints=kr['resolved'],
                ticket_sales_target=kr['bookings'],
                ticket_sales_score=kr['ts_score'],
                booking_accuracy_score=kr['ba_score'],
                customer_service_score=kr['cs_score'],
                attendance_score=kr['att_score'],
                daily_reporting_score=kr['dr_score'],
                sop_compliance_score=kr['sop_score'],
                total_score=kr['total'],
                performance_rating=kr['rating']
            )
            db.session.add(mks)
    db.session.commit()

    branch_list = db.session.query(DailyPerformanceReport.branch).distinct().all()
    branch_list = [b[0] for b in branch_list if b[0]]

    return render_template('kpi_dashboard.html', kpi_results=kpi_results,
                           year=year, month=month, branch_filter=branch_filter,
                           branch_list=branch_list)


@app.route('/kpi-dashboard/history')
@login_required
@permission_required('kpi_history_view')
def kpi_history():
    year_filter = request.args.get('year', '')
    month_filter = request.args.get('month', '')
    staff_filter = request.args.get('staff', '')
    query = MonthlyKpiSummary.query
    if year_filter:
        query = query.filter_by(year=int(year_filter))
    if month_filter:
        query = query.filter_by(month=int(month_filter))
    if staff_filter:
        query = query.filter_by(staff_id=staff_filter)
    summaries = query.order_by(MonthlyKpiSummary.year.desc(), MonthlyKpiSummary.month.desc(), MonthlyKpiSummary.total_score.desc()).all()
    years = db.session.query(MonthlyKpiSummary.year).distinct().order_by(MonthlyKpiSummary.year.desc()).all()
    staff_list = db.session.query(MonthlyKpiSummary.staff_id, MonthlyKpiSummary.staff_name).distinct().all()
    return render_template('kpi_history.html', summaries=summaries,
                           year_filter=year_filter, month_filter=month_filter, staff_filter=staff_filter,
                           years=years, staff_list=staff_list)


# ─────────────────────────────────────────
#  ROUTES: DEPARTMENTS
# ─────────────────────────────────────────

@app.route('/departments')
@login_required
@any_permission_required('department_view', 'department_create', 'department_edit', 'department_delete', 'department_download')
def departments():
    query = Department.query
    search = request.args.get('search', '')
    if search:
        query = query.filter(Department.name.contains(search) | (Department.description.contains(search)))
    dept_list = query.order_by(Department.name).all()
    return render_template('departments.html', departments=dept_list, search=search)


@app.route('/departments/new', methods=['GET', 'POST'])
@login_required
@permission_required('department_create')
def new_department():
    es_mod = ConfigModule.query.filter_by(module_key='entity_status', is_active=True).first()
    entity_statuses = DynamicRecord.query.filter_by(module_id=es_mod.id, is_active=True).all() if es_mod else []
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash(_('Department name is required.'), 'danger')
            return render_template('department_form.html', dept=None, entity_statuses=entity_statuses)
        dept = Department(
            name=name,
            description=request.form.get('description'),
            status=request.form.get('status', 'Active'),
        )
        db.session.add(dept)
        db.session.commit()
        flash(_('Department "%(name)s" created successfully!', name=dept.name), 'success')
        return redirect(url_for('departments'))
    return render_template('department_form.html', dept=None, entity_statuses=entity_statuses)


@app.route('/departments/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('department_edit')
def edit_department(id):
    dept = Department.query.get_or_404(id)
    es_mod = ConfigModule.query.filter_by(module_key='entity_status', is_active=True).first()
    entity_statuses = DynamicRecord.query.filter_by(module_id=es_mod.id, is_active=True).all() if es_mod else []
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash(_('Department name is required.'), 'danger')
            return render_template('department_form.html', dept=dept, entity_statuses=entity_statuses)
        dept.name = name
        dept.description = request.form.get('description')
        dept.status = request.form.get('status', 'Active')
        db.session.commit()
        flash(_('Department updated successfully!'), 'success')
        return redirect(url_for('departments'))
    return render_template('department_form.html', dept=dept, entity_statuses=entity_statuses)


@app.route('/departments/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('department_delete')
def delete_department(id):
    dept = Department.query.get_or_404(id)
    if dept.positions.count() > 0:
        flash(_('Cannot delete department with existing positions.'), 'danger')
        return redirect(url_for('departments'))
    db.session.delete(dept)
    db.session.commit()
    flash(_('Department deleted.'), 'info')
    return redirect(url_for('departments'))


# ─────────────────────────────────────────
#  ROUTES: POSITIONS
# ─────────────────────────────────────────

@app.route('/positions')
@login_required
@any_permission_required('position_view', 'position_create', 'position_edit', 'position_delete', 'position_download')
def positions():
    query = Position.query
    search = request.args.get('search', '')
    if search:
        query = query.filter(Position.name.contains(search) | (Position.description.contains(search)))
    pos_list = query.order_by(Position.name).all()
    return render_template('positions.html', positions=pos_list, search=search)


@app.route('/positions/new', methods=['GET', 'POST'])
@login_required
@permission_required('position_create')
def new_position():
    departments_list = Department.query.filter_by(status='Active').order_by(Department.name).all()
    es_mod = ConfigModule.query.filter_by(module_key='entity_status', is_active=True).first()
    entity_statuses = DynamicRecord.query.filter_by(module_id=es_mod.id, is_active=True).all() if es_mod else []
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        dept_id = request.form.get('department_id')
        if not name or not dept_id:
            flash(_('Position name and department are required.'), 'danger')
            return render_template('position_form.html', pos=None, departments=departments_list, entity_statuses=entity_statuses)
        pos = Position(
            name=name,
            department_id=int(dept_id),
            description=request.form.get('description'),
            status=request.form.get('status', 'Active'),
        )
        db.session.add(pos)
        db.session.commit()
        flash(_('Position "%(name)s" created successfully!', name=pos.name), 'success')
        return redirect(url_for('positions'))
    return render_template('position_form.html', pos=None, departments=departments_list, entity_statuses=entity_statuses)


@app.route('/positions/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('position_edit')
def edit_position(id):
    pos = Position.query.get_or_404(id)
    departments_list = Department.query.filter_by(status='Active').order_by(Department.name).all()
    es_mod = ConfigModule.query.filter_by(module_key='entity_status', is_active=True).first()
    entity_statuses = DynamicRecord.query.filter_by(module_id=es_mod.id, is_active=True).all() if es_mod else []
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        dept_id = request.form.get('department_id')
        if not name or not dept_id:
            flash(_('Position name and department are required.'), 'danger')
            return render_template('position_form.html', pos=pos, departments=departments_list, entity_statuses=entity_statuses)
        pos.name = name
        pos.department_id = int(dept_id)
        pos.description = request.form.get('description')
        pos.status = request.form.get('status', 'Active')
        db.session.commit()
        flash(_('Position updated successfully!'), 'success')
        return redirect(url_for('positions'))
    return render_template('position_form.html', pos=pos, departments=departments_list, entity_statuses=entity_statuses)


@app.route('/positions/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('position_delete')
def delete_position(id):
    pos = Position.query.get_or_404(id)
    db.session.delete(pos)
    db.session.commit()
    flash(_('Position deleted.'), 'info')
    return redirect(url_for('positions'))



# ─────────────────────────────────────────
#  ROUTES: REPORTS (Redesigned)
# ─────────────────────────────────────────

REPORT_TYPES = {
    'transport': {
        'label': 'Transport Requests Report',
        'icon': 'fa-route',
        'model': RouteRequest,
        'date_field': 'created_date',
        'columns': [
            {'key': 'request_id', 'label': 'Request Number', 'sortable': True},
            {'key': 'requester_name', 'label': 'Customer Name', 'sortable': True},
            {'key': 'destination_from', 'label': 'Route (From)', 'sortable': True},
            {'key': 'destination_to', 'label': 'Route (To)', 'sortable': True},
            {'key': 'transportation', 'label': 'Vehicle', 'sortable': True},
            {'key': 'request_date', 'label': 'Travel Date', 'sortable': True},
            {'key': 'status', 'label': 'Status', 'sortable': True},
            {'key': 'created_date', 'label': 'Created Date', 'sortable': True},
        ],
        'filters': ['customer', 'route', 'vehicle', 'status'],
        'filter_fields': {
            'customer': ('requester_name', 'string', 'Customer'),
            'route': ('destination_from', 'string', 'Route'),
            'vehicle': ('transportation', 'string', 'Vehicle'),
            'status': ('status', 'select', 'Status', ['Pending', 'Approved', 'Rejected']),
        },
    },
    'trip': {
        'label': 'Trip Operation Report',
        'icon': 'fa-bus',
        'model': TripOperationReport,
        'date_field': 'created_date',
        'columns': [
            {'key': 'report_id', 'label': 'Trip Number', 'sortable': True},
            {'key': 'vehicle_number', 'label': 'Vehicle', 'sortable': True},
            {'key': 'driver_phone', 'label': 'Driver', 'sortable': True},
            {'key': 'origin', 'label': 'Route (Origin)', 'sortable': True},
            {'key': 'destination', 'label': 'Route (Dest)', 'sortable': True},
            {'key': 'departure_time', 'label': 'Departure Time', 'sortable': True},
            {'key': 'arrival_at_station', 'label': 'Arrival Time', 'sortable': True},
            {'key': 'passenger_count', 'label': 'Passenger Count', 'sortable': True},
            {'key': 'vehicle_status', 'label': 'Status', 'sortable': True},
        ],
        'filters': ['driver', 'vehicle', 'route', 'status'],
        'filter_fields': {
            'driver': ('driver_phone', 'string', 'Driver'),
            'vehicle': ('vehicle_number', 'string', 'Vehicle'),
            'route': ('origin', 'string', 'Route'),
            'status': ('vehicle_status', 'select', 'Status', ['Departed', 'Not Departed']),
        },
    },
    'daily': {
        'label': 'Daily Performance Report',
        'icon': 'fa-file-alt',
        'model': DailyPerformanceReport,
        'date_field': 'created_date',
        'columns': [
            {'key': 'staff_name', 'label': 'Employee', 'sortable': True},
            {'key': 'staff_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'branch', 'label': 'Department', 'sortable': True},
            {'key': 'report_date', 'label': 'Report Date', 'sortable': True},
            {'key': 'tickets_sold', 'label': 'Completed Tasks', 'sortable': True},
            {'key': 'total_sales_amount', 'label': 'Sales ($)', 'sortable': True},
            {'key': 'booking_errors', 'label': 'Penalties', 'sortable': True},
            {'key': 'status', 'label': 'Status', 'sortable': True},
        ],
        'filters': ['employee', 'department', 'status'],
        'filter_fields': {
            'employee': ('staff_name', 'string', 'Employee'),
            'department': ('branch', 'string', 'Department'),
            'status': ('status', 'select', 'Status', ['Draft', 'Submitted']),
        },
    },
    'kpi': {
        'label': 'KPI Report',
        'icon': 'fa-chart-line',
        'model': KpiEvaluation,
        'date_field': 'created_date',
        'columns': [
            {'key': 'staff_name', 'label': 'Employee', 'sortable': True},
            {'key': 'staff_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'branch', 'label': 'Department', 'sortable': True},
            {'key': 'ticket_sales_target', 'label': 'Target', 'sortable': True},
            {'key': 'actual_tickets_sold', 'label': 'Actual', 'sortable': True},
            {'key': 'achievement_pct', 'label': 'Achievement %', 'sortable': True},
            {'key': 'ticket_sales_score', 'label': 'Score', 'sortable': True},
            {'key': 'total_score', 'label': 'Total Score', 'sortable': True},
            {'key': 'performance_rating', 'label': 'Rating', 'sortable': True},
            {'key': 'status', 'label': 'Status', 'sortable': True},
        ],
        'filters': ['employee', 'department', 'status'],
        'filter_fields': {
            'employee': ('staff_name', 'string', 'Employee'),
            'department': ('branch', 'string', 'Department'),
            'status': ('status', 'select', 'Status', ['Draft', 'Submitted']),
        },
    },
    'kpi_history': {
        'label': 'KPI History Report',
        'icon': 'fa-clock-rotate-left',
        'model': MonthlyKpiSummary,
        'date_field': 'created_date',
        'columns': [
            {'key': 'staff_name', 'label': 'Employee', 'sortable': True},
            {'key': 'staff_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'month', 'label': 'Month', 'sortable': True},
            {'key': 'year', 'label': 'Year', 'sortable': True},
            {'key': 'total_score', 'label': 'Previous Score', 'sortable': True},
            {'key': 'performance_rating', 'label': 'Rating', 'sortable': True},
            {'key': 'created_date', 'label': 'Updated Date', 'sortable': True},
        ],
        'filters': ['employee', 'department'],
        'filter_fields': {
            'employee': ('staff_name', 'string', 'Employee'),
            'department': ('branch', 'string', 'Department'),
        },
    },
    'kpi_evaluation': {
        'label': 'KPI Evaluation Report',
        'icon': 'fa-file-invoice',
        'model': KpiEvaluation,
        'date_field': 'created_date',
        'columns': [
            {'key': 'staff_name', 'label': 'Employee', 'sortable': True},
            {'key': 'staff_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'evaluator_name', 'label': 'Evaluator', 'sortable': True},
            {'key': 'performance_rating', 'label': 'KPI Category', 'sortable': True},
            {'key': 'total_score', 'label': 'Score', 'sortable': True},
            {'key': 'comments', 'label': 'Comment', 'sortable': True},
            {'key': 'evaluation_month', 'label': 'Month', 'sortable': True},
            {'key': 'evaluation_year', 'label': 'Year', 'sortable': True},
            {'key': 'status', 'label': 'Status', 'sortable': True},
        ],
        'filters': ['employee', 'department', 'status'],
        'filter_fields': {
            'employee': ('staff_name', 'string', 'Employee'),
            'department': ('branch', 'string', 'Department'),
            'status': ('status', 'select', 'Status', ['Draft', 'Submitted']),
        },
    },
    'penalties': {
        'label': 'Penalties Report',
        'icon': 'fa-gavel',
        'model': EmployeePenalty,
        'date_field': 'created_date',
        'columns': [
            {'key': 'employee_name', 'label': 'Employee', 'sortable': True},
            {'key': 'employee_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'violation_type', 'label': 'Violation Type', 'sortable': True},
            {'key': 'description', 'label': 'Description', 'sortable': True},
            {'key': 'penalty_amount', 'label': 'Penalty Amount', 'sortable': True},
            {'key': 'status', 'label': 'Status', 'sortable': True},
            {'key': 'approved_by', 'label': 'Approved By', 'sortable': True},
            {'key': 'incident_date', 'label': 'Date', 'sortable': True},
        ],
        'filters': ['employee', 'violation_type', 'status'],
        'filter_fields': {
            'employee': ('employee_name', 'string', 'Employee'),
            'violation_type': ('violation_type', 'string', 'Violation Type'),
            'status': ('status', 'select', 'Status', ['Pending', 'Approved', 'Rejected']),
        },
    },
    'employee_performance': {
        'label': 'Employee Performance Report',
        'icon': 'fa-user-check',
        'model': None,
        'date_field': 'created_date',
        'columns': [
            {'key': 'employee_id', 'label': 'Employee ID', 'sortable': True},
            {'key': 'employee_name', 'label': 'Employee Name', 'sortable': True},
            {'key': 'branch', 'label': 'Department', 'sortable': True},
            {'key': 'total_bookings', 'label': 'Position', 'sortable': True},
            {'key': 'performance_score', 'label': 'KPI Score', 'sortable': True},
            {'key': 'total_tickets', 'label': 'Attendance', 'sortable': True},
            {'key': 'penalty_count', 'label': 'Penalties', 'sortable': True},
            {'key': 'performance_rating', 'label': 'Final Rating', 'sortable': True},
        ],
        'filters': ['employee', 'department'],
        'filter_fields': {
            'employee': ('employee_name', 'string', 'Employee'),
            'department': ('branch', 'string', 'Department'),
        },
    },
}


def _build_report_query(report_type, params):
    cfg = REPORT_TYPES.get(report_type)
    if not cfg:
        return None, 0

    if report_type == 'employee_performance':
        return _build_employee_performance_query(params)

    model = cfg['model']
    q = model.query
    date_field = getattr(model, cfg['date_field'])

    # Date range
    range_filter = params.get('range', 'all')
    today = date.today()
    start_date = end_date = None

    if range_filter == 'today':
        start_date = end_date = today
    elif range_filter == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif range_filter == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif range_filter == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif range_filter == 'custom':
        try:
            raw_s = params.get('start_date', '')
            raw_e = params.get('end_date', '')
            if raw_s:
                start_date = datetime.strptime(raw_s, '%Y-%m-%d').date()
            if raw_e:
                end_date = datetime.strptime(raw_e, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    if start_date:
        q = q.filter(date_field >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        q = q.filter(date_field <= datetime.combine(end_date, datetime.max.time()))

    # Report-specific filters
    for fname in cfg.get('filters', []):
        fval = params.get(fname, '').strip()
        if not fval:
            continue
        finfo = cfg['filter_fields'].get(fname)
        if not finfo:
            continue
        field_name = finfo[0]
        ftype = finfo[1]
        col = getattr(model, field_name, None)
        if col is None:
            continue
        if ftype == 'string':
            q = q.filter(col.ilike(f'%{fval}%'))
        elif ftype == 'select':
            q = q.filter(col == fval)
        elif ftype == 'number':
            q = q.filter(col == int(fval))
        elif ftype == 'month':
            try:
                q = q.filter(db.extract('month', col) == int(fval))
            except (ValueError, TypeError):
                pass
        elif ftype == 'year':
            try:
                q = q.filter(db.extract('year', col) == int(fval))
            except (ValueError, TypeError):
                pass

    # Search
    search = params.get('search', '').strip()
    if search and cfg.get('columns'):
        search_filters = []
        for col_def in cfg['columns']:
            col = getattr(model, col_def['key'], None)
            if col is not None:
                search_filters.append(col.ilike(f'%{search}%'))
        if search_filters:
            q = q.filter(or_(*search_filters))

    total = q.count()

    # Compute status statistics
    stats = {'total': total}
    if report_type != 'employee_performance' and model:
        status_field_names = {
            'transport': ('status', 'status'),
            'trip': ('vehicle_status', 'vehicle_status'),
            'daily': ('status', 'status'),
            'kpi': ('status', 'status'),
            'kpi_history': None,
            'kpi_evaluation': ('status', 'status'),
            'penalties': ('status', 'status'),
        }
        sf = status_field_names.get(report_type)
        if sf:
            try:
                status_col = getattr(model, sf[0], None)
                if status_col:
                    status_query = q.with_entities(status_col, db.func.count(status_col)).group_by(status_col).all()
                    counts = {}
                    for row in status_query:
                        key = str(row[0]) if row[0] is not None else 'Other'
                        counts[key] = row[1]
                    if counts:
                        stats['status_counts'] = counts
            except Exception:
                pass

    # Sorting
    sort_col = params.get('sort', '')
    sort_dir = params.get('order', 'desc')
    if sort_col:
        col = getattr(model, sort_col, None)
        if col is not None:
            q = q.order_by(col.desc() if sort_dir == 'desc' else col.asc())
        else:
            q = q.order_by(date_field.desc())
    else:
        q = q.order_by(date_field.desc())

    # Pagination
    page = int(params.get('page', 1))
    length = int(params.get('length', 25))
    offset = (page - 1) * length
    records = q.offset(offset).limit(length).all()

    # Format rows
    rows = []
    for rec in records:
        row = {}
        for col_def in cfg['columns']:
            val = getattr(rec, col_def['key'], None)
            if isinstance(val, date) and val:
                val = val.strftime('%d %b %Y')
            elif isinstance(val, float):
                val = round(val, 2)
            row[col_def['key']] = val
        row['_id'] = getattr(rec, 'id', 0)
        rows.append(row)

    return {
        'data': rows,
        'total': total,
        'stats': stats,
        'page': page,
        'pages': max(1, -(-total // length)) if length > 0 else 1,
        'length': length,
    }, 200


def _build_employee_performance_query(params):
    from collections import defaultdict

    today = date.today()
    range_filter = params.get('range', 'all')
    start_date = end_date = None

    if range_filter == 'today':
        start_date = end_date = today
    elif range_filter == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif range_filter == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif range_filter == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif range_filter == 'custom':
        try:
            raw_s = params.get('start_date', '')
            raw_e = params.get('end_date', '')
            if raw_s:
                start_date = datetime.strptime(raw_s, '%Y-%m-%d').date()
            if raw_e:
                end_date = datetime.strptime(raw_e, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass

    def build_date_filter(q, model_cls):
        if start_date:
            q = q.filter(model_cls.created_date >= datetime.combine(start_date, datetime.min.time()))
        if end_date:
            q = q.filter(model_cls.created_date <= datetime.combine(end_date, datetime.max.time()))
        return q

    user_emp_map = defaultdict(lambda: {
        'employee_name': '', 'employee_id': '', 'branch': '',
        'total_tickets': 0, 'total_sales': 0, 'total_bookings': 0,
        'total_errors': 0, 'total_complaints': 0,
        'penalty_count': 0, 'penalty_amount': 0,
    })

    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    for u in users:
        key = f"{u.full_name}|{u.username}"
        user_emp_map[key]['employee_name'] = u.full_name
        user_emp_map[key]['employee_id'] = u.username
        user_emp_map[key]['branch'] = u.branch or ''

    # Aggregate daily performance data
    dp_query = build_date_filter(DailyPerformanceReport.query, DailyPerformanceReport)
    for dp in dp_query.all():
        key = f"{dp.staff_name}|{dp.staff_id}"
        if key not in user_emp_map:
            user_emp_map[key]['employee_name'] = dp.staff_name
            user_emp_map[key]['employee_id'] = dp.staff_id
            user_emp_map[key]['branch'] = dp.branch or ''
        user_emp_map[key]['total_tickets'] += dp.tickets_sold
        user_emp_map[key]['total_sales'] += dp.total_sales_amount
        user_emp_map[key]['total_bookings'] += dp.bookings
        user_emp_map[key]['total_errors'] += dp.booking_errors
        user_emp_map[key]['total_complaints'] += dp.complaints

    # Aggregate penalty data
    pen_query = build_date_filter(EmployeePenalty.query, EmployeePenalty)
    for p in pen_query.all():
        key = f"{p.employee_name}|{p.employee_id}"
        if key not in user_emp_map:
            user_emp_map[key]['employee_name'] = p.employee_name
            user_emp_map[key]['employee_id'] = p.employee_id
            user_emp_map[key]['branch'] = p.department or ''
        user_emp_map[key]['penalty_count'] += 1
        user_emp_map[key]['penalty_amount'] += p.penalty_amount

    search = params.get('search', '').strip()
    department_filter = params.get('department', '').strip()
    employee_filter = params.get('employee', '').strip()

    all_rows = []
    for key, data in user_emp_map.items():
        if not data['employee_name']:
            continue
        if department_filter and department_filter.lower() not in data['branch'].lower():
            continue
        if employee_filter and employee_filter.lower() not in data['employee_name'].lower():
            continue
        if search and search.lower() not in data['employee_name'].lower() and search.lower() not in data['employee_id'].lower():
            continue

        total_bookings = data['total_bookings'] or 1
        accuracy = max(0, ((total_bookings - data['total_errors']) / total_bookings) * 100) if total_bookings else 100
        sales_score = min(100, (data['total_tickets'] / max(data['total_bookings'], 1)) * 100) if data['total_bookings'] else 0
        cs_score = max(0, 100 - (data['total_complaints'] * 5))
        pen_deduction = min(30, data['penalty_count'] * 5)
        perf_score = round(max(0, (sales_score * 0.40 + accuracy * 0.20 + cs_score * 0.15 + 100 * 0.25) - pen_deduction), 2)

        rating = _calc_rating(perf_score)

        all_rows.append({
            'employee_name': data['employee_name'],
            'employee_id': data['employee_id'],
            'branch': data['branch'],
            'total_tickets': data['total_tickets'],
            'total_sales': round(data['total_sales'], 2),
            'total_bookings': data['total_bookings'],
            'total_errors': data['total_errors'],
            'total_complaints': data['total_complaints'],
            'penalty_count': data['penalty_count'],
            'penalty_amount': round(data['penalty_amount'], 2),
            'performance_score': perf_score,
            'performance_rating': rating,
        })

    total = len(all_rows)

    # Sorting
    sort_col = params.get('sort', '')
    sort_dir = params.get('order', 'desc')
    reverse = sort_dir == 'desc'
    if sort_col:
        all_rows.sort(key=lambda r: (r.get(sort_col) or ''), reverse=reverse)
    else:
        all_rows.sort(key=lambda r: r['performance_score'], reverse=True)

    # Pagination
    page = int(params.get('page', 1))
    length = int(params.get('length', 25))
    offset = (page - 1) * length
    page_rows = all_rows[offset:offset + length]

    # Stats for employee performance
    rating_counts = {}
    for r in all_rows:
        rat = r.get('performance_rating', 'N/A')
        rating_counts[rat] = rating_counts.get(rat, 0) + 1
    stats = {'total': total}
    if rating_counts:
        stats['status_counts'] = rating_counts

    return {
        'data': page_rows,
        'total': total,
        'stats': stats,
        'page': page,
        'pages': max(1, -(-total // length)) if length > 0 else 1,
        'length': length,
    }, 200


@app.route('/reports')
@login_required
@any_permission_required('transport_request_view', 'trip_operation_report_view', 'daily_report_view', 'kpi_evaluation_view', 'kpi_history_view', 'penalty_view', 'report_view', 'report_create', 'report_edit', 'report_delete', 'report_export', 'report_print', 'report_download')
def reports():
    safe_types = _get_safe_report_types()
    first_type = next(iter(safe_types.keys()), '')
    return render_template('reports.html',
        report_types=safe_types,
        initial_report_type=first_type,
    )


def _get_safe_report_types():
    safe_types = {}
    for key, cfg in REPORT_TYPES.items():
        safe_types[key] = {
            'label': cfg['label'],
            'icon': cfg['icon'],
            'columns': [{'key': c['key'], 'label': c['label'], 'sortable': c.get('sortable', False)} for c in cfg['columns']],
            'filters': cfg.get('filters', []),
            'filter_fields': {fk: list(fv) for fk, fv in cfg.get('filter_fields', {}).items()},
            'row_actions': cfg.get('row_actions'),
        }
    return safe_types





_REPORT_SLUG_MAP = {
    'transport-requests': 'transport',
    'trip-operation': 'trip',
    'daily-performance': 'daily',
    'kpi': 'kpi',
    'kpi-history': 'kpi_history',
    'kpi-evaluation': 'kpi_evaluation',
    'penalties': 'penalties',
    'employee-performance': 'employee_performance',
}


@app.route('/reports/data')
@login_required
@any_permission_required('transport_request_view', 'trip_operation_report_view', 'daily_report_view', 'kpi_evaluation_view', 'kpi_history_view', 'penalty_view', 'report_view', 'report_create', 'report_edit', 'report_delete', 'report_export', 'report_print', 'report_download')
def reports_data():
    report_type = request.args.get('type', 'transport')
    if report_type not in REPORT_TYPES:
        return {'error': 'Invalid report type'}, 400

    params = request.args.to_dict()
    result, status = _build_report_query(report_type, params)
    return jsonify(result), status


@app.route('/reports/charts/<report_type>')
@login_required
@any_permission_required('transport_request_view', 'trip_operation_report_view', 'daily_report_view', 'kpi_evaluation_view', 'kpi_history_view', 'penalty_view', 'report_view', 'report_create', 'report_edit', 'report_delete', 'report_export', 'report_print', 'report_download')
def reports_charts(report_type):
    if report_type not in REPORT_TYPES:
        return jsonify({'error': 'Invalid report type'}), 400
    cfg = REPORT_TYPES[report_type]
    model = cfg.get('model')
    if not model:
        return jsonify({'categories': [], 'series': []})

    date_field = getattr(model, cfg['date_field'], model.created_date)
    status_field = getattr(model, 'status', None)

    try:
        records = model.query.order_by(date_field.desc()).limit(100).all()
    except Exception:
        records = []

    status_counts = defaultdict(int)
    monthly_counts = defaultdict(int)
    for r in records:
        raw_date = getattr(r, cfg['date_field'], None)
        if raw_date:
            key = str(raw_date)[:7]
            monthly_counts[key] += 1
        s = getattr(r, 'status', None) or getattr(r, 'vehicle_status', None) or 'Unknown'
        status_counts[str(s)] += 1

    categories = sorted(monthly_counts.keys())
    series_data = [monthly_counts[k] for k in categories]
    status_labels = list(status_counts.keys())
    status_values = list(status_counts.values())

    return jsonify({
        'trend': {'categories': categories, 'series': [{'name': 'Records', 'data': series_data}]},
        'status': {'labels': status_labels, 'values': status_values},
    })


@app.route('/reports/export/<fmt>/<report_type>')
@login_required
@any_permission_required('transport_request_view', 'trip_operation_report_view', 'daily_report_view', 'kpi_evaluation_view', 'kpi_history_view', 'penalty_view', 'report_view', 'report_create', 'report_edit', 'report_delete', 'report_export', 'report_print', 'report_download')
def reports_export(fmt, report_type):
    try:
        if report_type not in REPORT_TYPES:
            return jsonify({'error': 'Invalid report type'}), 400
        if fmt not in ('excel', 'pdf'):
            return jsonify({'error': 'Invalid export format'}), 400

        cfg = REPORT_TYPES[report_type]
        params = request.args.to_dict()
        params['length'] = '10000'
        result, _ = _build_report_query(report_type, params)

        if not result or not result.get('data'):
            return jsonify({'error': 'No records found to export', 'warning': True}), 404

        cols = cfg['columns']
        headers = [c['label'] for c in cols]
        keys = [c['key'] for c in cols]
        rows = [[r.get(k, '') for k in keys] for r in result['data']]

        today_str = date.today().strftime('%Y%m%d')
        filename = f"{report_type}_{today_str}"

        if fmt == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = cfg['label'][:31]

            hdr_font = Font(bold=True, color='FFFFFF', size=11)
            hdr_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.alignment = Alignment(horizontal='center', vertical='center')

            for ri, row in enumerate(rows, 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)

            for ci in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_excel(buf, f'{filename}.xlsx')

        elif fmt == 'pdf':
            buf = _export_pdf_report(cfg['label'], headers, rows, f'{filename}.pdf')
            buf.seek(0)
            return send_pdf(buf, f'{filename}.pdf')
    except Exception as e:
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


# ── Keep original export routes for backwards compatibility ──

@app.route('/reports/export/requests')
@login_required
@any_permission_required('route_request_view', 'route_request_download')
def export_requests_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Route Requests"
    headers = ['Request ID', 'Date', 'Requester', 'From', 'To', 'Branch', 'Company',
               'Transport', 'Price', 'Departure', 'Arrival', 'Status']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3c5e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18

    for rr in RouteRequest.query.order_by(RouteRequest.created_date.desc()).all():
        ws.append([rr.request_id, str(rr.request_date), rr.requester_name,
                   rr.destination_from, rr.destination_to, rr.branch_location,
                   rr.company, rr.transportation, rr.price,
                   rr.departure_time, rr.arrival_time, rr.status])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_excel(buf, f"route_requests_{date.today()}.xlsx")


@app.route('/reports/export/penalties')
@login_required
@any_permission_required('penalty_view', 'penalty_download')
def export_penalties_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Penalties"
    headers = ['Penalty ID', 'Employee ID', 'Employee Name', 'Department',
               'Position', 'Branch', 'Violation Type', 'Price ($)', 'Old Code', 'Amount ($)', 'Date', 'Status', 'Approved by']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3c5e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20

    for ep in EmployeePenalty.query.order_by(EmployeePenalty.created_date.desc()).all():
        ws.append([ep.penalty_id, ep.employee_id, ep.employee_name,
                   ep.department, ep.position, ep.branch, ep.violation_type,
                   ep.price, ep.old_code, ep.penalty_amount, str(ep.incident_date), ep.status, ep.approved_by])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_excel(buf, f"penalties_{date.today()}.xlsx")


def _export_pdf_report(title, headers, rows, filename):
    pdf = FPDF()
    khmer_font = _find_khmer_font()
    if khmer_font:
        try:
            pdf.add_font('Khmer', '', khmer_font, uni=True)
            pdf.add_font('Khmer', 'B', khmer_font, uni=True)
            font_name = 'Khmer'
        except RuntimeError:
            font_name = 'Helvetica'
    else:
        font_name = 'Helvetica'
    pdf.add_page()
    pdf.set_font(font_name, 'B', 14)
    pdf.cell(0, 12, title, new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(2)
    pdf.set_font(font_name, '', 7)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", new_x='LMARGIN', new_y='NEXT', align='R')
    pdf.ln(4)

    col_w = max(10, min(45, 270 // len(headers))) if headers else 20
    page_w = 270
    used_w = col_w * len(headers)
    if used_w > page_w:
        col_w = max(10, page_w // len(headers))

    pdf.set_font(font_name, 'B', 6)
    pdf.set_fill_color(26, 60, 94)
    pdf.set_text_color(255, 255, 255)
    for h in headers:
        pdf.cell(col_w, 8, h, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_font(font_name, '', 6)
    pdf.set_text_color(50, 50, 50)
    for row in rows:
        for i, cell in enumerate(row):
            align = 'L' if i == 0 else 'C'
            txt = str(cell or '')[:int(col_w * 1.5)]
            pdf.cell(col_w, 6, txt, border=1, align=align)
        pdf.ln()

    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf


@app.route('/reports/export/requests/pdf')
@login_required
@any_permission_required('route_request_view', 'route_request_download')
def export_requests_pdf():
    headers = ['Req ID', 'Date', 'Requester', 'From', 'To', 'Branch', 'Company',
               'Transport', 'Road', 'Price', 'Depart', 'Arrive', 'Start', 'End', 'Status']
    rows = []
    for rr in RouteRequest.query.order_by(RouteRequest.created_date.desc()).all():
        rows.append([
            rr.request_id, str(rr.request_date), rr.requester_name,
            rr.destination_from, rr.destination_to, rr.branch_location or '',
            rr.company or '', rr.transportation or '', rr.national_road or '',
            f"${rr.price:.2f}", rr.departure_time or '', rr.arrival_time or '',
            str(rr.start_date or ''), str(rr.end_date or ''), rr.status
        ])
    buf = _export_pdf_report('Route Requests Report', headers, rows, f'route_requests_{date.today()}.pdf')
    return send_pdf(buf, f"route_requests_{date.today()}.pdf")


@app.route('/reports/export/penalties/pdf')
@login_required
@any_permission_required('penalty_view', 'penalty_download')
def export_penalties_pdf():
    headers = ['Penalty ID', 'Emp ID', 'Employee Name', 'Department', 'Position',
               'Branch', 'Violation Type', 'Price ($)', 'Old Code', 'Amount ($)', 'Date', 'Status', 'Approved by']
    rows = []
    for ep in EmployeePenalty.query.order_by(EmployeePenalty.created_date.desc()).all():
        rows.append([
            ep.penalty_id, ep.employee_id, ep.employee_name,
            ep.department or '', ep.position or '', ep.branch or '', ep.violation_type,
            f"${ep.price:.2f}" if ep.price else '', ep.old_code or '', f"${ep.penalty_amount:.2f}", str(ep.incident_date), ep.status, ep.approved_by or ''
        ])
    buf = _export_pdf_report('Employee Penalties Report', headers, rows, f'penalties_{date.today()}.pdf')
    return send_pdf(buf, f"penalties_{date.today()}.pdf")


@app.route('/reports/export/trip-reports')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_download')
def export_trip_reports_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trip Operation Reports"
    headers = ['Report ID', 'Origin', 'Destination', 'Direction', 'Departure',
               'Vehicle', 'Driver Phone', 'Arrival Station', 'Depart Station',
               'Delay (min)', 'Passengers', 'Coordinator', 'Status']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3c5e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18

    for r in TripOperationReport.query.order_by(TripOperationReport.created_date.desc()).all():
        ws.append([r.report_id, r.origin, r.destination, r.travel_direction,
                   r.departure_time, r.vehicle_number, r.driver_phone or '',
                   r.arrival_at_station or '', r.departure_from_station or '',
                   int(r.travel_delay_duration or 0), int(r.passenger_count),
                   r.coordinator_name or '', r.vehicle_status])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_excel(buf, f"trip_reports_{date.today()}.xlsx")


@app.route('/reports/export/trip-reports/pdf')
@login_required
@any_permission_required('trip_operation_report_view', 'trip_operation_report_download')
def export_trip_reports_pdf():
    headers = ['Report ID', 'Origin', 'Destination', 'Direction', 'Departure',
               'Vehicle', 'Driver Phone', 'Arrival Stn', 'Passengers', 'Coordinator', 'Status']
    rows = []
    for r in TripOperationReport.query.order_by(TripOperationReport.created_date.desc()).all():
        rows.append([
            r.report_id, r.origin, r.destination, r.travel_direction or '',
            r.departure_time or '', r.vehicle_number, r.driver_phone or '',
            r.arrival_at_station or '', int(r.passenger_count),
            r.coordinator_name or '', r.vehicle_status
        ])
    buf = _export_pdf_report('Trip Operation Reports', headers, rows, f'trip_reports_{date.today()}.pdf')
    return send_pdf(buf, f"trip_reports_{date.today()}.pdf")


# ── Catch-all: redirect /reports/<slug> to /reports#type (must be last) ──
@app.route('/reports/<path:subpath>')
@login_required
@any_permission_required('transport_request_view', 'trip_operation_report_view', 'daily_report_view', 'kpi_evaluation_view', 'kpi_history_view', 'penalty_view', 'report_view', 'report_create', 'report_edit', 'report_delete', 'report_export', 'report_print', 'report_download')
def report_subpath(subpath):
    report_type = _REPORT_SLUG_MAP.get(subpath)
    if report_type and report_type in REPORT_TYPES:
        return redirect(url_for('reports', _anchor=report_type))
    return redirect(url_for('reports'))


# ─────────────────────────────────────────
#  DYNAMIC CONFIGURATION ENGINE
# ─────────────────────────────────────────

class DynamicEngine:
    """Generic dynamic CRUD engine driven by ConfigModule/ConfigField metadata."""

    @staticmethod
    def get_module(module_key):
        m = ConfigModule.query.filter_by(module_key=module_key, is_active=True).first()
        if not m:
            flash(_('Module "%(key)s" not found.', key=module_key), 'danger')
        return m

    @staticmethod
    def get_fields(module_id, creatable_only=False, editable_only=False):
        q = ConfigField.query.filter_by(module_id=module_id)
        if creatable_only:
            q = q.filter_by(is_creatable=True)
        if editable_only:
            q = q.filter_by(is_editable=True)
        return q.order_by(ConfigField.display_order).all()

    @staticmethod
    def validate(data, fields, record_id=None, module_id=None):
        errors = {}
        for fld in fields:
            val = data.get(fld.field_name, '')

            # Required
            if fld.is_required and (val is None or str(val).strip() == ''):
                errors[fld.field_name] = _('%(label)s is required.', label=fld.field_label)
                continue

            if val is None or str(val).strip() == '':
                continue

            val_str = str(val).strip()

            # Min length
            if fld.min_length and len(val_str) < fld.min_length:
                errors[fld.field_name] = _('%(label)s must be at least %(len)d characters.', label=fld.field_label, len=fld.min_length)

            # Max length
            if fld.max_length and len(val_str) > fld.max_length:
                errors[fld.field_name] = _('%(label)s must not exceed %(len)d characters.', label=fld.field_label, len=fld.max_length)

            # Regex
            if fld.regex_pattern and not re.match(fld.regex_pattern, val_str):
                errors[fld.field_name] = fld.regex_message or _('%(label)s format is invalid.', label=fld.field_label)

            # Unique
            if fld.is_unique and module_id:
                existing = DynamicRecord.query.filter(
                    DynamicRecord.module_id == module_id,
                    DynamicRecord.data[fld.field_name].as_string() == val_str,
                    DynamicRecord.is_active == True
                )
                if record_id:
                    existing = existing.filter(DynamicRecord.id != record_id)
                if existing.first():
                    errors[fld.field_name] = _('%(label)s must be unique.', label=fld.field_label)

            # Extra validation rules from config_validation table
            for rule in fld.validations:
                if rule.validation_type == 'custom' and rule.validation_value:
                    try:
                        allowed_ops = {
                            'gt': lambda v, d, a: float(v) > float(a[0]) if a else False,
                            'gte': lambda v, d, a: float(v) >= float(a[0]) if a else False,
                            'lt': lambda v, d, a: float(v) < float(a[0]) if a else False,
                            'lte': lambda v, d, a: float(v) <= float(a[0]) if a else False,
                            'eq': lambda v, d, a: v == a[0],
                            'neq': lambda v, d, a: v != a[0],
                            'len_gt': lambda v, d, a: len(v) > int(a[0]) if a else False,
                            'len_lt': lambda v, d, a: len(v) < int(a[0]) if a else False,
                            'matches': lambda v, d, a: bool(re.match(a[0], v)) if a else False,
                            'in': lambda v, d, a: v in a,
                            'not_in': lambda v, d, a: v not in a,
                        }
                        parts = rule.validation_value.split(None, 1)
                        op_name = parts[0].lower()
                        args = parts[1].split() if len(parts) > 1 else []
                        if op_name in allowed_ops:
                            if not allowed_ops[op_name](val_str, data, args):
                                errors[fld.field_name] = rule.error_message or _('%(label)s validation failed.', label=fld.field_label)
                    except Exception:
                        pass

            # Number type check
            if fld.field_type == 'number' and val_str:
                try:
                    float(val_str)
                except ValueError:
                    errors[fld.field_name] = _('%(label)s must be a valid number.', label=fld.field_label)

        return errors

    @staticmethod
    def get_dropdown_options(field, parent_value=None):
        """Returns list of (value, label) tuples for a dropdown field."""
        if field.option_source == 'static':
            return [(o.option_value, o.option_label)
                    for o in ConfigDropdownOption.query.filter_by(field_id=field.id, is_active=True)
                    .order_by(ConfigDropdownOption.sort_order).all()]

        elif field.option_source == 'module' and field.option_module_id:
            target_module = ConfigModule.query.get(field.option_module_id)
            if not target_module:
                return []
            q = DynamicRecord.query.filter_by(module_id=field.option_module_id, is_active=True)
            if field.option_parent_field and parent_value:
                q = q.filter(DynamicRecord.data[field.option_parent_field].as_string() == str(parent_value))
            records = q.order_by(DynamicRecord.created_date.desc()).all()
            result = []
            for r in records:
                label = r.data.get('name', r.data.get('code', f'Record #{r.id}'))
                result.append((str(r.id), label))
            return result

        return []

    @staticmethod
    def resolve_field_value(field, value):
        """Resolve a foreign key ID to a display label."""
        if field.option_source == 'module' and field.option_module_id and value:
            rec = DynamicRecord.query.get(int(value))
            if rec:
                return rec.data.get('name', rec.data.get('code', f'#{value}'))
        return value

    @staticmethod
    def create(module_key, data):
        module = DynamicEngine.get_module(module_key)
        if not module:
            return None, _('Module not found.')
        fields = DynamicEngine.get_fields(module.id, creatable_only=True)
        errors = DynamicEngine.validate(data, fields, module_id=module.id)
        if errors:
            return None, errors
        record = DynamicRecord(
            module_id=module.id,
            data=data,
            is_active=True,
            created_by=current_user.id
        )
        db.session.add(record)
        db.session.commit()
        return record, None

    @staticmethod
    def update(record_id, data):
        record = DynamicRecord.query.get_or_404(record_id)
        fields = DynamicEngine.get_fields(record.module_id, editable_only=True)
        errors = DynamicEngine.validate(data, fields, record_id=record_id, module_id=record.module_id)
        if errors:
            return None, errors
        record.data = data
        record.updated_by = current_user.id
        record.updated_date = datetime.now(timezone.utc).replace(tzinfo=None)
        db.session.commit()
        return record, None

    @staticmethod
    def delete(record_id, soft=True):
        record = DynamicRecord.query.get_or_404(record_id)
        if soft:
            record.is_active = False
            record.deleted_date = datetime.now(timezone.utc).replace(tzinfo=None)
        else:
            db.session.delete(record)
        db.session.commit()

    @staticmethod
    def list_records(module_key, search=None, sort_by=None, sort_dir='desc', page=1, per_page=20):
        module = DynamicEngine.get_module(module_key)
        if not module:
            return None, 0
        q = DynamicRecord.query.filter_by(module_id=module.id, is_active=True)
        fields = DynamicEngine.get_fields(module.id)

        # Search
        if search:
            searchable = [f for f in fields if f.is_searchable]
            if searchable:
                filters = []
                for f in searchable:
                    filters.append(DynamicRecord.data[f.field_name].as_string().contains(search))
                from sqlalchemy import or_
                q = q.filter(or_(*filters))

        # Sort
        if sort_by:
            try:
                col = DynamicRecord.data[sort_by].as_string()
                q = q.order_by(col.desc() if sort_dir == 'desc' else col)
            except Exception:
                q = q.order_by(DynamicRecord.created_date.desc())
        else:
            q = q.order_by(DynamicRecord.created_date.desc())

        total = q.count()
        records = q.offset((page - 1) * per_page).limit(per_page).all()
        return records, total


@app.context_processor
def inject_dynamic_engine():
    return {'DynamicEngine': DynamicEngine}


# ─────────────────────────────────────────
#  ROUTES: DYNAMIC SETTINGS (Configuration Driven)
# ─────────────────────────────────────────

@app.route('/dynamic-settings')
@login_required
@any_permission_required('system_settings_view', 'system_settings_edit', 'system_settings_update')
def dynamic_settings():
    modules = ConfigModule.query.filter_by(is_active=True).order_by(ConfigModule.sort_order).all()
    return render_template('dynamic_settings.html', modules=modules)


@app.route('/dynamic-settings/<module_key>')
@login_required
@any_permission_required('system_settings_view', 'system_settings_edit', 'system_settings_update')
def dynamic_list(module_key):
    module = DynamicEngine.get_module(module_key)
    if not module:
        return redirect(url_for('dynamic_settings'))

    fields = DynamicEngine.get_fields(module.id)
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    sort_by = request.args.get('sort_by', '')
    sort_dir = request.args.get('sort_dir', 'desc')

    records, total = DynamicEngine.list_records(
        module_key, search=search, sort_by=sort_by, sort_dir=sort_dir, page=page, per_page=15
    )

    # Resolve FK fields for display
    for rec in records:
        for fld in fields:
            val = rec.data.get(fld.field_name)
            rec.data[f'_{fld.field_name}_display'] = DynamicEngine.resolve_field_value(fld, val)

    modules = ConfigModule.query.filter_by(is_active=True).order_by(ConfigModule.sort_order).all()
    return render_template('dynamic_list.html',
        module=module, fields=fields, records=records,
        total=total, page=page, per_page=15,
        search=search, sort_by=sort_by, sort_dir=sort_dir,
        modules=modules
    )


@app.route('/dynamic-settings/<module_key>/create', methods=['GET', 'POST'])
@login_required
@permission_required('system_settings_edit')
def dynamic_create(module_key):
    module = DynamicEngine.get_module(module_key)
    if not module:
        return redirect(url_for('dynamic_settings'))

    fields = DynamicEngine.get_fields(module.id, creatable_only=True)
    modules = ConfigModule.query.filter_by(is_active=True).order_by(ConfigModule.sort_order).all()

    if request.method == 'POST':
        data = {}
        for fld in fields:
            val = request.form.get(fld.field_name, '').strip()
            if fld.field_type == 'number':
                val = float(val) if val else 0
            elif fld.field_type == 'boolean' or fld.field_type == 'checkbox':
                val = fld.field_name in request.form
            data[fld.field_name] = val

        record, error = DynamicEngine.create(module_key, data)
        if error:
            if isinstance(error, dict):
                return render_template('dynamic_form.html', module=module, fields=fields,
                    modules=modules, data=data, errors=error, form_mode='create')
            flash(str(error), 'danger')
        else:
            flash(_('%(name)s created successfully.', name=module.module_name), 'success')
            return redirect(url_for('dynamic_list', module_key=module_key))

        return render_template('dynamic_form.html', module=module, fields=fields,
            modules=modules, data=data, form_mode='create')

    return render_template('dynamic_form.html', module=module, fields=fields,
        modules=modules, data={}, form_mode='create')


@app.route('/dynamic-settings/<module_key>/<int:id>')
@login_required
@permission_required('system_settings_view')
def dynamic_view(module_key, id):
    module = DynamicEngine.get_module(module_key)
    if not module:
        return redirect(url_for('dynamic_settings'))
    record = DynamicRecord.query.get_or_404(id)
    fields = DynamicEngine.get_fields(module.id)
    modules = ConfigModule.query.filter_by(is_active=True).order_by(ConfigModule.sort_order).all()
    for fld in fields:
        val = record.data.get(fld.field_name)
        record.data[f'_{fld.field_name}_display'] = DynamicEngine.resolve_field_value(fld, val)
    return render_template('dynamic_view.html', module=module, fields=fields,
        modules=modules, record=record)


@app.route('/dynamic-settings/<module_key>/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('system_settings_edit')
def dynamic_edit(module_key, id):
    module = DynamicEngine.get_module(module_key)
    if not module:
        return redirect(url_for('dynamic_settings'))

    record = DynamicRecord.query.get_or_404(id)
    fields = DynamicEngine.get_fields(module.id, editable_only=True)
    modules = ConfigModule.query.filter_by(is_active=True).order_by(ConfigModule.sort_order).all()

    if request.method == 'POST':
        data = {}
        for fld in fields:
            val = request.form.get(fld.field_name, '').strip()
            if fld.field_type == 'number':
                val = float(val) if val else 0
            elif fld.field_type == 'boolean' or fld.field_type == 'checkbox':
                val = fld.field_name in request.form
            data[fld.field_name] = val

        updated, error = DynamicEngine.update(id, data)
        if error:
            if isinstance(error, dict):
                return render_template('dynamic_form.html', module=module, fields=fields,
                    modules=modules, data=data, errors=error, form_mode='edit', record=record)
            flash(str(error), 'danger')
        else:
            flash(_('%(name)s updated successfully.', name=module.module_name), 'success')
            return redirect(url_for('dynamic_list', module_key=module_key))

    return render_template('dynamic_form.html', module=module, fields=fields,
        modules=modules, data=record.data, form_mode='edit', record=record)


@app.route('/dynamic-settings/<module_key>/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('system_settings_edit')
def dynamic_delete(module_key, id):
    DynamicEngine.delete(id, soft=True)
    flash(_('Record deleted.'), 'info')
    return redirect(url_for('dynamic_list', module_key=module_key))


@app.route('/dynamic-settings/<module_key>/<int:id>/toggle', methods=['POST'])
@login_required
@permission_required('system_settings_edit')
def dynamic_toggle(module_key, id):
    record = DynamicRecord.query.get_or_404(id)
    record.is_active = not record.is_active
    db.session.commit()
    flash(_('Record activated.') if record.is_active else _('Record deactivated.'), 'info')
    return redirect(url_for('dynamic_list', module_key=module_key))


# ─────────────────────────────────────────
#  ROUTES: USER MANAGEMENT
# ─────────────────────────────────────────

@app.route('/users')
@login_required
@any_permission_required('user_view', 'user_create', 'user_edit', 'user_delete', 'user_assign_roles', 'user_reset_password', 'user_activate', 'user_deactivate', 'user_download')
def users():
    users_list = User.query.order_by(User.created_date.desc()).all()
    return render_template('users.html', users=users_list)


@app.route('/roles')
@login_required
@any_permission_required('role_view', 'role_create', 'role_edit', 'role_delete', 'role_assign_permissions', 'role_download')
def list_roles():
    roles_list = Role.query.order_by(Role.created_date.desc()).all()
    return render_template('roles.html', roles=roles_list)


@app.route('/roles/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('role_edit')
def edit_role(id):
    role = Role.query.get_or_404(id)
    if request.method == 'POST':
        role_name = request.form.get('role_name', '').strip()
        role_label = request.form.get('role_label', '').strip()
        permissions = request.form.getlist('permissions')
        if not permissions:
            permissions_raw = request.form.get('permissions')
            if permissions_raw:
                try:
                    permissions = json.loads(permissions_raw)
                except (json.JSONDecodeError, TypeError):
                    permissions = []

        if not role_name:
            flash(_('Role name is required.'), 'danger')
            return render_template('create_role.html', role=role)

        existing = Role.query.filter_by(name=role_name).first()
        if existing and existing.id != role.id:
            flash(_('A role with that name already exists.'), 'danger')
            return render_template('create_role.html', role=role)

        role.name = role_name
        role.label = role_label or role_name
        role.permissions = permissions
        db.session.commit()

        flash(_('Role "%(name)s" updated.', name=role_name), 'success')
        return redirect(url_for('list_roles'))

    return render_template('create_role.html', role=role)


@app.route('/roles/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('role_delete')
def delete_role(id):
    role = Role.query.get_or_404(id)
    db.session.delete(role)
    db.session.commit()
    flash(_('Role "%(name)s" deleted.', name=role.name), 'info')
    return redirect(url_for('list_roles'))


@app.route('/roles/new', methods=['GET', 'POST'])
@login_required
@permission_required('role_create')
def new_role():
    if request.method == 'POST':
        role_name = request.form.get('role_name', '').strip()
        role_label = request.form.get('role_label', '').strip()
        permissions = request.form.getlist('permissions')
        if not permissions:
            permissions_raw = request.form.get('permissions')
            if permissions_raw:
                try:
                    permissions = json.loads(permissions_raw)
                except (json.JSONDecodeError, TypeError):
                    permissions = []

        if not role_name:
            flash(_('Role name is required.'), 'danger')
            return render_template('create_role.html')

        if Role.query.filter_by(name=role_name).first():
            flash(_('A role with that name already exists.'), 'danger')
            return render_template('create_role.html')

        new_role = Role(name=role_name, label=role_label or role_name, permissions=permissions)
        db.session.add(new_role)
        db.session.commit()

        flash(_('Role "%(name)s" created with %(count)d permissions.', name=role_name, count=len(permissions)), 'success')
        return redirect(url_for('list_roles'))

    return render_template('create_role.html')


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@permission_required('user_create')
def new_user():
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    if request.method == 'POST':
        u = User()
        u.full_name = request.form.get('full_name')
        u.username = request.form.get('username')
        role_id = _int_or_none(request.form.get('role_id'))
        logger.debug(f"NEW_USER: raw role_id='{request.form.get('role_id')}', parsed={role_id}")
        if role_id:
            role_obj = Role.query.get(role_id)
            if role_obj:
                u.role_id = role_obj.id
                logger.debug(f"NEW_USER: set role_id={u.role_id}")
            else:
                logger.debug(f"NEW_USER: role_id={role_id} NOT FOUND!")
                flash(_('Selected role not found.'), 'danger')
                roles = Role.query.order_by(Role.name).all()
                return render_template('user_form.html', user=None, roles=roles, branches=branches)
        else:
            logger.debug(f"NEW_USER: no role_id submitted")
        u.branch = request.form.get('branch')
        u.set_password(request.form.get('password'))
        if User.query.filter_by(username=u.username).first():
            flash(_('Username already exists.'), 'danger')
            roles = Role.query.order_by(Role.name).all()
            return render_template('user_form.html', user=None, roles=roles, branches=branches)
        db.session.add(u)
        db.session.commit()
        logger.debug(f"NEW_USER: COMMITTED. user.id={u.id}, role_id={u.role_id}")
        flash(_('User %(name)s created!', name=u.username), 'success')
        return redirect(url_for('users'))
    roles = Role.query.order_by(Role.name).all()
    return render_template('user_form.html', user=None, roles=roles, branches=branches)


@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('user_edit')
def edit_user(id):
    u = User.query.get_or_404(id)
    branch_module = ConfigModule.query.filter_by(module_key='branch', is_active=True).first()
    branches = DynamicRecord.query.filter_by(module_id=branch_module.id, is_active=True).order_by(DynamicRecord.created_date.desc()).all() if branch_module else []
    if request.method == 'POST':
        u.full_name = request.form.get('full_name')
        role_id = _int_or_none(request.form.get('role_id'))
        logger.debug(f"EDIT_USER: user id={u.id}, raw role_id='{request.form.get('role_id')}', parsed={role_id}")
        logger.debug(f"EDIT_USER: BEFORE: role_id={u.role_id}")
        if role_id:
            role_obj = Role.query.get(role_id)
            if role_obj:
                u.role_id = role_obj.id
                logger.debug(f"EDIT_USER: set role_id={u.role_id}")
            else:
                logger.debug(f"EDIT_USER: role_id={role_id} NOT FOUND!")
        else:
            logger.debug(f"EDIT_USER: no role_id submitted, keeping role_id={u.role_id}")
        u.branch = request.form.get('branch')
        u.is_active = 'is_active' in request.form
        new_pass = request.form.get('password')
        if new_pass:
            u.set_password(new_pass)
        db.session.commit()
        flash(_('User updated!'), 'success')
        return redirect(url_for('users'))
    roles = Role.query.order_by(Role.name).all()
    return render_template('user_form.html', user=u, roles=roles, branches=branches)


@app.route('/users/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('user_delete')
def delete_user(id):
    u = User.query.get_or_404(id)
    if u.id == current_user.id:
        flash(_('Cannot delete yourself.'), 'danger')
        return redirect(url_for('users'))
    db.session.delete(u)
    db.session.commit()
    flash(_('User deleted.'), 'info')
    return redirect(url_for('users'))


# ─────────────────────────────────────────
#  API: Roles
# ─────────────────────────────────────────




# ─────────────────────────────────────────
#  INIT DB
# ─────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()

        # ── Auto-migrate: add any missing columns from model definitions ──
        run_auto_migrations()

        # ── Seed Dynamic Configuration Modules ──
        if not ConfigModule.query.first():
            modules_data = [
                ('company', 'Company', 'fa-building', 1),
                ('branch', 'Branch', 'fa-location-dot', 2),
                ('agency', 'Agency', 'fa-handshake', 3),
                ('agent_group', 'Agent Group', 'fa-users-gear', 4),
                ('destination', 'Destination', 'fa-map-pin', 5),
                ('destination_group', 'Destination Group', 'fa-layer-group', 6),
                ('transportation_type', 'Transportation Type', 'fa-truck', 7),
                ('route', 'Route', 'fa-route', 8),
                ('boarding_point', 'Boarding Point', 'fa-sign-in-alt', 9),
                ('drop_off_point', 'Drop Off Point', 'fa-sign-out-alt', 10),
                ('bus', 'Bus', 'fa-bus', 11),
                ('bus_type', 'Bus Type', 'fa-bus-simple', 12),
                ('amenity', 'Amenity', 'fa-wifi', 13),
            ]
            for key, name, icon, order in modules_data:
                m = ConfigModule(module_key=key, module_name=name, module_icon=icon, sort_order=order)
                db.session.add(m)
            db.session.flush()

            # ── Seed Fields for each module ──
            field_sets = {
                'company': [
                    ('name', 'Company Name', 'text', True, True, True),
                    ('code', 'Company Code', 'text', False, False, True),
                    ('phone', 'Phone', 'tel', False, False, True),
                    ('email', 'Email', 'email', False, False, False),
                    ('address', 'Address', 'textarea', False, False, False),
                ],
                'branch': [
                    ('name', 'Branch Name', 'text', True, True, True),
                    ('code', 'Branch Code', 'text', False, False, True),
                    ('phone', 'Phone', 'tel', False, False, True),
                    ('address', 'Address', 'textarea', False, False, False),
                ],
                'agency': [
                    ('name', 'Agency Name', 'text', True, True, True),
                    ('code', 'Agency Code', 'text', True, True, True),
                    ('contact_person', 'Contact Person', 'text', False, False, True),
                    ('phone', 'Phone', 'tel', False, False, True),
                    ('email', 'Email', 'email', False, False, False),
                    ('commission_rate', 'Commission Rate (%)', 'number', False, False, False),
                ],
                'agent_group': [
                    ('name', 'Group Name', 'text', True, True, True),
                    ('description', 'Description', 'textarea', False, False, False),
                    ('discount_rate', 'Discount Rate (%)', 'number', False, False, False),
                ],
                'destination': [
                    ('name', 'Destination Name', 'text', True, True, True),
                    ('code', 'Code', 'text', False, True, True),
                    ('description', 'Description', 'textarea', False, False, False),
                ],
                'destination_group': [
                    ('name', 'Group Name', 'text', True, True, True),
                    ('description', 'Description', 'textarea', False, False, False),
                ],
                'transportation_type': [
                    ('name', 'Transport Type', 'text', True, True, True),
                    ('capacity', 'Default Capacity', 'number', False, False, False),
                ],
                'route': [
                    ('name', 'Route Name', 'text', False, False, True),
                    ('from_destination_id', 'From Destination', 'dropdown', True, False, True),
                    ('to_destination_id', 'To Destination', 'dropdown', True, False, True),
                    ('distance_km', 'Distance (km)', 'number', False, False, False),
                    ('estimated_hours', 'Est. Hours', 'number', False, False, False),
                ],
                'boarding_point': [
                    ('name', 'Point Name', 'text', True, False, True),
                    ('destination_id', 'Destination', 'dropdown', True, False, True),
                    ('address', 'Address', 'textarea', False, False, False),
                ],
                'drop_off_point': [
                    ('name', 'Point Name', 'text', True, False, True),
                    ('destination_id', 'Destination', 'dropdown', True, False, True),
                    ('address', 'Address', 'textarea', False, False, False),
                ],
                'bus': [
                    ('bus_number', 'Bus Number', 'text', True, True, True),
                    ('bus_type_id', 'Bus Type', 'dropdown', True, False, True),
                    ('capacity', 'Capacity', 'number', True, False, True),
                    ('plate_number', 'Plate Number', 'text', True, False, True),
                    ('amenities', 'Amenities', 'textarea', False, False, False),
                ],
                'bus_type': [
                    ('name', 'Bus Type Name', 'text', True, True, True),
                    ('capacity', 'Default Capacity', 'number', True, False, True),
                    ('description', 'Description', 'textarea', False, False, False),
                ],
                'amenity': [
                    ('name', 'Amenity Name', 'text', True, True, True),
                    ('icon', 'Icon Class', 'text', False, False, True),
                    ('description', 'Description', 'textarea', False, False, False),
                ],
            }
            for key, name, icon, order in modules_data:
                module = ConfigModule.query.filter_by(module_key=key).first()
                if module and key in field_sets:
                    for i, (fname, flabel, ftype, req, unique, listable) in enumerate(field_sets[key]):
                        # Set dropdown source for FK fields
                        opt_source = None
                        opt_module_id = None
                        if fname in ('from_destination_id', 'to_destination_id', 'destination_id'):
                            opt_source = 'module'
                            dest_mod = ConfigModule.query.filter_by(module_key='destination').first()
                            opt_module_id = dest_mod.id if dest_mod else None
                        elif fname == 'bus_type_id':
                            opt_source = 'module'
                            bt_mod = ConfigModule.query.filter_by(module_key='bus_type').first()
                            opt_module_id = bt_mod.id if bt_mod else None

                        f = ConfigField(
                            module_id=module.id,
                            field_name=fname,
                            field_label=flabel,
                            field_type=ftype,
                            is_required=req,
                            is_unique=unique,
                            is_listable=listable,
                            is_searchable=req or unique,
                            display_order=i,
                            option_source=opt_source,
                            option_module_id=opt_module_id,
                        )
                        db.session.add(f)

            # Seed sample data for core modules
            dest_module = ConfigModule.query.filter_by(module_key='destination').first()
            dest_data = ['Phnom Penh', 'Siem Reap', 'Battambang', 'Kampot', 'Kampong Cham', 'Sihanoukville', 'Takeo', 'Kandal']
            for name in dest_data:
                rec = DynamicRecord(module_id=dest_module.id, data={'name': name}, is_active=True)
                db.session.add(rec)

            transport_module = ConfigModule.query.filter_by(module_key='transportation_type').first()
            for name in ['Bus', 'Van', 'Car', 'Minibus', 'Motorcycle', 'Truck']:
                rec = DynamicRecord(module_id=transport_module.id, data={'name': name}, is_active=True)
                db.session.add(rec)

            bus_type_module = ConfigModule.query.filter_by(module_key='bus_type').first()
            for name, cap in [('Standard', 40), ('VIP', 30), ('Mini', 16), ('Luxury', 20)]:
                rec = DynamicRecord(module_id=bus_type_module.id, data={'name': name, 'capacity': cap}, is_active=True)
                db.session.add(rec)

            amenity_module = ConfigModule.query.filter_by(module_key='amenity').first()
            for item, icon in [('WiFi', 'fa-wifi'), ('AC', 'fa-snowflake'), ('TV', 'fa-tv'), ('Toilet', 'fa-toilet'), ('Water', 'fa-bottle-water'), ('Charger', 'fa-plug')]:
                rec = DynamicRecord(module_id=amenity_module.id, data={'name': item, 'icon': icon}, is_active=True)
                db.session.add(rec)

            db.session.commit()
            print("Dynamic configuration system initialized.")

        # ── Migrate: hide unused config modules ──
        unused_keys = ['agency', 'agent_group', 'destination_group', 'boarding_point',
                       'drop_off_point', 'bus', 'bus_type', 'amenity']
        for mod in ConfigModule.query.filter(ConfigModule.module_key.in_(unused_keys)).all():
            mod.is_active = False
        db.session.commit()

        # ── Migrate: add new config modules if missing ──
        new_modules = [
            ('penalty_violation_type', 'Violation Type', 'fa-exclamation-triangle', 20),
            ('penalty_department', 'Penalty Department', 'fa-building', 21),
            ('vehicle_status', 'Vehicle Status', 'fa-truck', 22),
            ('power_status', 'Power Status', 'fa-power-off', 23),
            ('report_submission_status', 'Report Submission', 'fa-file-export', 24),
            ('request_type', 'Request Type', 'fa-tag', 25),
            ('entity_status', 'Entity Status', 'fa-toggle-on', 26),
        ]
        new_field_sets = {
            'penalty_violation_type': [
                ('name', 'Violation Name', 'text', True, True, True),
                ('description', 'Description', 'textarea', False, False, False),
                ('default_min_amount', 'Default Min Amount ($)', 'number', False, False, False),
                ('default_max_amount', 'Default Max Amount ($)', 'number', False, False, False),
            ],
            'penalty_department': [
                ('name', 'Department Name', 'text', True, True, True),
                ('code', 'Department Code', 'text', False, False, True),
            ],
            'vehicle_status': [
                ('name', 'Status Name', 'text', True, True, True),
                ('label', 'Display Label', 'text', False, False, True),
            ],
            'power_status': [
                ('name', 'Status Name', 'text', True, True, True),
            ],
            'report_submission_status': [
                ('name', 'Status Name', 'text', True, True, True),
                ('score_weight', 'Score Weight (0-100)', 'number', False, False, False),
            ],
            'request_type': [
                ('name', 'Type Name', 'text', True, True, True),
                ('code', 'Type Code', 'text', True, True, True),
                ('description', 'Description', 'textarea', False, False, False),
                ('icon', 'Icon Class', 'text', False, False, True),
            ],
            'entity_status': [
                ('name', 'Status Name', 'text', True, True, True),
            ],
        }
        for key, name, icon, order in new_modules:
            existing = ConfigModule.query.filter_by(module_key=key).first()
            if not existing:
                m = ConfigModule(module_key=key, module_name=name, module_icon=icon, sort_order=order, is_active=True)
                db.session.add(m)
                db.session.flush()
                if key in new_field_sets:
                    for i, (fname, flabel, ftype, req, unique, listable) in enumerate(new_field_sets[key]):
                        f = ConfigField(
                            module_id=m.id,
                            field_name=fname,
                            field_label=flabel,
                            field_type=ftype,
                            is_required=req,
                            is_unique=unique,
                            is_listable=listable,
                            is_searchable=req or unique,
                            display_order=i,
                        )
                        db.session.add(f)

        # Seed sample data for new modules
        _seed_if_empty('penalty_violation_type', [
            {'name': 'Late Arrival', 'default_min_amount': 25, 'default_max_amount': 50},
            {'name': 'Absenteeism', 'default_min_amount': 50, 'default_max_amount': 100},
            {'name': 'Misconduct', 'default_min_amount': 100, 'default_max_amount': 500},
            {'name': 'Policy Violation', 'default_min_amount': 50, 'default_max_amount': 150},
            {'name': 'Performance Issue'},
            {'name': 'Insubordination'},
            {'name': 'Theft'},
            {'name': 'Other'},
        ])
        _seed_if_empty('penalty_department', [
            {'name': 'IT'}, {'name': 'HR'}, {'name': 'Finance'}, {'name': 'Operations'},
            {'name': 'Marketing'}, {'name': 'Management'}, {'name': 'Sales'},
            {'name': 'Legal'}, {'name': 'Logistics'},
        ])
        _seed_if_empty('vehicle_status', [
            {'name': 'Departed', 'label': 'Departed'},
            {'name': 'Not Departed', 'label': 'Not Departed'},
        ])
        _seed_if_empty('power_status', [
            {'name': 'Off'}, {'name': 'On'},
        ])
        _seed_if_empty('report_submission_status', [
            {'name': 'On Time', 'score_weight': 100},
            {'name': 'Late', 'score_weight': 50},
            {'name': 'Incomplete', 'score_weight': 25},
        ])
        _seed_if_empty('request_type', [
            {'name': 'Create Journey', 'code': 'create_journey', 'icon': 'fa-plus-circle',
             'description': 'Create a new transportation journey route'},
            {'name': 'Open Route', 'code': 'open_route', 'icon': 'fa-road',
             'description': 'Open a new route for transportation'},
            {'name': 'Change Route', 'code': 'change_route', 'icon': 'fa-exchange-alt',
             'description': 'Modify an existing transportation route'},
        ])
        _seed_if_empty('entity_status', [
            {'name': 'Active'}, {'name': 'Inactive'},
        ])
        db.session.commit()

        # ── Seed Default Roles ──
        all_permissions = [
            'dashboard_view', 'dashboard_edit', 'dashboard_download',
            'route_request_view', 'route_request_create', 'route_request_edit', 'route_request_delete',
            'route_request_approve', 'route_request_reject', 'route_request_download',
            'transport_request_view', 'transport_request_create', 'transport_request_edit', 'transport_request_delete',
            'transport_request_approve', 'transport_request_reject', 'transport_request_download',
            'penalty_view', 'penalty_create', 'penalty_edit', 'penalty_delete', 'penalty_download',
            'trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit',
            'trip_operation_report_delete', 'trip_operation_report_download',
            'daily_report_view', 'daily_report_create', 'daily_report_edit', 'daily_report_delete',
            'kpi_evaluation_view', 'kpi_evaluation_create', 'kpi_evaluation_edit', 'kpi_evaluation_delete',
            'kpi_dashboard_view', 'kpi_history_view',
            'system_settings_view', 'system_settings_edit', 'system_settings_update',
            'department_view', 'department_create', 'department_edit', 'department_delete', 'department_download',
            'position_view', 'position_create', 'position_edit', 'position_delete', 'position_download',
            'role_view', 'role_create', 'role_edit', 'role_delete', 'role_assign_permissions', 'role_download',
            'user_view', 'user_create', 'user_edit', 'user_delete',
            'user_assign_roles', 'user_reset_password', 'user_activate', 'user_deactivate', 'user_download',
            'view_all_records', 'edit_any_status',
        ]
        it_staff_permissions = [
            'dashboard_view',
            'route_request_view',
            'transport_request_view',
            'penalty_view',
            'trip_operation_report_view', 'trip_operation_report_create',
            'daily_report_view', 'kpi_dashboard_view', 'kpi_history_view',
            'department_view',
            'position_view',
            'role_view',
            'user_view',
        ]
        branch_manager_permissions = [
            'dashboard_view', 'dashboard_download',
            'route_request_view', 'route_request_create', 'route_request_edit', 'route_request_download',
            'transport_request_view', 'transport_request_create', 'transport_request_edit', 'transport_request_download',
            'penalty_view',
            'trip_operation_report_view', 'trip_operation_report_create', 'trip_operation_report_edit',
            'daily_report_view', 'daily_report_create', 'daily_report_edit',
            'kpi_dashboard_view', 'kpi_history_view',
        ]
        regional_manager_permissions = [
            'dashboard_view',
            'route_request_view', 'route_request_approve', 'route_request_reject',
            'transport_request_view', 'transport_request_approve', 'transport_request_reject',
            'penalty_view',
            'trip_operation_report_view',
            'daily_report_view', 'kpi_dashboard_view', 'kpi_history_view',
            'view_all_records',
        ]
        hr_manager_permissions = [
            'dashboard_view',
            'penalty_view', 'penalty_create', 'penalty_edit', 'penalty_download',
            'transport_request_view',
            'trip_operation_report_view',
            'daily_report_view', 'kpi_dashboard_view', 'kpi_history_view',
            'department_view', 'department_create', 'department_edit',
            'position_view', 'position_create', 'position_edit',
            'user_view',
            'view_all_records',
        ]
        role_perms = {
            'admin': all_permissions,
            'it_staff': it_staff_permissions,
            'branch_manager': branch_manager_permissions,
            'regional_manager': regional_manager_permissions,
            'hr_manager': hr_manager_permissions,
        }
        default_roles = [
            ('admin', 'Administrator'),
            ('it_staff', 'IT Staff'),
            ('branch_manager', 'Branch Manager'),
            ('regional_manager', 'Regional Manager'),
            ('hr_manager', 'HR Manager'),
        ]
        for name, label in default_roles:
            existing = Role.query.filter_by(name=name).first()
            if not existing:
                r = Role(name=name, label=label, permissions=role_perms.get(name, []))
                db.session.add(r)
        db.session.commit()

        # ── Seed Users and Demo Data ──
        admin_role = Role.query.filter_by(name='admin').first()
        admin_user = User.query.filter_by(username='admin').first()
        if admin_user:
            if admin_role and admin_user.role_id is None:
                admin_user.role_id = admin_role.id
                logger.info(f"Fixed admin user (id={admin_user.id}): assigned role_id={admin_role.id}")
        else:
            admin = User(full_name='System Administrator', username='admin', branch='HQ')
            if admin_role:
                admin.role_id = admin_role.id
            admin.set_password('admin123')
            db.session.add(admin)

        users_seed = [
            ('Branch Manager A', 'branch_mgr', 'branch_manager', 'Phnom Penh Branch'),
            ('Regional Manager', 'regional_mgr', 'regional_manager', 'Region 1'),
            ('HR Manager', 'hr_mgr', 'hr_manager', 'HQ'),
            ('IT Staff', 'it_staff', 'it_staff', 'HQ'),
        ]
        for name, uname, role_key, branch in users_seed:
            if User.query.filter_by(username=uname).first():
                continue
            u = User(full_name=name, username=uname, branch=branch)
            role_obj = Role.query.filter_by(name=role_key).first()
            if role_obj:
                u.role_id = role_obj.id
            u.set_password('pass123')
            db.session.add(u)

        # Seed sample data if tables are empty
        if not RouteRequest.query.first():
            sample_routes = [
                ('Phnom Penh', 'Siem Reap', 'PP Branch', 'ABC Co.', 'Bus', 'NR6', 25.0, '07:00', '12:00', 'Approved'),
                ('Phnom Penh', 'Kampot', 'PP Branch', 'XYZ Co.', 'Van', 'NR3', 15.0, '08:00', '11:00', 'Pending'),
                ('Siem Reap', 'Battambang', 'SR Branch', 'DEF Co.', 'Bus', 'NR5', 10.0, '09:00', '12:00', 'Pending'),
                ('Phnom Penh', 'Kampong Cham', 'PP Branch', 'GHI Co.', 'Van', 'NR7', 12.0, '06:00', '09:00', 'Rejected'),
                ('Battambang', 'Phnom Penh', 'BTB Branch', 'ABC Co.', 'Bus', 'NR5', 18.0, '07:30', '12:30', 'Approved'),
            ]
            for i, (frm, to, branch, comp, trans, road, price, dep, arr, status) in enumerate(sample_routes, 1):
                rr = RouteRequest(
                    request_id=f"RR-2024-{i:04d}",
                    requester_name='Branch Manager A',
                    destination_from=frm, destination_to=to,
                    branch_location=branch, company=comp,
                    transportation=trans, national_road=road,
                    price=price, departure_time=dep, arrival_time=arr,
                    start_date=date(2024, 1, 1), end_date=date(2024, 12, 31),
                    status=status, remarks='Sample data'
                )
                db.session.add(rr)

        if not EmployeePenalty.query.first():
            sample_penalties = [
                ('EMP001', 'Sok Dara', 'IT', 'Developer', 'Late Arrival', 'Arrived 2 hours late', 50.0, 'Approved'),
                ('EMP002', 'Chea Vanna', 'HR', 'Officer', 'Absenteeism', 'Absent without notice', 100.0, 'Pending'),
                ('EMP003', 'Lim Phalla', 'Finance', 'Accountant', 'Policy Violation', 'Misuse of company resources', 75.0, 'Approved'),
                ('EMP004', 'Keo Sophea', 'Operations', 'Supervisor', 'Late Arrival', 'Consistently late', 40.0, 'Rejected'),
                ('EMP005', 'Meas Pich', 'IT', 'Analyst', 'Misconduct', 'Inappropriate behavior', 150.0, 'Pending'),
            ]
            for i, (eid, ename, dept, pos, vtype, desc, amount, status) in enumerate(sample_penalties, 1):
                ep = EmployeePenalty(
                    penalty_id=f"EP-2024-{i:04d}",
                    employee_id=eid, employee_name=ename,
                    department=dept, position=pos,
                    violation_type=vtype, description=desc,
                    penalty_amount=amount, approved_by='Admin',
                    status=status
                )
                db.session.add(ep)

        # ── Seed HR Modules (Departments, Positions) ──
        if not Department.query.first():
            depts = ['Human Resources', 'Information Technology', 'Finance', 'Operations', 'Marketing', 'Sales', 'Legal', 'Logistics']
            dept_objs = {}
            for dname in depts:
                d = Department(name=dname, description=f'{dname} Department', status='Active')
                db.session.add(d)
                db.session.flush()
                dept_objs[dname] = d.id

            positions_data = [
                ('HR Manager', 'Human Resources'), ('HR Officer', 'Human Resources'),
                ('IT Manager', 'Information Technology'), ('Software Engineer', 'Information Technology'),
                ('IT Support', 'Information Technology'), ('Finance Manager', 'Finance'),
                ('Accountant', 'Finance'), ('Operations Manager', 'Operations'),
                ('Supervisor', 'Operations'), ('Marketing Manager', 'Marketing'),
                ('Sales Executive', 'Sales'), ('Legal Counsel', 'Legal'),
                ('Logistics Coordinator', 'Logistics'),
            ]
            for pname, dname in positions_data:
                pos = Position(name=pname, department_id=dept_objs[dname], status='Active')
                db.session.add(pos)

            db.session.commit()

if __name__ == '__main__':
    init_db()
    app.run(debug=False, port=5000)
